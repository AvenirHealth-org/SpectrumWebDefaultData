import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_MST_ID                = '<Master ID>'  
IC_CONSTANT              = '<Constant>'
IC_NAME_ENGLISH          = '<Name in English>'
IC_DRUG_SUPPLY_STR_CONST = '<String Constant>'
IC_TYPE                  = '<Type>'
IC_UNIT_COST             = '<Unit Cost>'
IC_FAMILY_LG             = '<Logistics - Family>'
IC_WEIGHT_LG             = '<Logistics - Weight (kg)>'
IC_WEIGHT_UOHM_LG        = '<Logistics - UoM_wt>'
IC_VOLUME_LG             = '<Logistics - Cubic Ft (m)>'
IC_VOLUME_UOM_LG         = '<Logistics - UoM_cubic>'
IC_FORECAST_ERROR_LG     = '<Logistics - ForecastError>'
IC_IMPORTANCE_LG         = '<Logistics - Importance>'
IC_NON_CRIT_FRACTION_LG  = '<Logistics - NonCriticalFraction>'
IC_STD_FRACTION_LG       = '<Logistics - StandardFraction>'
IC_CRIT_FRACTION_LG      = '<Logistics - CriticalFraction>'
IC_COST_PER_PACK_LG      = '<Logistics - Cost per Pack>'
IC_UNITS_PER_PACK_LG     = '<Logistics - Units per Pack>'
IC_COST_TYPE_ID_LG       = '<Logistics - CostTypeID>'
IC_SOURCE                = '<Source>'
IC_PROG_AREA_FILTERS     = '<Program Area Filters>'

def create_drug_supply_DB_IC(version = str):

    log('Creating IC drug supply DB')

    drug_supply_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_DRUG_SUPPLY_DB_NAME]
    
    '''
        IMPORTANT: Max drug/supply master ID: 1208

        New drugs and supplies must have a master ID greater than this. Please update this maximum if you add a new one.

        Drugs and supplies are in alphabetical order for easy searching.

        Type: Drug = 1, Supply = 2
    '''

    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    mstID_col = gbc.GB_NOT_FOUND
    constant_col = gbc.GB_NOT_FOUND
    str_const_col = gbc.GB_NOT_FOUND
    english_col = gbc.GB_NOT_FOUND
    type_col = gbc.GB_NOT_FOUND
    unit_cost_col = gbc.GB_NOT_FOUND
    source_col = gbc.GB_NOT_FOUND
    prog_area_filters_col = gbc.GB_NOT_FOUND
    cost_type_ID_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == IC_MST_ID:
            mstID_col = c

        elif tag == IC_CONSTANT:
            constant_col = c

        elif tag == IC_DRUG_SUPPLY_STR_CONST:
            str_const_col = c

        elif tag == IC_NAME_ENGLISH:
            english_col = c

        elif tag == IC_TYPE:
            type_col = c

        elif tag == IC_UNIT_COST:
            unit_cost_col = c

        elif tag == IC_SOURCE:
            source_col = c

        elif tag == IC_PROG_AREA_FILTERS:
            prog_area_filters_col = c

        elif tag == IC_COST_TYPE_ID_LG:
            cost_type_ID_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        drug_supply_dict = {}

        drug_supply_dict[icdbc.IC_MST_ID_KEY_DSDB] = row[mstID_col - 1]
        drug_supply_dict[icdbc.IC_CONSTANT_KEY_DSDB] = row[constant_col - 1]
        drug_supply_dict[icdbc.IC_STR_CONST_KEY_DSDB] = row[str_const_col - 1]
        drug_supply_dict[icdbc.IC_ENGLISH_STRING_KEY_DSDB] = row[english_col - 1]
        drug_supply_dict[icdbc.IC_TYPE_MST_ID_KEY_DSCB] = row[type_col - 1]
        drug_supply_dict[icdbc.IC_UNIT_COST_KEY_DSCB] = row[unit_cost_col - 1]
        drug_supply_dict[icdbc.IC_SOURCE_KEY_DSCB] = row[source_col - 1]
        if row[prog_area_filters_col - 1] != None:
            drug_supply_dict[icdbc.IC_PROG_AREA_FILTER_MST_IDS_KEY_DSCB] = str(row[prog_area_filters_col - 1]).split(', ')
        else:
            drug_supply_dict[icdbc.IC_PROG_AREA_FILTER_MST_IDS_KEY_DSCB] = []
        drug_supply_dict[icdbc.IC_COST_TYPE_MST_ID_KEY_LG_DSCB] = row[cost_type_ID_col - 1]
        drug_supply_list.append(drug_supply_dict)

    j = ujson.dumps(drug_supply_list)
    # for debugging
    # for i, obj in enumerate(drug_supply_list):
    #     log(obj)
    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC)
    JSON_file_name = icc.IC_DRUG_SUPPLY_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    with open(IC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished IC drug supply DB')

def upload_drug_supply_DB_IC(version):
    JSON_file_name = icc.IC_DRUG_SUPPLY_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_IC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + JSON_file_name) 
    log('uploaded IC drug supply DB')
            
