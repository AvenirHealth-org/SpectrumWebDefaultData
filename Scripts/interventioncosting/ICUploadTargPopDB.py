import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_CONSTANT        = '<Constant>'
IC_MST_ID          = '<Master ID>'  
IC_NAME            = '<Name>'
IC_STR_CONST       = '<String Constant>'
IC_IHT_PROG_AREAS  = '<IHT - Programme Area(s)>'
IC_TB_PROG_AREAS   = '<TB - Programme Area(s)>'
IC_LIST_PROG_AREAS = '<LiST - Programme Area(s)>'
IC_REQ_MODULES     = '<Required Module(s)>'

def create_targ_pop_DB_IC(version = str):

    log('Creating IC target pop DB')

    targ_pop_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_TARG_POP_DB_NAME]
    
    '''
        IMPORTANT: Max target population master ID:338

        New target populations must have a master ID greater than this. Please update this maximum if you add a new one.

    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    constant_col = gbc.GB_NOT_FOUND
    mstID_col = gbc.GB_NOT_FOUND
    name_col = gbc.GB_NOT_FOUND
    str_const_col = gbc.GB_NOT_FOUND
    IHT_prog_areas_col = gbc.GB_NOT_FOUND
    TB_prog_areas_col = gbc.GB_NOT_FOUND
    LiST_prog_areas_col = gbc.GB_NOT_FOUND
    req_modules_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == IC_CONSTANT:
            constant_col = c

        elif tag == IC_MST_ID:
            mstID_col = c

        elif tag == IC_NAME:
            name_col = c

        elif tag == IC_STR_CONST:
            str_const_col = c

        elif tag == IC_IHT_PROG_AREAS:
            IHT_prog_areas_col = c

        elif tag == IC_TB_PROG_AREAS:
            TB_prog_areas_col = c

        elif tag == IC_LIST_PROG_AREAS:
            LiST_prog_areas_col = c

        elif tag == IC_REQ_MODULES:
            req_modules_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        targ_pop_dict = {}

        targ_pop_dict[icdbc.IC_CONSTANT_KEY_TPDB] = row[constant_col - 1]
        targ_pop_dict[icdbc.IC_MST_ID_KEY_TPDB] = row[mstID_col - 1]
        targ_pop_dict[icdbc.IC_NAME_KEY_TPDB] = row[name_col - 1]
        targ_pop_dict[icdbc.IC_STR_CONST_KEY_TPDB] = row[str_const_col - 1]

        if row[IHT_prog_areas_col - 1] != None:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_IHT_KEY_TPDB] = str(row[IHT_prog_areas_col - 1]).split(', ')
        else:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_IHT_KEY_TPDB] = []

        if row[TB_prog_areas_col - 1] != None:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_TB_KEY_TPDB] = str(row[TB_prog_areas_col - 1]).split(', ')
        else:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_TB_KEY_TPDB] = []

        if row[LiST_prog_areas_col - 1] != None:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_LIST_KEY_TPDB] = str(row[LiST_prog_areas_col - 1]).split(', ')
        else:
            targ_pop_dict[icdbc.IC_APPLIC_PROG_AREA_IDS_LIST_KEY_TPDB] = []

        # handle special 'or' cases in column H later.
        if row[req_modules_col - 1] != None:
            targ_pop_dict[icdbc.IC_REQ_MOD_MST_IDS_KEY_TPDB] = str(row[req_modules_col - 1]).split(', ')
        else:
            targ_pop_dict[icdbc.IC_REQ_MOD_MST_IDS_KEY_TPDB] = []

        targ_pop_list.append(targ_pop_dict)

    j = ujson.dumps(targ_pop_list)
    # for debugging
    # for i, obj in enumerate(targ_pop_list):
    #     log(obj)
    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC)
    JSON_file_name = icc.IC_TARG_POP_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    with open(IC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished IC target pop DB')

    with open(os.getcwd() + '\SpectrumCommon\Const\IC\ICTargetPopulationIDs.py', 'w') as f:
        for targ_pop_dict in targ_pop_list:
            f.write(targ_pop_dict[icdbc.IC_CONSTANT_KEY_TPDB] + ' = ' + str(targ_pop_dict[icdbc.IC_MST_ID_KEY_TPDB]) + '\n')

    log('Updated target population master IDs in SpectrumCommon --> Const --> IC --> ICTargetPopulationIDs.py.')

def upload_targ_pop_DB_IC(version):
    JSON_file_name = icc.IC_TARG_POP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_IC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + JSON_file_name) 
    log('Uploaded IC target pop DB')
            
