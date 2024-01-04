import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

def create_treatment_input_and_group_DBs_IC(version_str):

    log('Creating IC treatment input and group DB')
    IC_path = os.getcwd() + '\Tools\DefaultDataManager\IC\\'
    IC_mod_data_FQ_name = IC_path + 'ModData\ICModData.xlsx'

    treatment_input_list = []
    treatment_input_group_list = []

    wb = load_workbook(IC_mod_data_FQ_name)
    sheet = wb[icc.IC_TREATMENT_INPUT_DB_NAME]

    '''
        <Notes>
        
            New groups within a particular intervention, channel, and item type must be given a master ID that is greater than any existing 
            group master ID within those confines.
  
            Special case types:
                1 = Standard (no special case); 
                2 = Detailed vaccine mode on in LiST and LiST Costing used; 
                3 = Detailed vaccine mode off in LiST and LiST Costing used

    '''

    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    interv_mst_ID_col = gbc.GB_NOT_FOUND
    #interv_name_col = gbc.GB_NOT_FOUND
    #special_case_type_mst_ID_col = gbc.GB_NOT_FOUND
    deliv_chan_mst_IDs_col = gbc.GB_NOT_FOUND
    treat_input_type_mst_ID_col = gbc.GB_NOT_FOUND
    group_mst_ID_col = gbc.GB_NOT_FOUND
    group_name_col = gbc.GB_NOT_FOUND
    group_str_const_col = gbc.GB_NOT_FOUND
    input_mst_ID_col = gbc.GB_NOT_FOUND
    #input_name_col = gbc.GB_NOT_FOUND
    perc_rec_treated_by_col = gbc.GB_NOT_FOUND
    note_col = gbc.GB_NOT_FOUND
    note_str_const_col = gbc.GB_NOT_FOUND
    num_items_col = gbc.GB_NOT_FOUND
    times_per_day_min_col = gbc.GB_NOT_FOUND
    days_per_case_num_days_col = gbc.GB_NOT_FOUND
    units_per_case_person_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == icdbc.IC_INTERV_MST_ID_TAG_TIDB:
            interv_mst_ID_col = c

        elif tag == icdbc.IC_DELIV_CHAN_MST_IDS_TAG_TIDB:
            deliv_chan_mst_IDs_col = c

        elif tag == icdbc.IC_TREAT_INPUT_TYPE_MST_ID_TAG_TIDB:
            treat_input_type_mst_ID_col = c

        elif tag == icdbc.IC_GROUP_MST_ID_TAG_TIDB:
            group_mst_ID_col = c

        elif tag == icdbc.IC_GROUP_NAME_TAG_TIDB:
            group_name_col = c

        elif tag == icdbc.IC_GROUP_STR_CONST_TAG_TIDB:
            group_str_const_col = c

        elif tag == icdbc.IC_INPUT_MST_ID_TAG_TIDB:
            input_mst_ID_col = c

        #elif tag == IC_INPUT_NAME:
        #    input_name_col = c

        elif tag == icdbc.IC_PERC_REC_TREATED_BY_TAG_TIDB:
            perc_rec_treated_by_col = c

        elif tag == icdbc.IC_NOTE_TAG_TIDB:
            note_col = c

        elif tag == icdbc.IC_NOTE_STR_CONST_TAG_TIDB:
            note_str_const_col = c

        elif tag == icdbc.IC_NUM_ITEMS_TAG_TIDB:
            num_items_col = c

        elif tag == icdbc.IC_TIMES_PER_DAY_MIN_TAG_TIDB:
            times_per_day_min_col = c

        elif tag == icdbc.IC_DAYS_PER_CASE_NUM_DAYS_TAG_TIDB:
            days_per_case_num_days_col = c
        
        elif tag == icdbc.IC_UNITS_PER_CASE_PERSON_TAG_TIDB:
            units_per_case_person_col = c

    current_interv_mst_ID = 0
    current_deliv_chan_mst_ID_list = []
    current_treat_input_type_mst_ID = 0
    currrent_group_mst_ID = 0

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        # If the dict with the current intervention master ID, delivery channel master ID, item type master ID, and group master ID 
        # does not exist, create a new group master ID. Otherwise, leave the treatment input groups alone.  and add to it. Otherwise, add to the one from the previous loop.

        interv_mst_ID = row[interv_mst_ID_col - 1]
        deliv_chan_mst_IDs = row[deliv_chan_mst_IDs_col - 1]
        treat_input_type_mst_ID = row[treat_input_type_mst_ID_col - 1]

        group_mst_ID = row[group_mst_ID_col - 1]
        group_name = row[group_name_col - 1]
        group_str_const = row[group_str_const_col - 1]
        
        input_mst_ID = row[input_mst_ID_col - 1]
        #input_name = row[input_name_col - 1]
        perc_rec_treated_by = row[perc_rec_treated_by_col - 1]
        note = row[note_col - 1]
        note_str_const = row[note_str_const_col - 1]
        num_items = row[num_items_col - 1]
        times_per_day_min = row[times_per_day_min_col - 1]
        days_per_case_num_days = row[days_per_case_num_days_col - 1]
        units_per_case_person = row[units_per_case_person_col - 1]

        # Create a treatment input object and treatment input group object for each delivery channel
        deliv_chan_list = str(deliv_chan_mst_IDs).split(', ')

        for deliv_chan_mst_ID in deliv_chan_list:
            treatment_input_dict = {}
            treatment_input_dict[icdbc.IC_INTERV_MST_ID_KEY_TIDB] = str(interv_mst_ID)
            treatment_input_dict[icdbc.IC_DELIV_CHAN_MST_ID_KEY_TIDB] = str(deliv_chan_mst_ID)
            treatment_input_dict[icdbc.IC_TREAT_INPUT_TYPE_MST_ID_KEY_TIDB] = treat_input_type_mst_ID
            treatment_input_dict[icdbc.IC_GROUP_MST_ID_KEY_TIDB] = str(group_mst_ID)
            treatment_input_dict[icdbc.IC_TREATMENT_INPUT_MST_ID_KEY_TIDB] = input_mst_ID if treat_input_type_mst_ID != icc.IC_TI_DRUGS_SUPPLIES_MST_ID else str(input_mst_ID)
            treatment_input_dict[icdbc.IC_PERC_APPLIED_KEY_TIDB] = perc_rec_treated_by
            treatment_input_dict[icdbc.IC_NOTE_KEY_TIDB] = note if note != None else ''
            treatment_input_dict[icdbc.IC_NOTE_STR_CONST_KEY_TIDB] = note_str_const
            treatment_input_dict[icdbc.IC_NUM_ITEMS_KEY_TIDB] = num_items
            treatment_input_dict[icdbc.IC_TIMES_PER_DAY_MIN_KEY_TIDB] = times_per_day_min
            treatment_input_dict[icdbc.IC_DAYS_PER_CASE_NUM_DAYS_KEY_TIDB] = days_per_case_num_days
            treatment_input_dict[icdbc.IC_UNITS_PER_CASE_PERSON_KEY_TIDB] = units_per_case_person
            treatment_input_list.append(treatment_input_dict)

            # Create a new group dict if the intervention, delivery channel, item type, or group IDs have changed.
            if not ((current_interv_mst_ID == interv_mst_ID) and 
                (deliv_chan_mst_ID in current_deliv_chan_mst_ID_list) and 
                (current_treat_input_type_mst_ID == treat_input_type_mst_ID) and 
                (currrent_group_mst_ID == group_mst_ID)):
                    treatment_input_group_dict = {}
                    treatment_input_group_dict[icdbc.IC_INTERV_MST_ID_KEY_TIGDB] = str(interv_mst_ID)
                    treatment_input_group_dict[icdbc.IC_DELIV_CHAN_MST_ID_KEY_TIGDB] = str(deliv_chan_mst_ID)
                    treatment_input_group_dict[icdbc.IC_TREAT_INPUT_TYPE_MST_ID_KEY_TIGDB] = treat_input_type_mst_ID
                    treatment_input_group_dict[icdbc.IC_GROUP_MST_ID_KEY_TIGDB] = str(group_mst_ID)
                    treatment_input_group_dict[icdbc.IC_GROUP_NAME_KEY_TIGDB] = group_name if group_name != None else ''
                    treatment_input_group_dict[icdbc.IC_GROUP_STR_CONST_KEY_TIGDB] = group_str_const
                    treatment_input_group_list.append(treatment_input_group_dict)
            
            # Clear the the delivery channels we're tracking if the intervention, treatment input type, or group changes
            if (current_interv_mst_ID != interv_mst_ID) or (current_treat_input_type_mst_ID != treat_input_type_mst_ID) or (currrent_group_mst_ID != group_mst_ID): 
                current_deliv_chan_mst_ID_list = []
            
            current_interv_mst_ID = interv_mst_ID
            if not (deliv_chan_mst_ID in current_deliv_chan_mst_ID_list):
                current_deliv_chan_mst_ID_list.append(deliv_chan_mst_ID)
            current_treat_input_type_mst_ID = treat_input_type_mst_ID
            currrent_group_mst_ID = group_mst_ID

    # Write treatment inputs

    j = ujson.dumps(treatment_input_list)
    # for debugging
    # for i, obj in enumerate(drug_supply_list):
    #     log(obj)

    with open(IC_path + 'JSON\\' + icc.IC_TREATMENT_INPUT_DB_NAME + '_' + version_str + '.' + gbc.GB_JSON, 'w') as f:
        f.write(j)

    # Write treatment input groups

    j = ujson.dumps(treatment_input_group_list)

    #fails after this point
    with open(IC_path + 'JSON\\' + icc.IC_TREATMENT_INPUT_GROUP_DB_NAME + '_' + version_str + '.' + gbc.GB_JSON, 'w') as f:
        f.write(j)
        
    log('Finished IC treatment input and group DB')

def upload_treatment_input_and_group_DBs_IC(version):
    connection =  os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV]
    container_name = gbc.GB_IC_CONTAINER

    FQName = os.getcwd() + '\Tools\DefaultDataManager\IC\JSON\\' + icc.IC_TREATMENT_INPUT_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(connection, container_name, icc.IC_TREATMENT_INPUT_DB_NAME + '_' + version + '.' + gbc.GB_JSON, FQName) 

    FQName = os.getcwd() + '\Tools\DefaultDataManager\IC\JSON\\' + icc.IC_TREATMENT_INPUT_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(connection, container_name, icc.IC_TREATMENT_INPUT_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON, FQName)
    
    log('Uploaded IC treatment input and group DB') 

    
            
