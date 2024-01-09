import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
IC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
IC_COUNTRY_NAME             = '<Country Name>'  
IC_SPAN_SINGLE              = '<Single>'
IC_SPAN_START               = '<Start>'
IC_SPAN_END                 = '<End>'
IC_SPAN_SKIP                = '<Skip>'
IC_INTERV_MST_ID            = '<Master Intervention ID>'  
IC_INTERV_NAME              = '<Intervention Name>'  
IC_AGE_GROUP                = '<Age Group (if applicable)>'  
IC_PIN_TYPE_STRUCTURE       = '<PIN Type Structure (if applicable)>'  

def create_PIN_DB_IC(version = str):

    log('Creating IC PIN DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_PIN_DB_NAME]
    
    num_rows = sheet.max_row
    #num_cols = sheet.max_column

    # Span row has <Start> and <End> tags to tell us when PINs by age group start and end.
    span_row = 1

    # Establish tag columns.  Not many, so don't bother searching for them.
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    first_data_col = col

    # Establish tag rows. Not many, so don't bother searching for them.
    row = 2
    interv_mstID_row = row
    row += 1
    #interv_name_row = row
    row += 1
    #age_group_row = row
    row += 1
    #PIN_type_structure_row = row
    row += 1
    first_data_row = row

    # Span row list
    span_row_list = []
    for row in islice(sheet.values, span_row - 1, span_row):
        span_row_list = row

    # Intervention master ID row
    interv_mstID_row_list = []
    for row in islice(sheet.values, interv_mstID_row - 1, interv_mstID_row):
        interv_mstID_row_list = row

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}
        intervention_list = []

        iso3 = row[iso_alpha_3_country_code_col - 1]

        if iso3 != None:
            country_dict[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PINDB] = iso3

            for col, col_val in enumerate(row):
                if col >= (first_data_col - 1):

                    # Single values
                    if span_row_list[col] == IC_SPAN_SINGLE:
                        intervention_dict = {}
                        intervention_dict[icdbc.IC_INTERV_MST_ID_KEY_PINDB] = interv_mstID_row_list[col]
                        intervention_dict[icdbc.IC_PIN_KEY_PINDB] = col_val
                        intervention_list.append(intervention_dict)
                        #log(intervention_dict)

                    # Values by age group
                    elif span_row_list[col] != IC_SPAN_SKIP:
                        # First column
                        if span_row_list[col] == IC_SPAN_START:
                            intervention_dict = {}
                            intervention_dict[icdbc.IC_INTERV_MST_ID_KEY_PINDB] = interv_mstID_row_list[col]
                            intervention_dict[icdbc.IC_PIN_KEY_PINDB] = []
                            intervention_dict[icdbc.IC_PIN_KEY_PINDB].append(col_val)
                        
                        # Last column
                        elif span_row_list[col] == IC_SPAN_END:
                            intervention_dict[icdbc.IC_PIN_KEY_PINDB].append(col_val)
                            intervention_list.append(intervention_dict)
                            #log(intervention_dict)

                        else:
                            intervention_dict[icdbc.IC_PIN_KEY_PINDB].append(col_val)

            country_dict[icdbc.IC_INTERVENTION_LIST_KEY_PINDB] = intervention_list
            country_list.append(country_dict)

    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_PIN_DB_DIR
    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PINDB]
        with open(IC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished IC PIN DB')

def upload_PIN_DB_IC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_PIN_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_IC_CONTAINER, walk_path, version)
    log('Uploaded IC PIN DB')
            
