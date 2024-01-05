import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
IC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
IC_COUNTRY_NAME             = '<Country Name>'
IC_INTERV_MST_ID            = '<Intervention Master ID>'
IC_INTERV_NAME              = '<Intervention Name>'
IC_DATA                     = '<Data>'
IC_YEARS                    = '<Years>'
IC_SPAN                     = '<Span>'
IC_START                    = '<Start>'
IC_END                      = '<End>'

def create_PIN_by_year_DB_IC(version = str):

    log('Creating IC PIN by year DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_PIN_BY_YEAR_DB_NAME]

    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Establish tag columns.
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    span_col = col
    col += 1
    interv_mstID_col = col
    col += 1
    #interv_name_col = col
    col += 1
    first_data_col = col

    row = 2
    year_row = row
    row += 1
    first_data_row = row

    # Get each row from the sheet and create a dictionary for each each country
    first_year = gbc.GB_NOT_FOUND
    final_year = gbc.GB_NOT_FOUND
    for row in islice(sheet.values, year_row - 1, year_row):
        first_year = row[first_data_col - 1]
        final_year = row[num_cols - 1]

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        
        span_val = row[span_col - 1]

        # Create new dict if we land on a new country, as well as a new intervention list
        if span_val == IC_START:
            country_dict = {}
            intervention_list = []

            country_dict[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PbYDB] = row[iso_alpha_3_country_code_col - 1]
            country_dict[icdbc.IC_FIRST_YEAR_KEY_PbYDB] = first_year
            country_dict[icdbc.IC_FINAL_YEAR_KEY_PbYDB] = final_year

        # Every row represents one intervention's PINs for a span of years
        intervention_dict = {}
        intervention_dict[icdbc.IC_INTERV_MST_ID_KEY_PbYDB] = row[interv_mstID_col - 1]
        intervention_dict[icdbc.IC_PIN_KEY_PbYDB] = []
        
        for col, col_val in enumerate(row):
            if col >= (first_data_col - 1):
                intervention_dict[icdbc.IC_PIN_KEY_PbYDB].append(col_val)

        intervention_list.append(intervention_dict)

        if span_val == IC_END:
            country_dict[icdbc.IC_INTERVENTION_LIST_KEY_PbYDB] = intervention_list
            country_list.append(country_dict)

    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_PIN_BY_YEAR_DB_DIR
    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PbYDB]
        with open(IC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished IC PIN by year DB')

def upload_PIN_by_year_DB_IC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_PIN_BY_YEAR_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_IC_CONTAINER, walk_path, version)
    log('Uploaded IC PIN by year DB')


