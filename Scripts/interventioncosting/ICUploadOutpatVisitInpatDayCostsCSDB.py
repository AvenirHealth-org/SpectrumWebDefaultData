import numpy as np
import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICConstLink as iccl
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
IC_OTHER_RECURRENT_COSTS    = '<Other-recurrent Costs>'
IC_CAPITAL_COSTS            = '<Capital Costs>'

def create_outpat_visit_inpat_day_costs_CS_DB_IC(version = str):

    log('Creating IC outpatient visit inpatient day costs CS DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_OUTPAT_VISIT_INPAT_DAY_COSTS_CS_DB_NAME]

    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Establish tag columns. 
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    #region_col = col
    col += 1
    first_data_col = col

    # Establish tag rows. 
    row = 1
    tag_row = row
    row += 1
    #visit_type_name_row = row
    row += 1
    #visit_type_mst_ID_row = row
    row += 1
    #default_deliv_chan_name_row = row
    row += 1
    #deliv_chan_ID_row = row
    row += 1

    first_data_row = row

    other_recurrent_costs_col = gbc.GB_NOT_FOUND
    capital_costs_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == IC_OTHER_RECURRENT_COSTS:
            other_recurrent_costs_col = c

        elif tag == IC_CAPITAL_COSTS:
            capital_costs_col = c

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        country_dict = {}

        other_recurrent_costs = np.full((icc.IC_NUM_VISIT_TYPES, iccl.IC_IH_NUM_DEFAULT_DELIV_CHANS_COSTED), 0.0).tolist()
        capital_costs = np.full((icc.IC_NUM_VISIT_TYPES, iccl.IC_IH_NUM_DEFAULT_DELIV_CHANS_COSTED), 0.0).tolist()

        other_recurrent_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_OUTREACH_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col - 1]
        other_recurrent_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_GEN_OUTPATIENT_SERV_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col]
        other_recurrent_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_FIRST_REFERRAL_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col + 1]
        other_recurrent_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_SEC_REFERRAL_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col + 2]
        other_recurrent_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_GEN_OUTPATIENT_SERV_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col + 3]
        other_recurrent_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_FIRST_REFERRAL_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col + 4]
        other_recurrent_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_SEC_REFERRAL_CHAN_CURR_ID - 1] = row[other_recurrent_costs_col + 5]

        capital_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_OUTREACH_CHAN_CURR_ID - 1] = row[capital_costs_col - 1]
        capital_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_GEN_OUTPATIENT_SERV_CHAN_CURR_ID - 1] = row[capital_costs_col]
        capital_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_FIRST_REFERRAL_CHAN_CURR_ID - 1] = row[capital_costs_col + 1]
        capital_costs[icc.IC_OUTPATIENT_VISIT_CURR_ID - 1][iccl.IC_IH_SEC_REFERRAL_CHAN_CURR_ID - 1] = row[capital_costs_col + 2]
        capital_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_GEN_OUTPATIENT_SERV_CHAN_CURR_ID - 1] = row[capital_costs_col + 3]
        capital_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_FIRST_REFERRAL_CHAN_CURR_ID - 1] = row[capital_costs_col + 4]
        capital_costs[icc.IC_INPATIENT_DAY_CURR_ID - 1][iccl.IC_IH_SEC_REFERRAL_CHAN_CURR_ID - 1] = row[capital_costs_col + 5]

        country_dict[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_OVIDCCSDB] = row[iso_alpha_3_country_code_col - 1]
        country_dict[icdbc.IC_OTHER_RECURRENT_COSTS_KEY_OVIDCCSDB] = other_recurrent_costs
        country_dict[icdbc.IC_CAPITAL_COSTS_KEY_OVIDCCSDB] = capital_costs

        country_list.append(country_dict)

    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_OUTPAT_VISIT_INPAT_DAY_COSTS_CS_DB_DIR
    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_OVIDCCSDB]
        with open(IC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished IC outpatient visit inpatient day costs CS DB')

def upload_outpat_visit_inpat_day_costs_CS_DB_IC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_OUTPAT_VISIT_INPAT_DAY_COSTS_CS_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_IC_CONTAINER, walk_path, version)
    log('Uploaded IC outpatient visit inpatient day costs CS DB')
            
  