import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_NUMERIC_COUNTRY_CODE              = '<ISO Numeric Country Code>'
IC_COUNTRY_NAME                          = '<Country Name>'
IC_ISO_ALPHA_3_COUNTRY_CODE              = '<ISO Alpha-3 Country Code>'
IC_SUPPLY_CHAIN_STATUS                   = '<Supply Chain Status>'
IC_SUPPLY_CHAIN_COST_PERC_COMMODITY_COST = '<Supply Chain Cost as a Percent of Commodity Cost>'
IC_INCOME_LEVEL                          = '<Income Level>'
IC_INFRASTRUCTURE_INVESTMENT_RATIO       = '<Infrastructure Investment Ratio>' 
IC_OTHER_TO_INVESTMENT_COSTS_RATIO       = '<Other to Intervention Costs Ratio>'

def create_health_system_costs_DB_IC(version = str):

    log('Creating IC health system costs DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IC) + '\ICModData.xlsx')
    sheet = wb[icc.IC_HEALTH_SYSTEM_COSTS_DB_NAME]
    
    # first row of data after tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    #iso_numeric_country_code_col = gbc.GB_NOT_FOUND
    #country_name_col = gbc.GB_NOT_FOUND
    iso_alpha_3_country_code_col = gbc.GB_NOT_FOUND
    supply_chain_status_col = gbc.GB_NOT_FOUND
    supply_chain_cost_perc_commodity_cost_col = gbc.GB_NOT_FOUND
    income_level_col = gbc.GB_NOT_FOUND
    infrastructure_investment_ratio_col = gbc.GB_NOT_FOUND
    other_to_investment_costs_ratio_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        #if tag == IC_ISO_NUMERIC_COUNTRY_CODE:
        #    iso_numeric_country_code_col = c

        #elif tag == IC_COUNTRY_NAME:
        #    country_name_col = c

        if tag == IC_ISO_ALPHA_3_COUNTRY_CODE:
            iso_alpha_3_country_code_col = c

        elif tag == IC_SUPPLY_CHAIN_STATUS:
            supply_chain_status_col = c

        elif tag == IC_SUPPLY_CHAIN_COST_PERC_COMMODITY_COST:
            supply_chain_cost_perc_commodity_cost_col = c

        elif tag == IC_INCOME_LEVEL:
            income_level_col = c

        elif tag == IC_INFRASTRUCTURE_INVESTMENT_RATIO:
            infrastructure_investment_ratio_col = c

        elif tag == IC_OTHER_TO_INVESTMENT_COSTS_RATIO:
            other_to_investment_costs_ratio_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        country_dict = {}

        country_dict[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_HSCDB] = row[iso_alpha_3_country_code_col - 1]
        country_dict[icdbc.IC_SUPPLY_CHAIN_COST_PERC_COMMODITY_COST_KEY_HSCDB] = row[supply_chain_cost_perc_commodity_cost_col - 1]
        country_dict[icdbc.IC_INFRASTRUCTURE_INVESTMENT_RATIO_KEY_HSCDB] = row[infrastructure_investment_ratio_col - 1]
        country_dict[icdbc.IC_OTHER_TO_INVESTMENT_COSTS_RATIO_KEY_HSCDB] = row[other_to_investment_costs_ratio_col - 1]
        country_dict[icdbc.IC_INCOME_LEVEL_KEY_HSCDB] = row[income_level_col - 1]
        country_dict[icdbc.IC_SUPPLY_CHAIN_STATUS_KEY_HSCDB] = row[supply_chain_status_col - 1]
        country_list.append(country_dict)

    IC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_HEALTH_SYSTEM_COSTS_DB_DIR
    os.makedirs(IC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_HSCDB]
        with open(IC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished IC health system costs DB')

def upload_health_system_costs_DB_IC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_IC) + '\\' + icc.IC_HEALTH_SYSTEM_COSTS_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_IC_CONTAINER, walk_path, version)
    log('Uploaded IC health system costs DB')
            
