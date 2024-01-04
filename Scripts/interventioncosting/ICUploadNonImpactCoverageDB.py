import os
import ujson
from itertools import islice
from openpyxl import load_workbook


from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IC.ICConst as icc
import SpectrumCommon.Const.IC.ICDatabaseConst as icdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
IC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>' 
IC_COUNTRY_NAME             = '<Country Name>'
IC_INTERV_MST_ID            = '<Intervention Master ID>'
IC_INTERV_NAME              = '<Intervention Name>'

def create_non_impact_coverage_DB_IC(version_str):
    log('Creating IC non impact coverage DB')
    IC_path = os.getcwd() + '\Tools\DefaultDataManager\IC\\'
    IC_mod_data_FQ_name = IC_path + 'ModData\ICModData.xlsx'

    country_list = []
    wb = load_workbook(IC_mod_data_FQ_name)
    sheet = wb[icc.IC_NON_IMPACT_COVERAGE_DB_NAME]
    
    num_rows = sheet.max_row
    #num_cols = sheet.max_column

    # Establish tag columns.  Not many, so don't bother searching for them.
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    first_data_col = col

    # Establish tag rows. Not many, so don't bother searching for them.
    row = 2
    interv_mstID_row = row
    row += 1
    #interv_name_row = row
    row += 1
    first_data_row = row

    # Intervention master ID row
    interv_mstID_row_list = []
    for row in islice(sheet.values, interv_mstID_row - 1, interv_mstID_row):
        interv_mstID_row_list = row

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}
        intervention_list = []

        iso3 = row[iso_alpha_3_country_code_col - 1]

        if iso3 != None:
            country_dict[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_NICDB] = iso3

            for col, col_val in enumerate(row):
                if col >= (first_data_col - 1):
                    intervention_dict = {}
                    intervention_dict[icdbc.IC_INTERV_MST_ID_KEY_NICDB] = interv_mstID_row_list[col]
                    intervention_dict[icdbc.IC_NON_IMPACT_COV_KEY_NICDB] = col_val
                    intervention_list.append(intervention_dict)
                    #log(intervention_dict)

            country_dict[icdbc.IC_INTERVENTION_LIST_KEY_NICDB] = intervention_list
            country_list.append(country_dict)

    os.makedirs(IC_path + '\JSON\\' + icc.IC_NON_IMPACT_COVERAGE_DB_DIR + '\\', exist_ok = True)
    for country_obj in country_list:    
        iso3 = country_obj[icdbc.IC_ISO_ALPHA_3_COUNTRY_CODE_KEY_NICDB]
        with open(IC_path + '\JSON\\' + icc.IC_NON_IMPACT_COVERAGE_DB_DIR + '\\' + iso3 + '_' + version_str + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)
    log('Finished IC non impact coverage DB')

def upload_non_impact_coverage_DB_IC(version):
    connection =  os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV]
    walk_path = os.getcwd() + '\Tools\DefaultDataManager\IC\JSON\\' + icc.IC_NON_IMPACT_COVERAGE_DB_DIR + '\\'

    for subdir, dirs, files in os.walk(walk_path):
        for file in files:
            FQName = os.path.join(subdir, file)
            #log(FQName)
            GB_upload_file(connection, gbc.GB_IC_CONTAINER, icc.IC_NON_IMPACT_COVERAGE_DB_DIR + '\\' + file, FQName)
    log('Uploaded IC non impact coverage DB')
