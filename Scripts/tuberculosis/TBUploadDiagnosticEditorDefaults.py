import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np

from AvenirCommon.Logger import log
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan



def create_TB_DiagEditorDefaults(version):
    
    # TB_Path = os.getcwd()+'\Tools\DefaultDataManager\TB\\'
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    who_tb_fqname = f'{default_path}\\SourceData\\tuberculosis\\TBDiagnosticEditorDefaults2024.xlsx'
    xlsx = openpyxl.load_workbook(who_tb_fqname, read_only=False, keep_vba=False, data_only=True, keep_links=True)
    
    diagnosticEditorDefaults = {}
    # for country_cell in xlsx['Countries']['C']:
        # iso3=country_cell.value

    patient_sheets = [
        'PulmTB HIVNeg Child',
        'PulmTB HIVNeg Adult',
        'PulmTB HIVPos Child 0t9',
        'PulmTB HIVPos Child 10t14',
        'PulmTB HIVPos Adult',
        'ExPulmTB HIVNeg Child',
        'ExPulmTB HIVNeg Adult',
        'ExPulmTB HIVPos Child 0t9',
        'ExPulmTB HIVPos Child 10t14',
        'ExPulmTB HIVPos Adult',
    ]

    
    household_sheets = [
        'Household_0_4',
        'Household_5_14',
        'Household_15+'
    ]

    art_sheets = [
        'ART_0_9',
        'ART_10_14',
        'ART_15+'
    ]

    hrg_sheets = [
        'HighRiskGroups'
    ]

    sectors = [(patient_sheets, 'patients'), 
               (household_sheets, 'households'), 
               (art_sheets, 'art'),
               (hrg_sheets, 'highriskgroups')]

    for sect in sectors:
        sheets = sect[0]
        name   = sect[1]
        diagnosticEditorDefaults[name] = {
            'methods':[],
            'IDs':[],
            'values':[],
            'sensitivity':[],
            'specificity':[]
        }
        
        for page_name in sheets:
            sheet = xlsx[page_name]

            methods = []
            IDs     = []
            values  = [] 
            sensitivity = [] 
            specificity = [] 
            for row in range(3, xlsx[page_name].max_row+1):
                row_str = str(row)
                method_id   = sheet['G'+row_str].value
                if  method_id != None:
                    methods.append(method_id)
                    IDs.append([])
                    values.append([])
                    sensitivity.append([])
                    specificity.append([])
                    continue
                if sheet['C'+row_str].value != None:
                    IDs[-1].append(sheet['C'+row_str].value)
                    values[-1].append(sheet['D'+row_str].value)
                    sensitivity[-1].append(sheet['E'+row_str].value)
                    specificity[-1].append(sheet['F'+row_str].value)
                
            #    if
            diagnosticEditorDefaults[name]['methods'].append(methods)
            diagnosticEditorDefaults[name]['IDs'].append(IDs)
            diagnosticEditorDefaults[name]['values'].append(values)
            diagnosticEditorDefaults[name]['sensitivity'].append(sensitivity)
            diagnosticEditorDefaults[name]['specificity'].append(specificity)
        #print(page_name)
        
    # os.makedirs(TB_Path+'\JSON\diag\\', exist_ok=True)
    
    os.makedirs(default_path+'\\JSONData\\tuberculosis\\diag\\', exist_ok=True)
    with open(default_path+'\\JSONData\\tuberculosis\\diag\\diagnosticEditorDefaults'+version+'.JSON', 'w') as f:
        ujson.dump(diagnosticEditorDefaults, f)
    # with open(TB_Path+'\JSON\diag\diagnosticEditorDefaults'+version+'.JSON', 'w') as f:

def upload_TB_DiagEditorDefaults(version):
    connection =  os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    FQName = default_path+'\\JSONData\\tuberculosis\\diag\\diagnosticEditorDefaults'+version+'.JSON'
    log(FQName)
    GB_upload_file(connection, 'tuberculosis', 'diagnosticDefaults_'+version+'.JSON', FQName)
        
    log('Uploaded TB DiagEditorDefault')
   