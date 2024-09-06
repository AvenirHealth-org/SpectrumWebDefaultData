import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np
import cProfile
import pstats

import logging
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan



def create_TB_DyanmicalModelData(version):
    
    # TB_Path = os.getcwd()+'\\Tools\DefaultDataManager\\TB\\'
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    who_tb_fqname = f'{default_path}\\SourceData\\tuberculosis\\WHO_TB_CountryData2022.xlsx'
    xlsx = openpyxl.load_workbook(who_tb_fqname, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    
    mort_inc_start_year = 2000
    noti_start_year    = 2015
    # profile = cProfile.Profile()
    # profile.enable()
    # for country_cell in xlsx['Countries']['C']:
    indicators = {}
    for row in range(2, xlsx['WHO_Data_Map'].max_row+1):
        row_str = str(row)
        field_name = xlsx['WHO_Data_Map']['A'+row_str].value
        sheet_name = xlsx['WHO_Data_Map']['B'+row_str].value
        years = xlsx['WHO_Data_Map']['C'+row_str].value
        if years:
           years = years.split(':')
        else:
            continue
        first_year = int(years[0])
        final_year = int(years[1])
        if sheet_name not in indicators:
            indicators[sheet_name] = {}
        indicators[sheet_name][field_name] = {'firstYear': first_year, 'finalYear':final_year}
    pass
        # iso3=country_cell.value
    for iso3 in ('IND', 'ZMB', 'ZWE'):
        if (iso3=='iso3'):# or (iso3=='GRL'):
            continue
        logging.debug(iso3)
        
        # pages = ['TB_burden_countries']
        country = {}
        for page_name in indicators:
            sector_name = page_name
            country[sector_name] = {}
            
            if True:
                # col_names = {}
                # for row in range(6, xlsx[page_name].max_row+1):
                #     row_str = str(row)
                #     col_name = xlsx[page_name]['A'+row_str].value
                #     var_name = xlsx[page_name]['B'+row_str].value

                #     if col_name=='year':
                #         start_year =  int(var_name.split(':')[0])
                #         country[sector_name]['startYear'] = start_year
                #     if var_name == None:
                #         continue
                #     elif var_name == 'x':
                #         var_name = col_name

                #     col_names[col_name] = var_name
                    

                col_letters = {}
                for col in xlsx[page_name].columns:
                    name = col[0]._value
                    if name in indicators[page_name]:
                        col_letters[col[0]._value] = col[0].column_letter
                    
                for row in range(2, xlsx[page_name].max_row+1):
                    row_str = str(row)
                    row_year = int(xlsx[page_name][f'F{row}'].value) 

                    if  iso3==xlsx[page_name]['C'+row_str].value:
                        for col_name in indicators[page_name]:
                            letter = col_letters[col_name]
                            
                            if row_year<indicators[page_name][col_name]['firstYear']: 
                                continue #Skip rows that are years before the start year

                            val = xlsx[page_name][letter+row_str].value
                            if not (col_name in country[sector_name]):
                                country[sector_name][col_name] = {'firstYear':indicators[page_name][col_name]['firstYear'], 'values':[val]}
                            else:
                                country[sector_name][col_name]['values'].append(val)    
                        #log(xlsx[page_name+'_2022']['F'+row_str].value)
                        if  iso3!=xlsx[page_name]['C'+str(row+1)].value:
                            break

                # country[sector_name]['startYear'] = xlsx[page_name+'_2022']['F2'].value 


            pass

        
        os.makedirs(default_path+'\\JSONData\\tuberculosis\\dynamicalmodel\\countries\\', exist_ok=True)
        with open(default_path+'\\JSONData\\tuberculosis\\dynamicalmodel\\countries\\'+iso3+'_'+version+'.JSON', 'w') as f:
            ujson.dump(country, f)
        if iso3 == 'KEN':
            iso3 = 'SAMP'
            with open(default_path+'\\JSONData\\tuberculosis\\dynamicalmodel\\countries\\'+iso3+'_'+version+'.JSON', 'w') as f:
                ujson.dump(country, f)
    # profile.disable()
    # ps = pstats.Stats(profile)
    # ps.sort_stats('calls', 'cumtime')
    # ps.print_stats(50)




def upload_TB_DyanmicalModelData(version):
    connection =  os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
    countries = []
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    json_path= default_path+'\\JSONData\\tuberculosis\\dynamicalmodel\\countries\\'
    for subdir, dirs, files in os.walk(json_path):
        for file in files:
        # for file in ('IND_V3.JSON'):
            FQName = os.path.join(subdir, file)
            if version in FQName:
                print(FQName)
                GB_upload_file(connection, 'tuberculosis', 'dynamicalmodel\\countries\\'+file, FQName)
        
    logging.debug('Uploaded TB who db json')
   