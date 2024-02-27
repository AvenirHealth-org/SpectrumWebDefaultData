import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file, GB_get_db_json
import ujson
import numpy as np

from AvenirCommon.Logger import log
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan
from AvenirCommon.Util import GBRange


def create_TB_fort_input(version):    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    who_tb_fqname = f'{default_path}\\SourceData\\tuberculosis\\WHO_TB_CountryData2022.xlsx'
    xlsx = openpyxl.load_workbook(who_tb_fqname, read_only=False, keep_vba=False, data_only=False, keep_links=True)

    failed_countries = []
    for country_cell in xlsx['Countries']['C']:
        iso3=country_cell.value
    # for iso3 in ('BFA',):# ('SWZ', 'ETH', 'ZMB', 'ZWE'):
        if iso3=='iso3':
            continue
        try:
            log(iso3)
            # pages = ['TB_burden_countries']
            fort_input = {
                "year": [],
                "iHat": [],
                "sEi" : [],
                "nHat": [],
                "sEn" : [],
                "mHat": [],
                "sEm" : [],
                "pHat": [],
                "sEp" : []
            }
            sheet = xlsx['TB_burden_countries_2022']
            for row in range(1, sheet.max_row+1):
                row_str = str(row)
                if  iso3==sheet['C'+row_str].value:
                    year          = sheet['F'+row_str].value
                    e_inc_num     = sheet['K'+row_str].value
                    e_inc_num_hi  = sheet['M'+row_str].value
                    e_inc_num_lo  = sheet['L'+row_str].value
                    e_mort_num    = sheet['AL'+row_str].value
                    e_mort_num_hi = sheet['AN'+row_str].value
                    e_mort_num_lo = sheet['AM'+row_str].value

                    
                    e_inc_num     = e_inc_num     if e_inc_num     != None else 0
                    e_inc_num_hi  = e_inc_num_hi  if e_inc_num_hi  != None else 0
                    e_inc_num_lo  = e_inc_num_lo  if e_inc_num_lo  != None else 0
                    e_mort_num    = e_mort_num    if e_mort_num    != None else 0
                    e_mort_num_hi = e_mort_num_hi if e_mort_num_hi != None else 0
                    e_mort_num_lo = e_mort_num_lo if e_mort_num_lo != None else 0

                    fort_input["year"].append(year)
                    fort_input["iHat"].append(e_inc_num)
                    fort_input["sEi" ].append((e_inc_num_hi-e_inc_num_lo)/3.92)
                    fort_input["mHat"].append(e_mort_num)
                    fort_input["sEm" ].append((e_mort_num_hi-e_mort_num_lo)/3.92)
                    fort_input["pHat"].append(None)
                    fort_input["sEp" ].append(0.0305) # 3.05%


            sheet = xlsx['TB_notifications_2022']
            for row in range(1, sheet.max_row+1):
                row_str = str(row)
                if  iso3==sheet['C'+row_str].value:
                    year = sheet['F'+row_str].value
                    if year>= fort_input["year"][0]:
                        c_newinc = sheet['X'+row_str].value
                        ret_nrel = sheet['U'+row_str].value
                        
                        # error checking
                        c_newinc = c_newinc if c_newinc!=None else 0 
                        ret_nrel = ret_nrel if ret_nrel!=None else 0 
                        
                        noti = c_newinc+ret_nrel 
                        noti_hi = (1+0.15)*noti
                        noti_lo = (1-0.15)*noti 
                        sEn = (noti_hi-noti_lo)/3.92

                        fort_input["nHat"].append(noti)
                        fort_input["sEn" ].append(sEn)
            pass


            fort_input["modelType"] = "IP"

            num_years = 2050-fort_input["year"][0]+1
            fort_input["sEp"]  = [None]*num_years
            # fort_input["pHat"] = [None]*num_years
            fort_input["hRd"]  = [1]*num_years
            fort_input["hRi"]  = [1]*num_years
            fort_input["oRt"]  = [1]*num_years    
            
            who_db  = GB_get_db_json(os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION'], 'tuberculosis', 'countries/'+iso3+'_V3.JSON')    
            
            #Scale up fort notifications to match Noti_Distribution WHO notifications
            noti_scale  = who_db['Noti_Distribution']['c_notification']/fort_input["nHat"][2022-2000]
            fort_input["nHat"] = (np.array(fort_input["nHat"])*noti_scale).tolist()   

            PropHIVp         = who_db['Noti_Distribution']['newrel_hivpos_perc']
            PropHIVn         = 1 - PropHIVp
            PropHIVponART    = who_db['Noti_Distribution']['newrel_art_perc']

            PropHIVpNoART    = PropHIVp * (1- PropHIVponART)
            PropHIVponART6m  = PropHIVp * 0.1 * PropHIVponART
            PropHIVponART12m = PropHIVp * 0.9 * PropHIVponART
            # check that     PropHIVn+PropHIVpNoART+PropHIVponART6m+PropHIVponART12m=1
            if (PropHIVn+PropHIVpNoART+PropHIVponART6m+PropHIVponART12m)!=1:
                print(f'PropHIVn+PropHIVpNoART+PropHIVponART6m+PropHIVponART12m!=1 {iso3}')
                failed_countries.append(iso3)
            TXf  = PropHIVn*(3/100)+PropHIVpNoART*(9/100)+PropHIVponART6m*(6/100)+PropHIVponART12m*(4/100)
            fort_input["tXf"] = [TXf]*num_years
            # check value for Ethiopia for which TXf ~ 0.035. Check value for Eswatini should closer to the HIV values

            for yr in GBRange(fort_input["year"][-1]+1, 2050):
                fort_input["year"].append(yr)    
                for key in ["iHat", "sEi", "nHat", "sEn", "mHat", "sEm"]:
                    fort_input[key].append(None)#fort_inputs[key][-1])

            os.makedirs(default_path+'\\JSONData\\tuberculosis\\fortinputs\\', exist_ok=True)
            with open(default_path+'\\JSONData\\tuberculosis\\fortinputs\\'+iso3+'_'+version+'.JSON', 'w') as f:
                ujson.dump(fort_input, f)
        except:
            print(f'{iso3} exception')

def upload_tb_fort_inputs_db(version):
    connection =  os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']

    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    json_path= default_path+'\\JSONData\\tuberculosis\\fortinputs\\'
    for subdir, dirs, files in os.walk(json_path):
        for file in files:
        # for file in ('BFA_'+version+'.JSON', ):

            FQName = os.path.join(subdir, file) 
            if version in FQName:
                log(FQName)
                GB_upload_file(connection, 'tuberculosis', 'fort\\inputs\\'+file, FQName)
    
        
    log('Uploaded fort inputs db json')
   