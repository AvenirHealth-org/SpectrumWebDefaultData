import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np

from AvenirCommon.Logger import log
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan



def create_TBScreeningDefaults(version):
    
    TB_Path = os.getcwd()+'\Tools\DefaultDataManager\TB\\'
    xlsx = openpyxl.load_workbook(
                '.\Tools\DefaultDataManager\TB\ModData\TBScreeningDefaults.xlsx', read_only=False, keep_vba=False, data_only=True, keep_links=True)
    
    diagnosticEditorDefaults = {
        'patients':{
            'methods':[],
            'IDs':[],
            'values':[],
            'sensitivity':[],
            'specificity':[]}
    }

    # for country_cell in xlsx['Countries']['C']:
        # iso3=country_cell.value

    patient_sheets = [
        'PulmTB HIVNeg Child',
        'PulmTB HIVNeg Adult',
        'PulmTB HIVPos Child 0t9',
        'PulmTB HIVPos Child 10t14',
        'PulmTB HIVPos Adult',
        'ExPulmTB HIVNeg Child',
        'ExPulmTB HIVNeg Adult',
        'ExPulmTB HIVPos Child 0t9',
        'ExPulmTB HIVPos Child 10t14',
        'ExPulmTB HIVPos Adult',
    ]
    for page_name in patient_sheets:
        sheet = xlsx[page_name]

        methods = []
        IDs     = []
        values  = [] 
        sensitivity = [] 
        specificity = [] 
        for row in range(3, xlsx[page_name].max_row+1):
            row_str = str(row)
            method_id   = sheet['G'+row_str].value
            if  method_id != None:
                methods.append(method_id)
                IDs.append([])
                values.append([])
                sensitivity.append([])
                specificity.append([])
                continue
            IDs[-1].append(sheet['C'+row_str].value)
            values[-1].append(sheet['D'+row_str].value)
            sensitivity[-1].append(sheet['E'+row_str].value)
            specificity[-1].append(sheet['F'+row_str].value)
            
        #    if
        diagnosticEditorDefaults['patients']['methods'].append(methods)
        diagnosticEditorDefaults['patients']['IDs'].append(IDs)
        diagnosticEditorDefaults['patients']['values'].append(values)
        diagnosticEditorDefaults['patients']['sensitivity'].append(sensitivity)
        diagnosticEditorDefaults['patients']['specificity'].append(specificity)
        #print(page_name)
        
    os.makedirs(TB_Path+'\JSON\screen\\', exist_ok=True)
    with open(TB_Path+'\JSON\screen\screeningDefaults'+version+'.JSON', 'w') as f:
        ujson.dump(diagnosticEditorDefaults, f)

def upload_TBScreeningDefaults(version):
    connection =  os.environ['AVENIR_SPEC_DEFAULT_DATA_CONNECTION']
    
    FQName = os.getcwd()+'\Tools\DefaultDataManager\TB\JSON\screen\screeningDefaults'+version+'.JSON'
    log(FQName)
    GB_upload_file(connection, 'tuberculosis', 'screeningDefaults_'+version+'.JSON', FQName)
        
    log('Uploaded TB who db json')
   