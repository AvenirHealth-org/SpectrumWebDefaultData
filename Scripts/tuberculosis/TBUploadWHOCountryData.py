import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np
import cProfile
import pstats

import logging
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan



def create_TB_WHOCountryData(version):
    
    # TB_Path = os.getcwd()+'\\Tools\DefaultDataManager\\TB\\'
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    who_tb_fqname = f'{default_path}\\SourceData\\tuberculosis\\WHO_TB_CountryData2023.xlsx'
    xlsx = openpyxl.load_workbook(who_tb_fqname, read_only=False, keep_vba=False, data_only=False, keep_links=True)
    
    # profile = cProfile.Profile()
    # profile.enable()
    countries = []
    for country_cell in xlsx['Countries']['C']:
        countries.append(country_cell.value)
    # countries = ('ZWE', 'DOM', 'SSD')
    for iso3 in countries: #81 GRL 137 ANT 174 SCG 197 TKL
        if (iso3=='iso3'):# or (iso3=='GRL'):
            continue
        print(iso3)
        try:
            # if os.path.exists(default_path+'\\JSONData\\tuberculosis\\countries\\'+iso3+'_'+version+'.JSON'):
                # continue #Skip over files that already exist
            
            pages = {'TB_burden_countries': 'burden',
                    'TB_notifications': 'notifications',
                    'Noti_Distribution':'Noti_Distribution',
                    'LTBI_estimates': 'estimates',
                    'TB_contact_tpt': 'contact',
                    'TB_dr_surveillance': 'dr_surveillance',
                    'MDR_RR_TB_burden_estimates': 'MDR_RR',
                    'TB_Projections':'projections',
                    'RR_Profile':'RR_Profile',
                    'DST_Coverage':'coverage',
                    'ScreeningSensSpec':'ScreeningSensSpec'
                    }
            # pages = ['TB_burden_countries']
            country = {}
            for page_name in pages:
                sector_name = pages[page_name]
                country[sector_name] = {}
                if page_name =='TB_Projections':
                    
                    for row in range(1, xlsx[page_name].max_row+1):
                        row_str = str(row)

                        if  iso3==xlsx[page_name]['C'+row_str].value:
                            if not ('startYear' in country[sector_name]):
                                country[sector_name]['startYear'] = xlsx[page_name]['F'+row_str].value
                                country[sector_name]['pop'] = []
                                country[sector_name]['notification'] = []
                                country[sector_name]['prevalence'] = []
                                country[sector_name]['prevalencePct'] = []  

                            country[sector_name]['pop'].append(xlsx[page_name]['G'+row_str].value)
                            country[sector_name]['notification'].append(xlsx[page_name]['H'+row_str].value)
                            country[sector_name]['prevalence'].append(xlsx[page_name]['I'+row_str].value)
                            country[sector_name]['prevalencePct'].append(xlsx[page_name]['J'+row_str].value)
                            
                    pass
                elif page_name =='ScreeningSensSpec':
                    country['Sensitivity'] = {}
                    country['Specificity'] = {}

                    country['Sensitivity']['Patient'] = []
                    country['Sensitivity']['HouseHold'] = []
                    country['Sensitivity']['ARTWithSerious'] = []
                    country['Sensitivity']['ARTWithoutSerious'] = []
                    country['Sensitivity']['HRG'] = []

                    country['Specificity']['Patient'] = []
                    country['Specificity']['HouseHold'] = []
                    country['Specificity']['ARTWithSerious'] = []
                    country['Specificity']['ARTWithoutSerious'] = []
                    country['Specificity']['HRG'] = []

                    for row in [7, 10, 11, 12, 13]:
                        row_str = str(row)
                        country['Sensitivity']['Patient'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['Patient'].append(xlsx[page_name]['E'+row_str].value)

                    for row in range(15, 20):
                        row_str = str(row)
                        country['Sensitivity']['Patient'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['Patient'].append(xlsx[page_name]['E'+row_str].value)
                        
                    for row in range(22, 25):
                        row_str = str(row)
                        country['Sensitivity']['HouseHold'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['HouseHold'].append(xlsx[page_name]['E'+row_str].value)

                    for row in [31, 34, 37]:
                        row_str = str(row)
                        country['Sensitivity']['ARTWithSerious'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['ARTWithSerious'].append(xlsx[page_name]['E'+row_str].value)

                    for row in [28, 33, 36]:
                        row_str = str(row)
                        country['Sensitivity']['ARTWithoutSerious'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['ARTWithoutSerious'].append(xlsx[page_name]['E'+row_str].value)

                    for row in [43, 48]:
                        row_str = str(row)
                        country['Sensitivity']['HRG'].append(xlsx[page_name]['D'+row_str].value)
                        country['Specificity']['HRG'].append(xlsx[page_name]['E'+row_str].value)
                    pass

                
                elif page_name =='Noti_Distribution':
                    for row in range(1, xlsx[page_name].max_row+1):
                        row_str = str(row)

                        if  iso3==xlsx[page_name]['C'+row_str].value:
                            tmp = row_str
                            # Numbers
                            country[sector_name]['c_notification']       = xlsx[page_name]['F'+row_str].value
                            country[sector_name]['c_newinc_num']         = xlsx[page_name]['G'+row_str].value
                            country[sector_name]['ret_nrel_num']         = xlsx[page_name]['H'+row_str].value
                            country[sector_name]['new_labconf_num']      = xlsx[page_name]['I'+row_str].value
                            country[sector_name]['new_clindx_num']       = xlsx[page_name]['J'+row_str].value
                            country[sector_name]['new_ep_num']           = xlsx[page_name]['K'+row_str].value
                            country[sector_name]['ret_rel_labconf_num']  = xlsx[page_name]['L'+row_str].value
                            country[sector_name]['ret_rel_clindx_num']   = xlsx[page_name]['M'+row_str].value
                            country[sector_name]['ret_rel_ep_num']       = xlsx[page_name]['N'+row_str].value
                            
                            country[sector_name]['newrel_04_num']     = xlsx[page_name]['Z'+row_str].value
                            country[sector_name]['newrel_59_num']     = xlsx[page_name]['AA'+row_str].value
                            country[sector_name]['newrel_1014_num']   = xlsx[page_name]['AB'+row_str].value
                            country[sector_name]['newrel_514_num']    = xlsx[page_name]['AC'+row_str].value
                            country[sector_name]['newrel_014_num']    = xlsx[page_name]['AD'+row_str].value
                            country[sector_name]['newrel_1519_num']   = xlsx[page_name]['AE'+row_str].value
                            country[sector_name]['newrel_2024_num']   = xlsx[page_name]['AF'+row_str].value
                            country[sector_name]['newrel_1524_num']   = xlsx[page_name]['AG'+row_str].value
                            country[sector_name]['newrel_2534_num']   = xlsx[page_name]['AH'+row_str].value
                            country[sector_name]['newrel_3544_num']   = xlsx[page_name]['AI'+row_str].value
                            country[sector_name]['newrel_4554_num']   = xlsx[page_name]['AJ'+row_str].value
                            country[sector_name]['newrel_5564_num']   = xlsx[page_name]['AK'+row_str].value
                            country[sector_name]['newrel_65_num']     = xlsx[page_name]['AL'+row_str].value
                            country[sector_name]['newrel_15plus_num'] = xlsx[page_name]['AM'+row_str].value

                            row_str = tmp
                            if xlsx[page_name]['BI'+row_str].value==0:
                                row_str = '221'

                            # Percents
                            country[sector_name]['c_newinc_perc']        = xlsx[page_name]['Q'+row_str].value
                            country[sector_name]['ret_nrel_perc']        = xlsx[page_name]['R'+row_str].value
                            country[sector_name]['new_labconf_perc']     = xlsx[page_name]['S'+row_str].value
                            country[sector_name]['new_clindx_perc']      = xlsx[page_name]['T'+row_str].value
                            country[sector_name]['new_ep_perc']          = xlsx[page_name]['U'+row_str].value
                            country[sector_name]['ret_rel_labconf_perc'] = xlsx[page_name]['V'+row_str].value
                            country[sector_name]['ret_rel_clindx_perc']  = xlsx[page_name]['W'+row_str].value
                            country[sector_name]['ret_rel_ep_perc']      = xlsx[page_name]['X'+row_str].value
                            
                            country[sector_name]['newrel_04_perc']     = xlsx[page_name]['AQ'+row_str].value
                            country[sector_name]['newrel_59_perc']     = xlsx[page_name]['AR'+row_str].value
                            country[sector_name]['newrel_1014_perc']   = xlsx[page_name]['AS'+row_str].value
                            country[sector_name]['newrel_514_perc']    = xlsx[page_name]['AT'+row_str].value
                            country[sector_name]['newrel_014_perc']    = xlsx[page_name]['AU'+row_str].value
                            country[sector_name]['newrel_1519_perc']   = xlsx[page_name]['AV'+row_str].value
                            country[sector_name]['newrel_2024_perc']   = xlsx[page_name]['AW'+row_str].value
                            country[sector_name]['newrel_1524_perc']   = xlsx[page_name]['AX'+row_str].value
                            country[sector_name]['newrel_2534_perc']   = xlsx[page_name]['AY'+row_str].value
                            country[sector_name]['newrel_3544_perc']   = xlsx[page_name]['AZ'+row_str].value
                            country[sector_name]['newrel_4554_perc']   = xlsx[page_name]['BA'+row_str].value
                            country[sector_name]['newrel_5564_perc']   = xlsx[page_name]['BB'+row_str].value
                            country[sector_name]['newrel_65_perc']     = xlsx[page_name]['BC'+row_str].value
                            country[sector_name]['newrel_15plus_perc'] = xlsx[page_name]['BD'+row_str].value
                            
                            country[sector_name]['newrel_hivpos_perc']   = xlsx[page_name]['BK'+row_str].value
                            country[sector_name]['newrel_art_perc']      = xlsx[page_name]['BL'+row_str].value

                            country[sector_name]['newrel_014_perc']      = ( xlsx[page_name]['AQ'+row_str].value
                                                                        + xlsx[page_name]['AR'+row_str].value
                                                                        + xlsx[page_name]['AS'+row_str].value)
                            
                            country[sector_name]['newrel_15plus_perc']  = 1-country[sector_name]['newrel_014_perc'] 

                            
                            break
        

                    pass


                elif page_name =='RR_Profile':
                    rr_profile = np.zeros((TB_NewRetreatLen, TB_ResistProfileLen))
                    for row in range(1, xlsx[page_name].max_row+1):
                        row_str = str(row)

                        if  iso3==xlsx[page_name]['C'+row_str].value:
                            if xlsx[page_name]['T'+row_str].value == '#N/A':
                                row_str = '221'
                            rr_profile[TB_NewCases][TB_Rif_Sen] = xlsx[page_name]['T'+row_str].value
                            rr_profile[TB_NewCases][TB_Rif_Res] = xlsx[page_name]['Q'+row_str].value

                            rr_profile[TB_NewCases][TB_INH_Sen] = xlsx[page_name]['U'+row_str].value
                            rr_profile[TB_NewCases][TB_INH_Res] = xlsx[page_name]['V'+row_str].value

                            rr_profile[TB_NewCases][TB_FQ_Sen ] = xlsx[page_name]['R'+row_str].value
                            rr_profile[TB_NewCases][TB_FQ_Res ] = xlsx[page_name]['S'+row_str].value

                            rr_profile[TB_NewCases][TB_XDR    ] = 0.01

                            rr_profile[TB_RetreatCases][TB_Rif_Sen] = xlsx[page_name]['AA'+row_str].value
                            rr_profile[TB_RetreatCases][TB_Rif_Res] = xlsx[page_name]['X'+row_str].value

                            rr_profile[TB_RetreatCases][TB_INH_Sen] = xlsx[page_name]['AB'+row_str].value
                            rr_profile[TB_RetreatCases][TB_INH_Res] = xlsx[page_name]['AC'+row_str].value

                            rr_profile[TB_RetreatCases][TB_FQ_Sen ] = xlsx[page_name]['Y'+row_str].value
                            rr_profile[TB_RetreatCases][TB_FQ_Res ] = xlsx[page_name]['Z'+row_str].value
                            rr_profile[TB_RetreatCases][TB_XDR    ] = 0.01
                            break

                    country[sector_name] = rr_profile.tolist()
                    pass
                elif page_name =='DST_Coverage':   
                    for row in range(1, xlsx[page_name].max_row+1):
                        row_str = str(row)

                        if  iso3==xlsx[page_name]['D'+row_str].value: 
                            country[sector_name]['DST'] = xlsx[page_name]['K'+row_str].value
                            break
                    pass
                else:
                    col_names = {}
                    for row in range(7, xlsx[page_name].max_row+1):
                        row_str = str(row)
                        col_name = xlsx[page_name]['A'+row_str].value
                        var_name = xlsx[page_name]['B'+row_str].value
                        if var_name == None:
                            continue
                        elif var_name == 'x':
                            var_name = col_name

                        col_names[col_name] = var_name

                    col_letters = {}
                    for col in xlsx[page_name+'_2023'].columns:
                        name = col[0]._value
                        if name in col_names:
                            col_letters[col[0]._value] = col[0].column_letter
                        
                    for row in range(1, xlsx[page_name+'_2023'].max_row+1):
                        row_str = str(row)

                        if  iso3==xlsx[page_name+'_2023']['C'+row_str].value:
                            for col_name in col_names:
                                letter = col_letters[col_name]
                                val = xlsx[page_name+'_2023'][letter+row_str].value
                                if not (col_name in country[sector_name]):
                                    country[sector_name][col_name] = [val]
                                else:
                                    country[sector_name][col_name].append(val)    
                            #log(xlsx[page_name+'_2023']['F'+row_str].value)
                            if  iso3!=xlsx[page_name+'_2023']['C'+str(row+1)].value:
                                break

                    country[sector_name]['startYear'] = xlsx[page_name+'_2023']['F2'].value 


                pass

            
            os.makedirs(default_path+'\\JSONData\\tuberculosis\\countries\\', exist_ok=True)
            with open(default_path+'\\JSONData\\tuberculosis\\countries\\'+iso3+'_'+version+'.JSON', 'w') as f:
                ujson.dump(country, f)
                
            if iso3 == 'KEN':
                iso3 = 'SAMP'
                print(iso3)
                with open(default_path+'\\JSONData\\tuberculosis\\countries\\'+iso3+'_'+version+'.JSON', 'w') as f:
                    ujson.dump(country, f)
        except:
            print(f'{iso3} exception')            
                    
    # profile.disable()
    # ps = pstats.Stats(profile)
    # ps.sort_stats('calls', 'cumtime')
    # ps.print_stats(50)
    pass



def upload_TB_WHOCountryData(version):
    connection =  os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
    countries = []
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    json_path= default_path+'\\JSONData\\tuberculosis\\countries\\'
    for subdir, dirs, files in os.walk(json_path):
        for file in files:
        # for iso3 in ('ZWE', 'DOM', 'SSD'):
            # file = iso3+'_'+version+'.JSON'
            FQName = os.path.join(subdir, file)
            if version in FQName:
                print(FQName)
                GB_upload_file(connection, 'tuberculosis', 'countries\\'+file, FQName)
        
    logging.debug('Uploaded TB who db json')
   