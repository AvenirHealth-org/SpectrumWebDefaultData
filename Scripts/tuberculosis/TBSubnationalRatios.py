import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np
import cProfile
import pstats

import logging
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan


def create_TB_subnationals(version):
    """
    Creates the TB subnational list JSON file to the specified database connection.
    
    Args:
        version (str): The version identifier for the subnational list.
    """
    default_path = os.path.join(os.getcwd(), __name__.split('.')[0])
    countries = ['NGA']  # Hardcoded for now, can be extended
    subnational_list = {}

    try:
        subnationals = {}
        for iso3 in countries:
            # Load the Excel file
            xlsx = openpyxl.load_workbook(f'{default_path}\\SourceData\\demproj\\NGA_SubNat_Data.xlsx', 
                                          read_only=False, 
                                          keep_vba=False, 
                                          data_only=True, 
                                          keep_links=True)

            # Extract subnational data
            distribution_sheet = xlsx['TB_IHT_SubNatData']
            first_subnat_row = 2
            last_subnat_row = 38
            subnationals[iso3] = {}
            for row in range(first_subnat_row, last_subnat_row + 1):
                subnat_id = distribution_sheet[f'A{row}'].value
                subnat_name = distribution_sheet[f'B{row}'].value
                subnationals[iso3][subnat_id] = {
                    'iso3': iso3,
                    'level': 2, #distribution_sheet[f'A{row}'].value,
                    'name': subnat_name,
                    'areaID': subnat_id,
                    'notif': distribution_sheet[f'C{row}'].value,
                    'notifProp': distribution_sheet[f'D{row}'].value,
                    'notif2023': distribution_sheet[f'E{row}'].value,
                }


        # Save the subnational list to a JSON file
        output_dir = os.path.join(default_path, 'JSONData', 'tuberculosis', 'subnationals')
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, f'{iso3}_subnat_{version}.JSON')

        with open(output_file, 'w') as f:
            ujson.dump(subnationals[iso3], f)

        logging.debug('Created subnational list successfully')

    except Exception as e:
        logging.debug(f'Error creating subnational list: {e}')

def upload_TB_subnationals(version):
    """
    Uploads the TB subnational list JSON file to the specified database connection.
    
    Args:
        version (str): The version identifier for the subnational list.
    """
    try:
        # Retrieve the database connection from environment variables
        connection = os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
        
        # Construct the file path for the subnational list JSON
        default_path = os.path.join(os.getcwd(), __name__.split('.')[0])
        json_path = os.path.join(default_path, 'JSONData', 'tuberculosis', 'subnationals')
        countries = ['NGA',]  # Hardcoded for now, can be extended
        for iso3 in countries:
            file_name = f'{iso3}_subnat_{version}.JSON'
            full_file_path = os.path.join(json_path, file_name)
            
            # Log the file path being uploaded
            logging.debug(f"Uploading file: {full_file_path}")
            
            # Upload the file to the database
            GB_upload_file(connection, 'tuberculosis', f'subnationals\\{iso3}_subnat_{version}.JSON', full_file_path)
            
            logging.debug('Uploaded TB subnational list JSON successfully')
    except Exception as e:
        logging.error(f"Failed to upload TB subnational list JSON: {e}")
    
    
def create_TB_subnational_list(version):
    """
    Creates the TB subnational list JSON file to the specified database connection.
    
    Args:
        version (str): The version identifier for the subnational list.
    """
    default_path = os.path.join(os.getcwd(), __name__.split('.')[0])
    countries = ['NGA']  # Hardcoded for now, can be extended
    subnational_list = {}

    try:
        subnationals = {}
        for iso3 in countries:
            # Load the Excel file
            xlsx = openpyxl.load_workbook(f'{default_path}\\SourceData\\demproj\\NGA_SubNat_Data.xlsx', 
                                          read_only=False, 
                                          keep_vba=False, 
                                          data_only=False, 
                                          keep_links=True)

            # Extract subnational data
            distribution_sheet = xlsx['TB_IHT_SubNatData']
            first_subnat_row = 2
            last_subnat_row = 38
            subnationals[iso3] = {}
            for row in range(first_subnat_row, last_subnat_row + 1):
                subnat_id = distribution_sheet[f'A{row}'].value
                subnat_name = distribution_sheet[f'B{row}'].value
                subnationals[iso3][subnat_id] = {
                    'iso3': iso3,
                    'level': 2, #distribution_sheet[f'A{row}'].value,
                    'name': subnat_name,
                    'areaID': subnat_id,
                }


        # Save the subnational list to a JSON file
        output_dir = os.path.join(default_path, 'JSONData', 'tuberculosis', 'subnationals')
        os.makedirs(output_dir, exist_ok=True)
        output_file = os.path.join(output_dir, f'TB_subnationalList_{version}.JSON')

        with open(output_file, 'w') as f:
            ujson.dump(subnationals, f)

        logging.debug('Created subnational list successfully')

    except Exception as e:
        logging.debug(f'Error creating subnational list: {e}')

def upload_TB_subnational_list(version):
    """
    Uploads the TB subnational list JSON file to the specified database connection.
    
    Args:
        version (str): The version identifier for the subnational list.
    """
    try:
        # Retrieve the database connection from environment variables
        connection = os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
        
        # Construct the file path for the subnational list JSON
        default_path = os.path.join(os.getcwd(), __name__.split('.')[0])
        json_path = os.path.join(default_path, 'JSONData', 'tuberculosis', 'subnationals')
        file_name = f'TB_subnationalList_{version}.JSON'
        full_file_path = os.path.join(json_path, file_name)
        
        # Log the file path being uploaded
        logging.debug(f"Uploading file: {full_file_path}")
        
        # Upload the file to the database
        GB_upload_file(connection, 'tuberculosis', f'subnationals\\TB_subnationalList_{version}.JSON', full_file_path)
        
        logging.debug('Uploaded TB subnational list JSON successfully')
    except Exception as e:
        logging.error(f"Failed to upload TB subnational list JSON: {e}")
   