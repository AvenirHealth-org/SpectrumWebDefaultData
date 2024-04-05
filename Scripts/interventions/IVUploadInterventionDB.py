import os
import ujson
from itertools import islice
from openpyxl import load_workbook
import numpy as np

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
from SpectrumCommon.Const.IV import IVConst as ivc
from SpectrumCommon.Const.IV import IVConstLink as ivcl
from SpectrumCommon.Const.IV import IVDatabaseKeys as ivdbk

# Column tag 'constants' for identifying columns.

# Columns shared by all modules

IV_ROOT_CONSTANT   = '<Root constant>'
IV_MST_ID          = '<Master ID>'
IV_MODULES         = '<Module(s)>'
IV_MODULES_ENGLISH = '<Module(s) - English>'

# Column blocks that are repeated for each module

# IV_GROUP_ORDER_IH      = '<UH - Group order>'  
# IV_GROUP_IH            = '<UH - Group>'      
IV_GROUP_MST_ID_IH     = '<UH - Group master ID>' 
# IV_SUBGROUP_CURR_ID_IH = '<UH - Subgroup>'       
IV_SUBGROUP_MST_ID_IH  = '<UH - Subgroup master ID>' 
IV_INDENTS_IH          = '<UH - Indents>'     
IV_STRING_CONSTANT_IH  = '<UH - String constant>' 
# IV_STRING_NUMBER_IH    = '<UH - String #>'       
IV_ENGLISH_STRING_IH   = '<UH - English>'     

# IV_GROUP_ORDER_TI      = '<TI - Group order>'  
# IV_GROUP_TI            = '<TI - Group>'      
IV_GROUP_MST_ID_TI     = '<TI - Group master ID>' 
# IV_SUBGROUP_CURR_ID_TI = '<TI - Subgroup>'       
IV_SUBGROUP_MST_ID_TI  = '<TI - Subgroup master ID>' 
# IV_INDENTS_TI          = '<TI - Indents>'     
IV_STRING_CONSTANT_TI  = '<TI - String constant>' 
# IV_STRING_NUMBER_TI    = '<TI - String #>'       
IV_ENGLISH_STRING_TI   = '<TI - English>'   

IV_GROUP_ORDER_CS      = '<CS - Group order>'  
IV_GROUP_CS            = '<CS - Group>'      
IV_GROUP_MST_ID_CS     = '<CS - Group master ID>' 
IV_SUBGROUP_CURR_ID_CS = '<CS - Subgroup>'       
IV_SUBGROUP_MST_ID_CS  = '<CS - Subgroup master ID>' 
IV_INDENTS_CS          = '<CS - Indents>'     
IV_STRING_CONSTANT_CS  = '<CS - String constant>' 
IV_STRING_NUMBER_CS    = '<CS - String #>'       
IV_ENGLISH_STRING_CS   = '<CS - English>'   
    
# IHT-specific columns

IV_INTERV_IMPACT = '<UH - Intervention with impact?>'      

IV_AVAIL_COMMUNITY_CHAN = '<UH - Available at Community channel?>' 
IV_AVAIL_OUTREACH_CHAN  = '<UH - Available at Outreach channel?>'  
IV_AVAIL_CLINIC_CHAN    = '<UH - Available at Clinic channel?>'   
IV_AVAIL_HOSPITAL_CHAN  = '<UH - Available at Hospital channel?>' 

IV_PIN_STRING_CONSTANT  = '<UH - Population in need string constant>'
IV_PIN_STRING_NUMBER    = '<UH - Population in need string #>'    
IV_TARG_POP_MST_ID      = '<UH - Target population master ID>'   
IV_TARG_POP_MST_ID_LIST = '<UH - Target population master ID, LiST>'  

IV_PERC_DELIV_COMMUNITY_CURR_YEAR        = '<UH - Percent delivered through Community, current year>'   
IV_PERC_DELIV_COMMUNITY_TARG_YEAR        = '<UH - Percent delivered through Community, target year>'      
IV_PERC_DELIV_OUTREACH_CURR_YEAR         = '<UH - Percent delivered through Outreach, current year>'   
IV_PERC_DELIV_OUTREACH_TARG_YEAR         = '<UH - Percent delivered through Outreach, target year>'    
IV_PERC_DELIV_PREHOSP_EMERG_CURR_YEAR    = '<UH - Percent delivered through Prehospital emergency, current year>'     
IV_PERC_DELIV_PREHOSP_EMERG_TARG_YEAR    = '<UH - Percent delivered through Prehospital emergency, target year>'        
IV_PERC_DELIV_GEN_OUTPAT_SERV_CURR_YEAR  = '<UH - Percent delivered through General outpatient services, current year>'  
IV_PERC_DELIV_GEN_OUTPAT_SERV_TARG_YEAR  = '<UH - Percent delivered through General outpatient services, target year>'         
IV_PERC_DELIV_FIRST_REF_CURR_YEAR        = '<UH - Percent delivered through First referral, current year>'      
IV_PERC_DELIV_FIRST_REF_TARG_YEAR        = '<UH - Percent delivered through First referral, target year>'     
IV_PERC_DELIV_SEC_REF_ABOVE_CURR_YEAR    = '<UH - Percent delivered through Second referral and above, current year>'      
IV_PERC_DELIV_SEC_REF_ABOVE_TARG_YEAR    = '<UH - Percent delivered through Second referral and above, target year>'  
IV_PERC_DELIV_WASH_CURR_YEAR             = '<UH - Percent delivered through WASH, current year>'            
IV_PERC_DELIV_WASH_TARG_YEAR             = '<UH - Percent delivered through WASH, target year>'              
IV_PERC_DELIV_OTHER_NON_HEALTH_CURR_YEAR = '<UH - Percent delivered through Other non-health, current year>' 
IV_PERC_DELIV_OTHER_NON_HEALTH_TARG_YEAR = '<UH - Percent delivered through Other non-health, target year>'
IV_PERC_DELIV_PRIVATE_SECTOR_CURR_YEAR   = '<UH - Percent delivered through Private sector, current year>' 
IV_PERC_DELIV_PRIVATE_SECTOR_TARG_YEAR   = '<UH - Percent delivered through Private sector, target year>'  

IV_CS_IMPACT_IH                     = '<UH - CS Impact>'
IV_AM_IMPACT_IH                     = '<UH - AM Impact>'
IV_RN_IMPACT_IH                     = '<UH - RN Impact>'
IV_FP_IMPACT_IH                     = '<UH - FP Impact>'
IV_TB_IMPACT_IH                     = '<UH - TB Impact>'
IV_MA_IMPACT_IH                     = '<UH - MA Impact>'
IV_NC_IMPACT_IH                     = '<UH - NC Impact>'
IV_IS_PROG_MGMT_IH                  = '<UH - IS Programme Management>'
IV_RN_PC_HIV_IH                     = '<UH - RN PC HIV>'
IV_RN_STIS_IH                       = '<UH - RN STIs>'
IV_NOT_DELIV_MAIN_CHANNELS_IH       = '<UH - Not Delivered in Main Delivery Channels>'
IV_BREAST_CANCER_TREATMENT_IH       = '<UH - Breast Cancer Treatment>'
IV_TB_NOTIF_TREATMENT_IH            = '<UH - TB Notification and Treatment>'
IV_CS_CHILDBIRTH_IH                 = '<UH - CS Childbirth>'
IV_PREG_WOMEN_MALARIA_IH            = '<UH - Pregnant Women Malaria>'
IV_ITN_IRS_MALARIA_IH               = '<UH - ITM and IRS Malaria>'
IV_MALARIA_PSEUDO_IMPACT_IH         = '<UH - Malaria Pseudo Impact>'
IV_ART_SET_1_IH                     = '<UH - ART Set 1>'
IV_ART_SET_2_IH                     = '<UH - ART Set 2>'
IV_PC_MNS_IH                        = '<UH - NC PC MNS>'
IV_NC_CANCER_IH                     = '<UH - NC Cancer>'
IV_BASIC_INTENS_DEPRESSION_IH       = '<UH - Basic Intensive Depression>'
IV_BASIC_INTENS_PSYCHOSIS_IH        = '<UH - Basic Intensive Psychosis>'
IV_BASIC_INTENS_BIPOLAR_IH          = '<UH - Basic Intensive Bipolar>'
IV_BASIC_INTENS_DEV_DISORDERS_IH    = '<UH - Basic Intensive Developmental Disorders>'
IV_BASIC_INTENS_BEHAV_DISORDERS_IH  = '<UH - Basic Intensive Behavioral Disorders>'
IV_PENTA_VACC_IH                    = '<UH - Pentavalent Vaccine>'
IV_CS_STANDARD_VACC_IH              = '<UH - CS Standard Vaccines>'
IV_CS_DET_VACC_IH                   = '<UH - CS Detailed Vaccines>'
IV_DEV_MODE_VACC_IH                 = '<UH - CS Developer Mode Vaccines>'
IV_BREAST_CANCER_TR_PINS_IH         = '<UH - Breast Cancer Treatment PINs>'
IV_WORKERS_HEALTH_IH                = '<UH - Workers Health>'
IV_ADOLESCENT_HEALTH_IH             = '<UH - Adolescent Health>'
IV_TB_DAT_DISAG_IH                  = '<UH - TB DAT Disaggregate>'
IV_TB_DAT_AG_IH                     = '<UH - TB DAT Aggregate>'
IV_PHC_IH                           = '<UH - PHC>'
IV_ALT_COVERAGE_IH                  = '<UH - Alternative Coverage>'
IV_HIDE_CONFIG_IH                   = '<UH - Hide Configuration>'
IV_HIDE_INPUT_OUTPUT_DATA_IH        = '<UH - Hide Input and Output Data>'
IV_HIDE_PINS_IH                     = '<UH - Hide PINs>'
IV_HIDE_TARG_POPS_IH                = '<UH - Hide Target Populations>'
IV_HIDE_COVERAGES_IH                = '<UH - Hide Coverages>'
IV_LOCK_COVERAGES_IH                = '<UH - Lock Coverages>'
IV_HIDE_TREAT_INPUTS_IH             = '<UH - Hide Treatment Inputs>'
IV_HIDE_DELIV_CHANNELS_IH           = '<UH - Hide Delivery Channels>'
IV_HIDE_RESULTS_IH                  = '<UH - Hide Results>'
IV_CALC_PINS_ALL_YEARS_LOCKED_IH    = '<UH - Calculate PINs for All Years, Locked>'
IV_CALC_PINS_ALL_YEARS_LOCKED_LIST  = '<UH - Calculate PINs for All Years, Locked, LiST>'
IV_CALC_PINS_ALL_YEARS_OVERRIDE_IH  = '<UH - Calculate PINs for All Years, Overridable>'
IV_CALC_PINS_TWO_PLUS_LOCKED_IH     = '<UH - Calculate PINs for Year Two Forward>'
IV_MULTIPLY_PINS_100_IH             = '<UH - Multiply PINs by 100>'
IV_PINS_OVER_100_IH                 = '<UH - PINs Over 100>'
IV_MALARIA_PINS_IH                  = '<UH - Malaria PINs>'

# TB-specific columns

# LiST (impact)-specific columns
IV_NUTRITION_CS          = '<CS - Is nutrition>'
IV_BIRTH_OUTCOME_CS      = '<CS - Is birth outcome>'
IV_INTERV_TYPE_CS        = '<CS - Intervention type>'
IV_QUALITY_MODE_CS       = '<CS - Quality mode>'
IV_OTHER_MODULE_CS       = '<CS - Other module intervention>'
IV_M_CS                  = '<CS - M>'
IV_SB_CS                 = '<CS - SB>'
IV_NN_CS                 = '<CS - NN>'
IV_PNN_CS                = '<CS - PNN>'
IV_ADOL_EFF_CS           = '<CS - Adol Eff>'
IV_COVERAGE_CS           = '<CS - Coverage>'
IV_DISPLAY_CS            = '<CS - Display>'
IV_MATERNAL_DISPLAY_CS   = '<CS - Maternal display>'
IV_STILLBIRTH_DISPLAY_CS = '<CS - Stillbirth display>'
IV_ADOLESCENT_DISPLAY_CS = '<CS -Adolescent display>'
IV_VACCINES_BEDNET_CS    = '<CS - Vaccines bednet>'
IV_IS_STUNT_INTERV_CS    = '<CS - Is stunt intervention>'
        
# Besides creating intervention databases to upload, this method will also create files with intervention constants. 
def create_intervention_DB_IV(version = str, mod_ID = int):
    log('Creating Intervention DB, ' + gbc.GB_MOD_STR[mod_ID])

    interventions_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IV) + '\IVModData.xlsx')
    sheet = wb[ivc.IV_INTERVENTION_DB_NAME]
    
    '''
        IMPORTANT: The max intervention master ID is 654. 
        
        New interventions must have a master ID greater than this. Please update this maximum if you add a new intervention.

        Last updated by Jared on 9/1/2023.  Added new food security intervention for LiST on desktop.
    '''

    # first row of intervention data after col descriptions and col tags
    tag_row = 2
    first_data_row = 3
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Columns shared by all modules
    root_constant_col = gbc.GB_NOT_FOUND
    mstID_col = gbc.GB_NOT_FOUND
    modules_col = gbc.GB_NOT_FOUND

    # IHT repeated block

    # group_order_col_IH = gbc.GB_NOT_FOUND 
    #group_col_IH = gbc.GB_NOT_FOUND     
    group_mstID_col_IH = gbc.GB_NOT_FOUND  
    #subgroup_col_IH = gbc.GB_NOT_FOUND   
    subgroup_mstID_col_IH = gbc.GB_NOT_FOUND   
    #indents_col_IH = gbc.GB_NOT_FOUND  
    string_constant_col_IH = gbc.GB_NOT_FOUND
    #string_number_col_IH = gbc.GB_NOT_FOUND   
    english_string_col_IH = gbc.GB_NOT_FOUND

    # TB repeated block

    # group_order_col_TI = gbc.GB_NOT_FOUND 
    #group_col_TI = gbc.GB_NOT_FOUND     
    group_mstID_col_TI = gbc.GB_NOT_FOUND  
    #subgroup_col_TI = gbc.GB_NOT_FOUND   
    subgroup_mstID_col_TI = gbc.GB_NOT_FOUND   
    #indents_col_TI = gbc.GB_NOT_FOUND  
    string_constant_col_TI = gbc.GB_NOT_FOUND
    #string_number_col_TI = gbc.GB_NOT_FOUND   
    english_string_col_TI = gbc.GB_NOT_FOUND

    # LiST repeated block
     
    group_order_col_CS = gbc.GB_NOT_FOUND 
    #group_col_CS = gbc.GB_NOT_FOUND     
    group_mstID_col_CS = gbc.GB_NOT_FOUND  
    #subgroup_col_CS = gbc.GB_NOT_FOUND   
    subgroup_mstID_col_CS = gbc.GB_NOT_FOUND   
    #indents_col_CS = gbc.GB_NOT_FOUND  
    string_constant_col_CS = gbc.GB_NOT_FOUND
    #string_number_col_CS = gbc.GB_NOT_FOUND   
    english_string_col_CS = gbc.GB_NOT_FOUND

    # IHT-specific columns

    interv_impact_col = gbc.GB_NOT_FOUND    

    #avail_community_chan_col = gbc.GB_NOT_FOUND

    PIN_string_constant_col = gbc.GB_NOT_FOUND
    #PIN_string_number_col = gbc.GB_NOT_FOUND  
    targ_pop_mstID_col = gbc.GB_NOT_FOUND    
    targ_pop_mstID_LIST_col = gbc.GB_NOT_FOUND    

    perc_deliv_community_curr_year_col = gbc.GB_NOT_FOUND  

    CS_impact_col_IH = gbc.GB_NOT_FOUND
    AM_impact_col_IH = gbc.GB_NOT_FOUND
    RN_impact_col_IH = gbc.GB_NOT_FOUND
    FP_impact_col_IH = gbc.GB_NOT_FOUND
    TB_impact_col_IH = gbc.GB_NOT_FOUND
    MA_impact_col_IH = gbc.GB_NOT_FOUND
    NC_impact_col_IH = gbc.GB_NOT_FOUND
    IS_prog_mgmt_col_IH = gbc.GB_NOT_FOUND
    RN_PC_HIV_col_IH = gbc.GB_NOT_FOUND
    RN_STIs_col_IH = gbc.GB_NOT_FOUND
    not_deliv_main_channels_col_IH = gbc.GB_NOT_FOUND
    breast_cancer_treatment_col_IH = gbc.GB_NOT_FOUND
    TB_notif_treatment_col_IH = gbc.GB_NOT_FOUND
    CS_childbirth_col_IH = gbc.GB_NOT_FOUND
    preg_women_malaria_col_IH = gbc.GB_NOT_FOUND
    itn_irs_malaria_col_IH = gbc.GB_NOT_FOUND
    malaria_pseudo_impact_col_IH = gbc.GB_NOT_FOUND
    art_set_1_col_IH = gbc.GB_NOT_FOUND
    pc_mns_col_IH = gbc.GB_NOT_FOUND
    nc_cancer_col_IH = gbc.GB_NOT_FOUND
    basic_intens_depression_col_IH = gbc.GB_NOT_FOUND
    basic_intens_psychosis_col_IH = gbc.GB_NOT_FOUND
    basic_intens_bipolar_col_IH = gbc.GB_NOT_FOUND
    basic_intens_dev_disorders_col_IH = gbc.GB_NOT_FOUND
    basic_intens_behav_disorders_col_IH = gbc.GB_NOT_FOUND
    penta_vacc_col_IH = gbc.GB_NOT_FOUND
    cs_standard_vacc_col_IH = gbc.GB_NOT_FOUND
    cs_det_vacc_col_IH = gbc.GB_NOT_FOUND
    dev_mode_vacc_col_IH = gbc.GB_NOT_FOUND
    breast_cancer_tr_pins_col_IH = gbc.GB_NOT_FOUND
    workers_health_col_IH = gbc.GB_NOT_FOUND
    adolescent_health_col_IH = gbc.GB_NOT_FOUND
    tb_dat_disag_col_IH = gbc.GB_NOT_FOUND
    tb_dat_ag_col_IH = gbc.GB_NOT_FOUND
    phc_col_IH = gbc.GB_NOT_FOUND
    alt_coverage_col_IH = gbc.GB_NOT_FOUND
    hide_config_col_IH = gbc.GB_NOT_FOUND
    hide_input_output_data_col_IH = gbc.GB_NOT_FOUND
    hide_pins_col_IH = gbc.GB_NOT_FOUND
    hide_targ_pops_col_IH = gbc.GB_NOT_FOUND
    hide_coverages_col_IH = gbc.GB_NOT_FOUND
    lock_coverages_col_IH = gbc.GB_NOT_FOUND
    hide_treat_inputs_col_IH = gbc.GB_NOT_FOUND
    hide_deliv_channels_col_IH = gbc.GB_NOT_FOUND
    hide_results_col_IH = gbc.GB_NOT_FOUND
    calc_pins_all_years_locked_col_IH = gbc.GB_NOT_FOUND
    calc_pins_all_years_locked_col_LiST = gbc.GB_NOT_FOUND
    calc_pins_all_years_override_col_IH = gbc.GB_NOT_FOUND
    calc_pins_two_plus_locked_col_IH = gbc.GB_NOT_FOUND
    multiply_pins_100_col_IH = gbc.GB_NOT_FOUND
    pins_over_100_col_IH = gbc.GB_NOT_FOUND
    malaria_pins_col_IH = gbc.GB_NOT_FOUND
    
    # LiST (impact)-specific columns

    nutrition_col_CS = gbc.GB_NOT_FOUND
    birth_outcome_col_CS = gbc.GB_NOT_FOUND
    interv_type_col_CS = gbc.GB_NOT_FOUND
    quality_mode_col_CS = gbc.GB_NOT_FOUND
    other_module_col_CS = gbc.GB_NOT_FOUND
    m_col_CS = gbc.GB_NOT_FOUND
    sb_col_CS = gbc.GB_NOT_FOUND
    nn_col_CS = gbc.GB_NOT_FOUND
    pnn_col_CS = gbc.GB_NOT_FOUND
    adol_eff_col_CS = gbc.GB_NOT_FOUND
    coverage_col_CS = gbc.GB_NOT_FOUND
    display_col_CS = gbc.GB_NOT_FOUND
    maternal_display_col_CS = gbc.GB_NOT_FOUND
    stillbirth_display_col_CS = gbc.GB_NOT_FOUND
    adolescent_display_col_CS = gbc.GB_NOT_FOUND
    vaccines_bednet_col_CS = gbc.GB_NOT_FOUND
    is_stunt_interv_col_CS = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # Columns shared by all modules

        if tag == IV_ROOT_CONSTANT:
            root_constant_col = c

        elif tag == IV_MST_ID:
            mstID_col = c

        elif tag == IV_MODULES:
            modules_col = c

        # elif tag == IV_MODULES_ENGLISH_KEY :
        #    modules_English_col = c

        # IHT repeated block

        # elif tag == IV_GROUP_ORDER_IH:
        #     group_order_col_IH = c

        #elif tag == IV_GROUP_IH:
        #    group_col_IH = c

        elif tag == IV_GROUP_MST_ID_IH:
            group_mstID_col_IH = c  

        #elif tag == IV_SUBGROUP_IH:
        #    subgroup_col_IH = c

        elif tag == IV_SUBGROUP_MST_ID_IH:
            subgroup_mstID_col_IH = c  

        #elif tag == IV_INDENTS_IH:
        #    indents_col_IH = c

        elif tag == IV_STRING_CONSTANT_IH:
            string_constant_col_IH = c

        # elif tag == IV_STRING_NUMBER_IH:
        #     string_number_col_IH = c

        elif tag == IV_ENGLISH_STRING_IH:
            english_string_col_IH = c

        # TB repeated block

        # elif tag == IV_GROUP_ORDER_TI:
        #     group_order_col_TI = c

        #elif tag == IV_GROUP_TI:
        #    group_col_TI = c

        elif tag == IV_GROUP_MST_ID_TI:
            group_mstID_col_TI = c  

        #elif tag == IV_SUBGROUP_TI:
        #    subgroup_col_TI = c

        elif tag == IV_SUBGROUP_MST_ID_TI:
            subgroup_mstID_col_TI = c  

        #elif tag == IV_INDENTS_TI:
        #    indents_col_TI = c

        elif tag == IV_STRING_CONSTANT_TI:
            string_constant_col_TI = c

        # elif tag == IV_STRING_NUMBER_TI:
        #     string_number_col_TI = c

        elif tag == IV_ENGLISH_STRING_TI:
            english_string_col_TI = c

        # LiST repeated block

        elif tag == IV_GROUP_ORDER_CS:
            group_order_col_CS = c

        #elif tag == IV_GROUP_CS:
        #    group_col_CS = c

        elif tag == IV_GROUP_MST_ID_CS:
            group_mstID_col_CS = c  

        #elif tag == IV_SUBGROUP_CS:
        #    subgroup_col_CS = c

        elif tag == IV_SUBGROUP_MST_ID_CS:
            subgroup_mstID_col_CS = c  

        #elif tag == IV_INDENTS_CS:
        #    indents_col_CS = c

        elif tag == IV_STRING_CONSTANT_CS:
            string_constant_col_CS = c

        # elif tag == IV_STRING_NUMBER_CS:
        #     string_number_col_CS = c

        elif tag == IV_ENGLISH_STRING_CS:
            english_string_col_CS = c

        # IHT-specific columns

        elif tag == IV_INTERV_IMPACT:
            interv_impact_col = c

        #elif tag == IV_AVAIL_COMMUNITY_CHAN:
        #    avail_community_chan_col = c

        elif tag == IV_PIN_STRING_CONSTANT:
            PIN_string_constant_col = c

        #elif tag == IV_PIN_STRING_NUMBER:
        #    PIN_string_number_col = c

        elif tag == IV_TARG_POP_MST_ID:
            targ_pop_mstID_col = c

        elif tag == IV_TARG_POP_MST_ID_LIST:
            targ_pop_mstID_LIST_col = c    

        elif tag == IV_PERC_DELIV_COMMUNITY_CURR_YEAR:
            perc_deliv_community_curr_year_col = c

        elif tag == IV_CS_IMPACT_IH:
            CS_impact_col_IH = c

        elif tag == IV_AM_IMPACT_IH:
            AM_impact_col_IH = c

        elif tag == IV_RN_IMPACT_IH:    
            RN_impact_col_IH = c

        elif tag == IV_FP_IMPACT_IH:    
            FP_impact_col_IH = c

        elif tag == IV_TB_IMPACT_IH:    
            TB_impact_col_IH = c

        elif tag == IV_MA_IMPACT_IH:    
            MA_impact_col_IH = c

        elif tag == IV_NC_IMPACT_IH:    
            NC_impact_col_IH = c

        elif tag == IV_IS_PROG_MGMT_IH:    
            IS_prog_mgmt_col_IH = c

        elif tag == IV_RN_PC_HIV_IH:
            RN_PC_HIV_col_IH = c

        elif tag == IV_RN_STIS_IH:    
            RN_STIs_col_IH = c

        elif tag == IV_NOT_DELIV_MAIN_CHANNELS_IH:    
            not_deliv_main_channels_col_IH = c

        elif tag == IV_BREAST_CANCER_TREATMENT_IH:    
            breast_cancer_treatment_col_IH = c

        elif tag == IV_TB_NOTIF_TREATMENT_IH:    
            TB_notif_treatment_col_IH = c

        elif tag == IV_CS_CHILDBIRTH_IH:    
            CS_childbirth_col_IH = c

        elif tag == IV_PREG_WOMEN_MALARIA_IH:    
            preg_women_malaria_col_IH = c

        elif tag == IV_ITN_IRS_MALARIA_IH:    
            itn_irs_malaria_col_IH = c

        elif tag == IV_MALARIA_PSEUDO_IMPACT_IH:    
            malaria_pseudo_impact_col_IH = c

        elif tag == IV_ART_SET_1_IH:    
            art_set_1_col_IH = c

        elif tag == IV_ART_SET_2_IH:    
            art_set_2_col_IH = c

        elif tag == IV_PC_MNS_IH:    
            pc_mns_col_IH = c

        elif tag == IV_NC_CANCER_IH:    
            nc_cancer_col_IH = c

        elif tag == IV_BASIC_INTENS_DEPRESSION_IH:    
            basic_intens_depression_col_IH = c

        elif tag == IV_BASIC_INTENS_PSYCHOSIS_IH:    
            basic_intens_psychosis_col_IH = c

        elif tag == IV_BASIC_INTENS_BIPOLAR_IH:    
            basic_intens_bipolar_col_IH = c

        elif tag == IV_BASIC_INTENS_DEV_DISORDERS_IH:    
            basic_intens_dev_disorders_col_IH = c

        elif tag == IV_BASIC_INTENS_BEHAV_DISORDERS_IH:    
            basic_intens_behav_disorders_col_IH = c

        elif tag == IV_PENTA_VACC_IH:    
            penta_vacc_col_IH = c

        elif tag == IV_CS_STANDARD_VACC_IH:    
            cs_standard_vacc_col_IH = c

        elif tag == IV_CS_DET_VACC_IH:    
            cs_det_vacc_col_IH = c

        elif tag == IV_DEV_MODE_VACC_IH:    
            dev_mode_vacc_col_IH = c

        elif tag == IV_BREAST_CANCER_TR_PINS_IH:
            breast_cancer_tr_pins_col_IH = c

        elif tag == IV_WORKERS_HEALTH_IH:    
            workers_health_col_IH = c

        elif tag == IV_ADOLESCENT_HEALTH_IH:    
            adolescent_health_col_IH = c

        elif tag == IV_TB_DAT_DISAG_IH:    
            tb_dat_disag_col_IH = c

        elif tag == IV_TB_DAT_AG_IH:    
            tb_dat_ag_col_IH = c

        elif tag == IV_PHC_IH:    
            phc_col_IH = c

        elif tag == IV_ALT_COVERAGE_IH:    
            alt_coverage_col_IH = c

        elif tag == IV_HIDE_CONFIG_IH:    
            hide_config_col_IH = c

        elif tag == IV_HIDE_INPUT_OUTPUT_DATA_IH:    
            hide_input_output_data_col_IH = c

        elif tag == IV_HIDE_PINS_IH:    
            hide_pins_col_IH = c

        elif tag == IV_HIDE_TARG_POPS_IH:    
            hide_targ_pops_col_IH = c

        elif tag == IV_HIDE_COVERAGES_IH:    
            hide_coverages_col_IH = c

        elif tag == IV_LOCK_COVERAGES_IH:
            lock_coverages_col_IH = c

        elif tag == IV_HIDE_TREAT_INPUTS_IH:    
            hide_treat_inputs_col_IH = c

        elif tag == IV_HIDE_DELIV_CHANNELS_IH:    
            hide_deliv_channels_col_IH = c

        elif tag == IV_HIDE_RESULTS_IH:    
            hide_results_col_IH = c

        elif tag == IV_CALC_PINS_ALL_YEARS_LOCKED_IH:    
            calc_pins_all_years_locked_col_IH = c

        elif tag == IV_CALC_PINS_ALL_YEARS_LOCKED_LIST:    
            calc_pins_all_years_locked_col_LiST = c

        elif tag == IV_CALC_PINS_ALL_YEARS_OVERRIDE_IH:    
            calc_pins_all_years_override_col_IH = c

        elif tag == IV_CALC_PINS_TWO_PLUS_LOCKED_IH:   
            calc_pins_two_plus_locked_col_IH = c

        elif tag == IV_MULTIPLY_PINS_100_IH:    
            multiply_pins_100_col_IH = c

        elif tag == IV_PINS_OVER_100_IH:    
            pins_over_100_col_IH = c

        elif tag == IV_MALARIA_PINS_IH:    
            malaria_pins_col_IH = c

        # LiST-specific columns

        elif tag == IV_NUTRITION_CS:    
            nutrition_col_CS = c

        elif tag == IV_BIRTH_OUTCOME_CS:    
            birth_outcome_col_CS = c

        elif tag == IV_INTERV_TYPE_CS:    
            interv_type_col_CS = c

        elif tag == IV_QUALITY_MODE_CS:    
            quality_mode_col_CS = c

        elif tag == IV_OTHER_MODULE_CS:    
            other_module_col_CS = c

        elif tag == IV_M_CS:    
            m_col_CS = c

        elif tag == IV_SB_CS:    
            sb_col_CS = c

        elif tag == IV_NN_CS:    
            nn_col_CS = c

        elif tag == IV_PNN_CS:    
            pnn_col_CS = c

        elif tag == IV_ADOL_EFF_CS:    
            adol_eff_col_CS = c

        elif tag == IV_COVERAGE_CS:    
            coverage_col_CS = c

        elif tag == IV_DISPLAY_CS:    
            display_col_CS = c

        elif tag == IV_MATERNAL_DISPLAY_CS:    
            maternal_display_col_CS = c

        elif tag == IV_STILLBIRTH_DISPLAY_CS:    
            stillbirth_display_col_CS = c

        elif tag == IV_ADOLESCENT_DISPLAY_CS:    
            adolescent_display_col_CS = c

        elif tag == IV_VACCINES_BEDNET_CS:    
            vaccines_bednet_col_CS = c

        elif tag == IV_IS_STUNT_INTERV_CS:    
            is_stunt_interv_col_CS = c

    num_deliv_chan_all = ivcl.IV_IH_NUM_DEFAULT_DELIV_CHANS_COSTED + ivcl.IV_IC_NUM_DELIV_CHANS

    # IC delivery channels come after the IH ones
    dc1_IC = ivcl.IV_IH_NUM_DEFAULT_DELIV_CHANS_COSTED + 1
    dc2_IC = dc1_IC + ivcl.IV_IC_NUM_DELIV_CHANS - 1

    # Get each row from the sheet and create a dictionary for each each intervention
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        # Need to use str for single modules like 15 so they are not seen as integers when using find. 
        modules_str = str(row[modules_col - 1])

        if modules_str.find(str(mod_ID)) != gbc.GB_NOT_FOUND:
            # Each intervention can exist for one or more modules. Save the position of the module in the list of specified modules.
            modules_list = modules_str.split(',')
            for m, string in enumerate(modules_list):
                modules_list[m] = string.strip()
            module_idx = modules_list.index(str(mod_ID))

            # All modules

            interv_dict = {}
            interv_dict[ivdbk.IV_ROOT_CONSTANT_KEY_IDB] = row[root_constant_col - 1]
            interv_dict[ivdbk.IV_MST_ID_KEY_IDB] = row[mstID_col - 1]
            #interv_dict[ivdbk.IV_MODULES_KEY] = modules_str

            # Only used for IHT

            if mod_ID == gbc.GB_IH:

                # interv_dict[ivdbk.IV_GROUP_ORDER_KEY_IDB] = row[group_order_col_IH - 1]
                #interv_dict[ivdbk.IV_GROUP_KEY_IH] = row[group_col_IH - 1]                     not needed?
                interv_dict[ivdbk.IV_GROUP_MST_ID_KEY_IDB] = row[group_mstID_col_IH - 1]
                #interv_dict[ivdbk.IV_SUBGROUP_KEY_IH] = row[subgroup_col_IH - 1]                not needed?
                interv_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_IDB] = row[subgroup_mstID_col_IH - 1]
                #interv_dict[ivdbk.IV_INDENTS_KEY_IH] = row[indents_col_IH - 1]                  always 2?
                interv_dict[ivdbk.IV_STRING_CONSTANT_KEY_IDB] = row[string_constant_col_IH - 1]
                #interv_dict[ivdbk.IV_STRING_NUMBER_KEY_IH] = row[string_number_col_IH - 1]
                interv_dict[ivdbk.IV_ENGLISH_STRING_KEY_IDB] = row[english_string_col_IH - 1]

                # Currently shares the same list of targ pops as TI
                interv_dict[ivdbk.IV_TARG_POP_MST_ID_KEY_IDB] = row[targ_pop_mstID_col - 1]

            # only used for TB

            elif mod_ID == gbc.GB_TI:
            
                # interv_dict[ivdbk.IV_GROUP_ORDER_KEY_IDB] = row[group_order_col_TI - 1]
                #interv_dict[ivdbk.IV_GROUP_KEY_IH] = row[group_col_TI - 1]                     not needed?
                interv_dict[ivdbk.IV_GROUP_MST_ID_KEY_IDB] = row[group_mstID_col_TI - 1]
                #interv_dict[ivdbk.IV_SUBGROUP_KEY_IH] = row[subgroup_col_TI - 1]                not needed?
                interv_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_IDB] = row[subgroup_mstID_col_TI - 1]
                #interv_dict[ivdbk.IV_INDENTS_KEY_IH] = row[indents_col_TI - 1]                  always 2?
                interv_dict[ivdbk.IV_STRING_CONSTANT_KEY_IDB] = row[string_constant_col_TI - 1]
                #interv_dict[ivdbk.IV_STRING_NUMBER_KEY_IH] = row[string_number_col_TI - 1]
                interv_dict[ivdbk.IV_ENGLISH_STRING_KEY_IDB] = row[english_string_col_TI - 1]

                # Currently shares the same list of targ pops as IH
                interv_dict[ivdbk.IV_TARG_POP_MST_ID_KEY_IDB] = row[targ_pop_mstID_col - 1]
            
            # Only used for LiST

            elif mod_ID == gbc.GB_CS:
            
                interv_dict[ivdbk.IV_GROUP_ORDER_KEY_IDB] = row[group_order_col_CS - 1]
                #interv_dict[ivdbk.IV_GROUP_KEY_IH] = row[group_col_CS - 1]                     not needed?
                interv_dict[ivdbk.IV_GROUP_MST_ID_KEY_IDB] = row[group_mstID_col_CS - 1]
                #interv_dict[ivdbk.IV_SUBGROUP_KEY_IH] = row[subgroup_col_CS - 1]                not needed?
                interv_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_IDB] = row[subgroup_mstID_col_CS - 1]
                #interv_dict[ivdbk.IV_INDENTS_KEY_IH] = row[indents_col_CS - 1]                  always 2?
                interv_dict[ivdbk.IV_STRING_CONSTANT_KEY_IDB] = row[string_constant_col_CS - 1]
                #interv_dict[ivdbk.IV_STRING_NUMBER_KEY_IH] = row[string_number_col_CS - 1]
                interv_dict[ivdbk.IV_ENGLISH_STRING_KEY_IDB] = row[english_string_col_CS - 1]

                # Has its own column of targ pops, since it shares some of the same interventions as TB,
                # and TB wants to use different targ pops for some.
                interv_dict[ivdbk.IV_TARG_POP_MST_ID_KEY_IDB] = row[targ_pop_mstID_LIST_col - 1]
                
                interv_dict[ivdbk.IV_NUTRITION_CS_KEY_IDB] = row[nutrition_col_CS - 1]
                interv_dict[ivdbk.IV_BIRTH_OUTCOME_CS_KEY_IDB] = row[birth_outcome_col_CS - 1]
                interv_dict[ivdbk.IV_INTERV_TYPE_CS_KEY_IDB] = row[interv_type_col_CS - 1]
                interv_dict[ivdbk.IV_QUALITY_MODE_CS_KEY_IDB] = row[quality_mode_col_CS - 1]
                interv_dict[ivdbk.IV_OTHER_MODULE_CS_KEY_IDB] = row[other_module_col_CS - 1]
                interv_dict[ivdbk.IV_M_CS_KEY_IDB] = row[m_col_CS - 1]
                interv_dict[ivdbk.IV_SB_CS_KEY_IDB] = row[sb_col_CS - 1]
                interv_dict[ivdbk.IV_NN_CS_KEY_IDB] = row[nn_col_CS - 1]
                interv_dict[ivdbk.IV_PNN_CS_KEY_IDB] = row[pnn_col_CS - 1]
                interv_dict[ivdbk.IV_ADOL_EFF_CS_KEY_IDB] = row[adol_eff_col_CS - 1]
                interv_dict[ivdbk.IV_COVERAGE_CS_KEY_IDB] = row[coverage_col_CS - 1]
                interv_dict[ivdbk.IV_DISPLAY_CS_KEY_IDB] = row[display_col_CS - 1]
                interv_dict[ivdbk.IV_MATERNAL_DISPLAY_CS_KEY_IDB] = row[maternal_display_col_CS - 1]
                interv_dict[ivdbk.IV_STILLBIRTH_DISPLAY_CS_KEY_IDB] = row[stillbirth_display_col_CS - 1]
                interv_dict[ivdbk.IV_ADOLESCENT_DISPLAY_CS_KEY_IDB] = row[adolescent_display_col_CS - 1]
                interv_dict[ivdbk.IV_VACCINES_BEDNET_CS_KEY_IDB] = row[vaccines_bednet_col_CS - 1]
                interv_dict[ivdbk.IV_IS_STUNT_INTERV_CS_KEY_IDB] = row[is_stunt_interv_col_CS - 1]
            
            # These are needed for IHT, TB Costing, and LiST Costing:

            interv_dict[ivdbk.IV_INTERV_IMPACT_KEY_IDB] = True if row[interv_impact_col - 1] == 1 else False
            #interv_dict[ivdbk.IV_AVAIL_COMMUNITY_CHAN_KEY] = row[avail_community_chan_col - 1]     not needed?
            interv_dict[ivdbk.IV_PIN_STRING_CONSTANT_KEY_IDB] = row[PIN_string_constant_col - 1]
            #interv_dict[ivdbk.IV_PIN_STRING_NUMBER_KEY] = row[PIN_string_number_col - 1]

            dc_col = perc_deliv_community_curr_year_col

            perc_deliv_list = np.full((num_deliv_chan_all, ivcl.IV_IC_NUM_YEAR_ENDPOINTS), 0.0).tolist()

            # IH delivery channels by year endpoint
            for dc in GBRange(1, ivcl.IV_IH_NUM_DEFAULT_DELIV_CHANS_COSTED):
                for yr in GBRange(1, ivcl.IV_IC_NUM_YEAR_ENDPOINTS):
                    if isinstance(row[dc_col - 1], int) or isinstance(row[dc_col - 1], float):
                        perc_deliv_list[dc - 1][yr - 1] = row[dc_col - 1]
                    else:
                        values_list = row[dc_col - 1].split(',')
                        for m, string in enumerate(values_list):
                            values_list[m] = int(string.strip())

                        perc_deliv_list[dc - 1][yr - 1] = values_list[module_idx]
                    dc_col += 1

            # IC delivery channels by year endpoint
            for dc in GBRange(dc1_IC, dc2_IC):
                for yr in GBRange(1, ivcl.IV_IC_NUM_YEAR_ENDPOINTS):
                    if isinstance(row[dc_col - 1], int) or isinstance(row[dc_col - 1], float):
                        perc_deliv_list[dc - 1][yr - 1] = row[dc_col - 1]
                    else:
                        values_list = row[dc_col - 1].split(',')
                        for m, string in enumerate(values_list):
                            values_list[m] = int(string.strip())

                        perc_deliv_list[dc - 1][yr - 1] = values_list[module_idx]
                    dc_col += 1

            interv_dict[ivdbk.IV_PERC_DELIV_KEY_IDB] = perc_deliv_list

            interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB] = []

            if (CS_impact_col_IH != gbc.GB_NOT_FOUND) and (row[CS_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CS_IMPACT)

            if (AM_impact_col_IH != gbc.GB_NOT_FOUND) and (row[AM_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_AM_IMPACT)

            if (RN_impact_col_IH != gbc.GB_NOT_FOUND) and (row[RN_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_RN_IMPACT)

            if (FP_impact_col_IH != gbc.GB_NOT_FOUND) and (row[FP_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_FP_IMPACT)

            if (TB_impact_col_IH != gbc.GB_NOT_FOUND) and (row[TB_impact_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_TB_IMPACT)

            if (MA_impact_col_IH != gbc.GB_NOT_FOUND) and (row[MA_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_MA_IMPACT)

            if (NC_impact_col_IH != gbc.GB_NOT_FOUND) and (row[NC_impact_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_NC_IMPACT)

            if (IS_prog_mgmt_col_IH != gbc.GB_NOT_FOUND) and (row[IS_prog_mgmt_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_IS_PROG_MGMT)

            if (RN_PC_HIV_col_IH != gbc.GB_NOT_FOUND) and (row[RN_PC_HIV_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_RN_PC_HIV)

            if (RN_STIs_col_IH != gbc.GB_NOT_FOUND) and (row[RN_STIs_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_RN_STIS)

            if (not_deliv_main_channels_col_IH != gbc.GB_NOT_FOUND) and (row[CS_impact_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_NOT_DELIV_MAIN_CHANNELS)

            if (breast_cancer_treatment_col_IH != gbc.GB_NOT_FOUND) and (row[breast_cancer_treatment_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BREAST_CANCER_TREATMENT)

            if (TB_notif_treatment_col_IH != gbc.GB_NOT_FOUND) and (row[TB_notif_treatment_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_TB_NOTIF_TREATMENT)

            if (CS_childbirth_col_IH != gbc.GB_NOT_FOUND) and (row[CS_childbirth_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CS_CHILDBIRTH)

            if (preg_women_malaria_col_IH != gbc.GB_NOT_FOUND) and (row[preg_women_malaria_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_PREG_WOMEN_MALARIA)

            if (itn_irs_malaria_col_IH != gbc.GB_NOT_FOUND) and (row[itn_irs_malaria_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_ITN_IRS_MALARIA)

            if (malaria_pseudo_impact_col_IH != gbc.GB_NOT_FOUND) and (row[malaria_pseudo_impact_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_MALARIA_PSEUDO_IMPACT)

            if (art_set_1_col_IH != gbc.GB_NOT_FOUND) and (row[art_set_1_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_ART_SET_1)

            if (art_set_2_col_IH != gbc.GB_NOT_FOUND) and (row[art_set_2_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_ART_SET_2)

            if (pc_mns_col_IH != gbc.GB_NOT_FOUND) and (row[pc_mns_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_PC_MNS)

            if (nc_cancer_col_IH != gbc.GB_NOT_FOUND) and (row[nc_cancer_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_NC_CANCER)

            if (basic_intens_depression_col_IH != gbc.GB_NOT_FOUND) and (row[basic_intens_depression_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BASIC_INTENS_DEPRESSION)

            if (basic_intens_psychosis_col_IH != gbc.GB_NOT_FOUND) and (row[basic_intens_psychosis_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BASIC_INTENS_PSYCHOSIS)

            if (basic_intens_bipolar_col_IH != gbc.GB_NOT_FOUND) and (row[basic_intens_bipolar_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BASIC_INTENS_BIPOLAR)

            if (basic_intens_dev_disorders_col_IH != gbc.GB_NOT_FOUND) and (row[basic_intens_dev_disorders_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BASIC_INTENS_DEV_DISORDERS)

            if (basic_intens_behav_disorders_col_IH != gbc.GB_NOT_FOUND) and (row[basic_intens_behav_disorders_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BASIC_INTENS_BEHAV_DISORDERS)

            if (penta_vacc_col_IH != gbc.GB_NOT_FOUND) and (row[penta_vacc_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_PENTA_VACC)

            if (cs_standard_vacc_col_IH != gbc.GB_NOT_FOUND) and (row[cs_standard_vacc_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CS_STANDARD_VACC)

            if (cs_det_vacc_col_IH != gbc.GB_NOT_FOUND) and (row[cs_det_vacc_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CS_DET_VACC)

            if (dev_mode_vacc_col_IH != gbc.GB_NOT_FOUND) and (row[dev_mode_vacc_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_DEV_MODE_VACC)

            if (breast_cancer_tr_pins_col_IH != gbc.GB_NOT_FOUND) and (row[breast_cancer_tr_pins_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_BREAST_CANCER_TR_PINS)

            if (workers_health_col_IH != gbc.GB_NOT_FOUND) and (row[workers_health_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_WORKERS_HEALTH)

            if (adolescent_health_col_IH != gbc.GB_NOT_FOUND) and (row[adolescent_health_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_ADOLESCENT_HEALTH)

            if (tb_dat_disag_col_IH != gbc.GB_NOT_FOUND) and (row[tb_dat_disag_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_TB_DAT_DISAG)

            if (tb_dat_ag_col_IH != gbc.GB_NOT_FOUND) and (row[tb_dat_ag_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_TB_DAT_AG)

            if (phc_col_IH != gbc.GB_NOT_FOUND) and (row[phc_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_PHC)

            if (alt_coverage_col_IH != gbc.GB_NOT_FOUND) and (row[alt_coverage_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_ALT_COVERAGE)

            if (hide_config_col_IH != gbc.GB_NOT_FOUND) and (row[hide_config_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_CONFIG)

            if (hide_input_output_data_col_IH != gbc.GB_NOT_FOUND) and (row[hide_input_output_data_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_INPUT_OUTPUT_DATA)

            if (hide_pins_col_IH != gbc.GB_NOT_FOUND) and (row[hide_pins_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_PINS)

            if (hide_targ_pops_col_IH != gbc.GB_NOT_FOUND) and (row[hide_targ_pops_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_TARG_POPS)

            if (hide_coverages_col_IH != gbc.GB_NOT_FOUND) and (row[hide_coverages_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_COVERAGES)

            if (lock_coverages_col_IH != gbc.GB_NOT_FOUND) and (row[lock_coverages_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_LOCK_COVERAGES)

            if (hide_treat_inputs_col_IH != gbc.GB_NOT_FOUND) and (row[hide_treat_inputs_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_TREAT_INPUTS)

            if (hide_deliv_channels_col_IH != gbc.GB_NOT_FOUND) and (row[hide_deliv_channels_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_DELIV_CHANNELS)

            if (hide_results_col_IH != gbc.GB_NOT_FOUND) and (row[hide_results_col_IH - 1] == 1):
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_HIDE_RESULTS)

            # As of 4/4/2024, LiST interventions are either not locked or locked at all years and not overrideable. IHT and TB share the original three options 
            # (locked all years, not overridable; locked all years, overridable; locked year 2+, not overridable)
            if mod_ID == gbc.GB_CS:
                if (calc_pins_all_years_locked_col_LiST != gbc.GB_NOT_FOUND) and (row[calc_pins_all_years_locked_col_LiST - 1] == 1):
                    interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CALC_PINS_ALL_YEARS_LOCKED)
            else:
                if (calc_pins_all_years_locked_col_IH != gbc.GB_NOT_FOUND) and (row[calc_pins_all_years_locked_col_IH - 1] == 1):
                    interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CALC_PINS_ALL_YEARS_LOCKED)

                if (calc_pins_all_years_override_col_IH != gbc.GB_NOT_FOUND) and (row[calc_pins_all_years_override_col_IH - 1] == 1): 
                    interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CALC_PINS_ALL_YEARS_OVERRIDE)

                if (calc_pins_two_plus_locked_col_IH != gbc.GB_NOT_FOUND) and (row[calc_pins_two_plus_locked_col_IH - 1] == 1):
                    interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_CALC_PINS_TWO_PLUS_LOCKED)

            if (multiply_pins_100_col_IH != gbc.GB_NOT_FOUND) and (row[multiply_pins_100_col_IH - 1] == 1):  
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_MULTIPLY_PINS_100)

            if (pins_over_100_col_IH != gbc.GB_NOT_FOUND) and (row[pins_over_100_col_IH - 1] == 1):   
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_PINS_OVER_100)

            if (malaria_pins_col_IH != gbc.GB_NOT_FOUND) and (row[malaria_pins_col_IH - 1] == 1): 
                interv_dict[ivdbk.IV_SPECIAL_CASES_IH_KEY_IDB].append(ivcl.IV_IC_MALARIA_PINS)

            interventions_list.append(interv_dict)

    # Write database out to IV folder regardless of which mode/modules it's for, since it's temporary anyway.

    j = ujson.dumps(interventions_list)

    IV_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IV) + '\\' + gbc.GB_CONTAINERS[mod_ID]
    JSON_file_name = ivc.IV_INTERVENTION_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(IV_JSON_data_path + '\\', exist_ok = True)
    with open(IV_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished Intervention DB, ' + gbc.GB_MOD_STR[mod_ID])

    # Write out a new intervention master ID constants file to the appropriate place depending on mode/module.

    if mod_ID == gbc.GB_IH:
        with open(os.getcwd() + '\SpectrumCommon\Const\IC\ICInterventionIDs.py', 'w') as f:
            for interv_dict in interventions_list:
                f.write('IC_' + interv_dict[ivdbk.IV_ROOT_CONSTANT_KEY_IDB] + '_MST_ID' + ' = \'' + str(interv_dict[ivdbk.IV_MST_ID_KEY_IDB]) + '\'' + '\n')

        log('Updated intervention master IDs in SpectrumCommon --> Const --> IC --> ICInterventionIDs.')

    elif mod_ID == gbc.GB_TI:
        with open(os.getcwd() + '\SpectrumCommon\Const\IC\IC_TI_InterventionIDs.py', 'w') as f:
            for interv_dict in interventions_list:
                f.write('IC_TI_' + interv_dict[ivdbk.IV_ROOT_CONSTANT_KEY_IDB] + '_MST_ID' + ' = \'' + str(interv_dict[ivdbk.IV_MST_ID_KEY_IDB]) + '\'' + '\n')

        log('Updated intervention master IDs in SpectrumCommon --> Const --> IC --> IC_TI_InterventionIDs.')

    elif mod_ID == gbc.GB_CS:
        with open(os.getcwd() + '\SpectrumCommon\Const\CS\CSInterventionIDs.py', 'w') as f:
            for interv_dict in interventions_list:
                f.write('CS_' + interv_dict[ivdbk.IV_ROOT_CONSTANT_KEY_IDB] + '_MST_ID' + ' = \'' + str(interv_dict[ivdbk.IV_MST_ID_KEY_IDB]) + '\'' + '\n')

        log('Updated intervention master IDs in SpectrumCommon --> Const --> CS --> CSInterventionIDs.')

def upload_intervention_DB_IV(version = str, mod_ID = int):
    # File will be uploaded from this location. Same location regardless of mode/module since it's temporary.
    container_name = ''
    destination_database_name = ''
    source_JSON_file_name = ivc.IV_INTERVENTION_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    # Store IHT and TB Costing intervention databases in the IHT container. TB doesn't use the TB interventions and groups from the IV database
    # outside of costing. On the other hand, LiST does, so LiST interventions will go in the LiST container.
    if (mod_ID == gbc.GB_IH) or (mod_ID == gbc.GB_TI):
        container_name = gbc.GB_IH_CORE_CONTAINER
        # Need to differentiate between IHT and TB Costing since they will be uploaded to the same container, so just add the module in the name.
        destination_database_name = ivc.IV_INTERVENTION_DB_NAME + '_' + gbc.GB_MOD_STR[mod_ID] + '_' + version + '.' + gbc.GB_JSON

    elif mod_ID == gbc.GB_CS:
        container_name = gbc.GB_CS_CONTAINER
        destination_database_name = ivc.IV_INTERVENTION_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], container_name, destination_database_name, ddu.get_JSON_data_path(gbc.GB_IV) + '\\' + gbc.GB_CONTAINERS[mod_ID] + '\\' + source_JSON_file_name) 

    log('Uploaded Intervention DB, ' + gbc.GB_MOD_STR[mod_ID])
            
