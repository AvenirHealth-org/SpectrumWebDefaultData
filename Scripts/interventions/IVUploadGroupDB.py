import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file, GB_file_exists
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.IV.IVConst as ivc
import SpectrumCommon.Const.IV.IVDatabaseKeys as ivdbk 

# Column tag 'constants' for identifying columns.

IV_ROOT_CONSTANT_CS    = '<CS - Root constant>'  
IV_GROUP_MST_ID_CS     = '<CS - Group master ID>'   
IV_SUBGROUP_MST_ID_CS  = '<CS - Subgroup master ID>'    
IV_GROUP_CURR_ID_CS    = '<CS - Group>' 
IV_SUBGROUP_CURR_ID_CS = '<CS - Subgroup>'       
IV_ENGLISH_STRING_CS   = '<CS - English>' 
IV_STRING_CONSTANT_CS  = '<CS - String constant>'     
IV_STRING_NUMBER_CS    = '<CS - String #>'  
    
IV_ROOT_CONSTANT_IH    = '<UH - Root constant>'  
IV_GROUP_MST_ID_IH     = '<UH - Group master ID>'
IV_SUBGROUP_MST_ID_IH  = '<UH - Subgroup master ID>'          
# IV_GROUP_CURR_ID_IH    = '<UH - Group>' 
# IV_SUBGROUP_CURR_ID_IH = '<UH - Subgroup>'       
IV_ENGLISH_STRING_IH   = '<UH - English>' 
IV_STRING_CONSTANT_IH  = '<UH - String constant>'     
# IV_STRING_NUMBER_IH    = '<UH - String #>'  

IV_ROOT_CONSTANT_TI    = '<TI - Root constant>'  
IV_GROUP_MST_ID_TI     = '<TI - Group master ID>'     
IV_SUBGROUP_MST_ID_TI  = '<TI - Subgroup master ID>'     
# IV_GROUP_CURR_ID_TI    = '<TI - Group>' 
# IV_SUBGROUP_CURR_ID_TI = '<TI - Subgroup>'       
IV_ENGLISH_STRING_TI   = '<TI - English>' 
IV_STRING_CONSTANT_TI  = '<TI - String constant>'     
# IV_STRING_NUMBER_TI    = '<TI - String #>'  

def create_group_DB_IV(version = str, mod_ID = int):
    # Testing
    # connection = os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV]
    # container = gbc.GB_CS_CONTAINER
    # file_name = destination_database_name = ivc.IV_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    # if GB_file_exists(connection, container, file_name):
    #     log('File exists')
    # else:
    #     log('File does not exist')

    log('Creating Intervention Group DB, ' + gbc.GB_MOD_STR[mod_ID])

    group_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IV) + '\IVModData.xlsx')
    sheet = wb[ivc.IV_GROUP_DB_NAME]
    
    '''
        IMPORTANT: The max group master ID is as follows:
            CS: 33
            IH: 107
            TI: 34
        
        New groups must have a master ID greater than this. Please update this maximum if you add a new group.

        Last updated by Jared on 8/28/2019.  Added three subgroups for the LiST childbirth tab.  
        Last updated by Mark on 2/7/2023.  Completely replaced groups for what is now TB Costing.
    '''

    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    root_constant_col_IH = gbc.GB_NOT_FOUND
    group_mstID_col_IH = gbc.GB_NOT_FOUND
    subgroup_mstID_col_IH = gbc.GB_NOT_FOUND
    string_const_col_IH = gbc.GB_NOT_FOUND
    english_string_col_IH = gbc.GB_NOT_FOUND

    root_constant_col_TI = gbc.GB_NOT_FOUND
    group_mstID_col_TI = gbc.GB_NOT_FOUND
    subgroup_mstID_col_TI = gbc.GB_NOT_FOUND
    string_const_col_TI = gbc.GB_NOT_FOUND
    english_string_col_TI = gbc.GB_NOT_FOUND

    root_constant_col_CS = gbc.GB_NOT_FOUND
    group_mstID_col_CS = gbc.GB_NOT_FOUND
    subgroup_mstID_col_CS = gbc.GB_NOT_FOUND
    string_const_col_CS = gbc.GB_NOT_FOUND
    english_string_col_CS = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # IHT

        if tag == IV_ROOT_CONSTANT_IH:
            root_constant_col_IH = c

        elif tag == IV_GROUP_MST_ID_IH:
            group_mstID_col_IH = c

        elif tag == IV_SUBGROUP_MST_ID_IH:
            subgroup_mstID_col_IH = c

        elif tag == IV_STRING_CONSTANT_IH:
            string_const_col_IH = c

        elif tag == IV_ENGLISH_STRING_IH:
            english_string_col_IH = c

        # TB Costing

        if tag == IV_ROOT_CONSTANT_TI:
            root_constant_col_TI = c

        elif tag == IV_GROUP_MST_ID_TI:
            group_mstID_col_TI = c

        elif tag == IV_SUBGROUP_MST_ID_TI:
            subgroup_mstID_col_TI = c

        elif tag == IV_STRING_CONSTANT_TI:
            string_const_col_TI = c

        elif tag == IV_ENGLISH_STRING_TI:
            english_string_col_TI = c

        # LiST Costing

        if tag == IV_ROOT_CONSTANT_CS:
            root_constant_col_CS = c

        elif tag == IV_GROUP_MST_ID_CS:
            group_mstID_col_CS = c

        elif tag == IV_SUBGROUP_MST_ID_CS:
            subgroup_mstID_col_CS = c

        elif tag == IV_STRING_CONSTANT_CS:
            string_const_col_CS = c

        elif tag == IV_ENGLISH_STRING_CS:
            english_string_col_CS = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        if mod_ID == gbc.GB_IH:

            if row[group_mstID_col_IH - 1] != None:
                group_dict = {}

                group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] = row[root_constant_col_IH - 1]
                group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB] = row[group_mstID_col_IH - 1]
                group_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_GDB] = row[subgroup_mstID_col_IH - 1]
                group_dict[ivdbk.IV_STRING_CONSTANT_KEY_GDB] = row[string_const_col_IH - 1]
                group_dict[ivdbk.IV_ENGLISH_STRING_KEY_GDB] = row[english_string_col_IH - 1]

                group_list.append(group_dict)

        elif mod_ID == gbc.GB_TI:

            if row[group_mstID_col_TI - 1] != None:
                group_dict = {}

                group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] = row[root_constant_col_TI - 1]
                group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB] = row[group_mstID_col_TI - 1]
                group_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_GDB] = row[subgroup_mstID_col_TI - 1]
                group_dict[ivdbk.IV_STRING_CONSTANT_KEY_GDB] = row[string_const_col_TI - 1]
                group_dict[ivdbk.IV_ENGLISH_STRING_KEY_GDB] = row[english_string_col_TI - 1]

                group_list.append(group_dict)

        elif mod_ID == gbc.GB_CS:
            
            if row[group_mstID_col_CS - 1] != None:
                group_dict = {}

                group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] = row[root_constant_col_CS - 1]
                group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB] = row[group_mstID_col_CS - 1]
                group_dict[ivdbk.IV_SUBGROUP_MST_ID_KEY_GDB] = row[subgroup_mstID_col_CS - 1]
                group_dict[ivdbk.IV_STRING_CONSTANT_KEY_GDB] = row[string_const_col_CS - 1]
                group_dict[ivdbk.IV_ENGLISH_STRING_KEY_GDB] = row[english_string_col_CS - 1]

                group_list.append(group_dict)

    j = ujson.dumps(group_list)

    IV_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IV)
    JSON_file_name = ivc.IV_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(IV_JSON_data_path + '\\', exist_ok = True)
    with open(IV_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished Intervention Group DB, ' + gbc.GB_MOD_STR[mod_ID])

    # Write out a new group master ID constants file to the appropriate place depending on mode/module.

    if mod_ID == gbc.GB_IH:
        with open(os.getcwd() + '\SpectrumCommon\Const\IH\IHProgAreaIDs.py', 'w') as f:
            group_mst_IDs = []
            for group_dict in group_list:
                if group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] != None:
                    group_mst_ID = 'IH_' + group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] + '_MST_ID'
                    f.write(group_mst_ID + ' = \'' + str(group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB]) + '\'' + '\n')
                    group_mst_IDs.append(group_mst_ID)
            group_mst_IDs_str = str(group_mst_IDs)
            group_mst_IDs_str = group_mst_IDs_str.replace('\'', '')
            f.write('IH_PROG_AREA_MST_IDS = ' + group_mst_IDs_str)

        log('Updated programme area master IDs in SpectrumCommon --> Const --> IH --> IHProgAreaIDs.')

    elif mod_ID == gbc.GB_TI:
        with open(os.getcwd() + '\SpectrumCommon\Const\IH\IH_TI_ProgAreaIDs.py', 'w') as f:
            group_mst_IDs = []
            for group_dict in group_list:
                if group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] != None:
                    group_mst_ID = 'IH_TI_' + group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] + '_MST_ID'
                    f.write(group_mst_ID + ' = \'' + str(group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB]) + '\'' + '\n')
                    group_mst_IDs.append(group_mst_ID)
            group_mst_IDs_str = str(group_mst_IDs)
            group_mst_IDs_str = group_mst_IDs_str.replace('\'', '')
            f.write('IH_TI_PROG_AREA_MST_IDS = ' + group_mst_IDs_str)

        log('Updated intervention master IDs in SpectrumCommon --> Const --> IH --> IH_TI_ProgAreaIDs.')

    elif mod_ID == gbc.GB_CS:
        with open(os.getcwd() + '\SpectrumCommon\Const\CS\CSGroupIDs.py', 'w') as f:
            group_mst_IDs = []
            for group_dict in group_list:
                if group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] != None:
                    group_mst_ID = 'CS_' + group_dict[ivdbk.IV_ROOT_CONSTANT_KEY_GDB] + '_MST_ID'
                    f.write(group_mst_ID + ' = \'' + str(group_dict[ivdbk.IV_GROUP_MST_ID_KEY_GDB]) + '\'' + '\n')
                    group_mst_IDs.append(group_mst_ID)
            group_mst_IDs_str = str(group_mst_IDs)
            group_mst_IDs_str = group_mst_IDs_str.replace('\'', '')
            f.write('CS_GROUP_MST_IDS = ' + group_mst_IDs_str)

        log('Updated intervention master IDs in SpectrumCommon --> Const --> CS --> CSGroupIDs.')

def upload_group_DB_IV(version = str, mod_ID = int):
    # File will be uploaded from this location. Same location regardless of mode/module since it's temporary.
    container_name = ''
    destination_database_name = ''
    source_JSON_file_name = ivc.IV_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    # Store IHT and TB Costing group databases in the IHT container. TB doesn't use the TB interventions and groups from the IV database
    # outside of costing. On the other hand, LiST does, so LiST groups will go in the LiST container.
    if (mod_ID == gbc.GB_IH) or (mod_ID == gbc.GB_TI):
        container_name = gbc.GB_IH_CORE_CONTAINER
        # Need to differentiate between IHT and TB Costing since they will be uploaded to the same container, so just add the module in the name.
        destination_database_name = ivc.IV_GROUP_DB_NAME + '_' + gbc.GB_MOD_STR[mod_ID] + '_' + version + '.' + gbc.GB_JSON

    elif mod_ID == gbc.GB_CS:
        container_name = gbc.GB_CS_CONTAINER
        destination_database_name = ivc.IV_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], container_name, destination_database_name, ddu.get_JSON_data_path(gbc.GB_IV) + '\\' + source_JSON_file_name) 

    log('Uploaded Intervention Group DB, ' + gbc.GB_MOD_STR[mod_ID])
            
