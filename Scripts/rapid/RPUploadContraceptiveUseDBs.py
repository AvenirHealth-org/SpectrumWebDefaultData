import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.RP.RPConst as rpc
import SpectrumCommon.Const.RP.RPDatabaseConst as rpdbc 

def create_contraceptive_use_DBs_RP():

    log('Creating RP contraceptive use DBs')

    def create_DB(DB_type = int):

        if DB_type == rpdbc.RP_MEDIAN_DB:
            DB_name = rpc.RP_CONTRACEPTIVE_USE_MED_DB_NAME
            DB_dir = rpc.RP_CONTRACEPTIVE_USE_MED_DB_DIR
            DB_version = rpc.RP_CONTRACEPTIVE_USE_MED_DB_CURR_VERSION

        else:
            DB_name = rpc.RP_CONTRACEPTIVE_USE_ACC_DB_NAME
            DB_dir = rpc.RP_CONTRACEPTIVE_USE_ACC_DB_DIR
            DB_version = rpc.RP_CONTRACEPTIVE_USE_ACC_DB_CURR_VERSION

        country_list = []
        wb = load_workbook(ddu.get_source_data_path(gbc.GB_RP) + '\RPModData.xlsx')
        sheet = wb[DB_name]
        
        first_data_row = 2
        first_data_col = 4
        num_rows = sheet.max_row

        # Get each row from the sheet and create a dictionary for each each country.
        for row in islice(sheet.values, first_data_row - 1, num_rows):
            country_dict = {}

            iso3 = row[1]

            if iso3 != None:
                country_dict[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_CUDB] = iso3
                contraceptive_use = []

                for col, col_val in enumerate(row):
                    if col >= (first_data_col - 1):
                        contraceptive_use.append(col_val)

                country_dict[rpdbc.RP_CONTRACEPTIVE_USE_KEY_CUDB] = contraceptive_use

                country_list.append(country_dict)

        RP_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + DB_dir
        os.makedirs(RP_JSON_data_path + '\\', exist_ok = True)
        
        for country_obj in country_list:    
            iso3 = country_obj[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_CUDB]
            with open(RP_JSON_data_path + '\\' + iso3 + '_' + DB_version + '.' + gbc.GB_JSON, 'w') as f:
                ujson.dump(country_obj, f)

    create_DB(rpdbc.RP_MEDIAN_DB)
    create_DB(rpdbc.RP_ACCELERATED_DB)

    log('Finished RP contraceptive use DBs')

def upload_contraceptive_use_DBs_RP():
    walk_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_CONTRACEPTIVE_USE_MED_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_RP_CONTAINER, walk_path, rpc.RP_CONTRACEPTIVE_USE_MED_DB_CURR_VERSION, rpc.RP_CONTRACEPTIVE_USE_MED_DB_DIR + '\\')

    walk_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_CONTRACEPTIVE_USE_ACC_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_RP_CONTAINER, walk_path, rpc.RP_CONTRACEPTIVE_USE_ACC_DB_CURR_VERSION, rpc.RP_CONTRACEPTIVE_USE_ACC_DB_DIR + '\\')

    log('Uploaded RP contraceptive use DBs')
            
