import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.RP.RPConst as rpc
import SpectrumCommon.Const.RP.RPDatabaseConst as rpdbc 

def create_urban_DB_RP(version):

    log('Creating RP urban DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_RP) + '\RPModData.xlsx')
    sheet = wb[rpc.RP_URBAN_DB_NAME]
    
    first_data_row = 2
    first_data_col = 4
    num_rows = sheet.max_row

    # Get each row from the sheet and create a dictionary for each each country.
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}

        iso3 = row[1]

        if iso3 != None:
            country_dict[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_UDB] = iso3
            perc_pop_living_in_urban_areas = []

            for col, col_val in enumerate(row):
                if col >= (first_data_col - 1):
                    perc_pop_living_in_urban_areas.append(col_val)

            country_dict[rpdbc.RP_PERC_POP_LIVING_URBAN_AREAS_KEY_UDB] = perc_pop_living_in_urban_areas

            country_list.append(country_dict)

    RP_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_URBAN_DB_DIR
    os.makedirs(RP_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_UDB]
        with open(RP_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished RP urban DB')

def upload_urban_DB_RP(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_URBAN_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_RP_CONTAINER, walk_path, version, rpc.RP_URBAN_DB_DIR + '\\')

    log('Uploaded RP urban DB')
            
