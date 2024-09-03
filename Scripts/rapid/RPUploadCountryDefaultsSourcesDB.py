import os
import ujson
from itertools import islice
from openpyxl import load_workbook
from copy import deepcopy

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.RP.RPConst as rpc
import SpectrumCommon.Const.RP.RPDatabaseConst as rpdbc 

def create_country_defaults_sources_DB_RP(version = str):

    log('Creating RP country defaults sources DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_RP) + '\RPModData.xlsx')
    sheet = wb[rpc.RP_COUNTRY_DEFAULTS_SOURCES_DB_NAME]
    
    tag_row = 1
    first_data_row = 2
    first_data_col = 4
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    DB_tag_columns = deepcopy(rpdbc.RP_tag_columns_CDDB)

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # If the tag is valid, save the column we find it at. This will allow us to move the columns around without needing to change
        # the code.
        if DB_tag_columns.get(tag):
            DB_tag_columns[tag] = c

    # Get each row from the sheet and create a dictionary for each each country.
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}

        iso3 = row[DB_tag_columns[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_TAG_CDDB] - 1]

        if iso3 != None:
            # For each tag in the database,... 
            for tag in DB_tag_columns:
                # If a corresponding column was found in the database,...
                if DB_tag_columns[tag] != gbc.GB_NOT_FOUND:
                    # If a field was set up for the key (fields may not be set up for columns we don't care about), save the
                    # value from the column we're at to the appropriate key in the country dict.
                    if rpdbc.RP_tag_fields_CDDB[tag] != gbc.GB_NOT_FOUND:
                        value = row[DB_tag_columns[tag] - 1]
                        country_dict[rpdbc.RP_tag_fields_CDDB[tag]] = value

            country_list.append(country_dict)

    RP_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_COUNTRY_DEFAULTS_SOURCES_DB_DIR
    os.makedirs(RP_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_CDDB]
        with open(RP_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished RP country defaults sources DB')

def upload_country_defaults_sources_DB_RP(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_COUNTRY_DEFAULTS_SOURCES_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_RP_CONTAINER, walk_path, version, rpc.RP_COUNTRY_DEFAULTS_SOURCES_DB_DIR + '\\')
    log('Uploaded RP country defaults sources DB')
            
