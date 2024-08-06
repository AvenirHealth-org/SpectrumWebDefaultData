import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.RP.RPConst as rpc
import SpectrumCommon.Const.RP.RPDatabaseConst as rpdbc 

RP_DB_tag_columns = {
    rpdbc.RP_ISO_NUMERIC_COUNTRY_CODE_DB_TAG        : 1,
    rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_DB_TAG        : 2,
    rpdbc.RP_COUNTRY_NAME_DB_TAG                    : 3,
    rpdbc.RP_CLASSIFICATION_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_REGION_DB_TAG                          : gbc.GB_NOT_FOUND,
    rpdbc.RP_EXCH_RATE_DB_TAG                       : gbc.GB_NOT_FOUND,
    rpdbc.RP_HH_SIZE_DB_TAG                         : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_URBAN_INFORMAL_DB_TAG             : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_STERILIZATION_DB_TAG              : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_IUD_DB_TAG                        : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_IMPLANT_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_INJECTIONS_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_PILL_DB_TAG                       : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_CONDOMS_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_INDIRECTS_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_AGE_START_PRIM_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_AGE_END_PRIM_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_AGE_START_SEC_DB_TAG                   : gbc.GB_NOT_FOUND,
    rpdbc.RP_AGE_END_SEC_DB_TAG                     : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_ENROLL_PRIM_DB_TAG                : gbc.GB_NOT_FOUND,
    rpdbc.RP_PRIM_ENROLL_FEMALES_DB_TAG             : gbc.GB_NOT_FOUND,
    rpdbc.RP_PRIM_ENROLL_MALES_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_ENROLL_SEC_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_SEC_ENROLL_FEMALES_DB_TAG              : gbc.GB_NOT_FOUND,
    rpdbc.RP_SEC_ENROLL_MALES_DB_TAG                : gbc.GB_NOT_FOUND,
    rpdbc.RP_TEACHER_RATIO_PRIM_DB_TAG              : gbc.GB_NOT_FOUND,
    rpdbc.RP_TEACHER_RATIO_SEC_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_PRIM_COSTS_DB_TAG                      : gbc.GB_NOT_FOUND,
    rpdbc.RP_SEC_COSTS_DB_TAG                       : gbc.GB_NOT_FOUND,
    rpdbc.RP_PRIM_TEACHER_SALARY_DB_TAG             : gbc.GB_NOT_FOUND,
    rpdbc.RP_SEC_TEACHER_SALARY_DB_TAG              : gbc.GB_NOT_FOUND,
    rpdbc.RP_SKILLED_BIRTHS_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_ANC_4_PLUS_DB_TAG                      : gbc.GB_NOT_FOUND,
    rpdbc.RP_PROFS_DB_TAG                           : gbc.GB_NOT_FOUND,
    rpdbc.RP_DOCTOR_SALARY_PUBLIC_DB_TAG            : gbc.GB_NOT_FOUND,
    rpdbc.RP_NURSE_SALARY_PUBLIC_DB_TAG             : gbc.GB_NOT_FOUND,
    rpdbc.RP_SKILLED_BIRTH_COSTS_DB_TAG             : gbc.GB_NOT_FOUND,
    rpdbc.RP_ANC_COSTS_DB_TAG                       : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_DOCTORS_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_NURSES_DB_TAG                     : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_IMPROVED_WATER_SOURCE_DB_TAG      : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_IMPROVED_SANI_DB_TAG              : gbc.GB_NOT_FOUND,
    rpdbc.RP_UNIT_COST_1000_L_WATER_DB_TAG          : gbc.GB_NOT_FOUND,
    rpdbc.RP_CAP_COST_PP_IMPROVED_WATER_DB_TAG      : gbc.GB_NOT_FOUND,
    rpdbc.RP_CAP_COST_PP_IMPROVED_SANI_DB_TAG       : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_PEOPLE_PHASE_4_PLUS_DB_TAG        : gbc.GB_NOT_FOUND,
    rpdbc.RP_SEVERE_INSECURITY_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_AREA_HARVESTED_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_STAPLE_PRODUCTION_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_DOMINANT_CROP_DB_TAG                   : gbc.GB_NOT_FOUND,
    rpdbc.RP_AGRICULTURAL_LAND_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_CROPLAND_HECTARES_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_STAPLE_ANNUAL_CONSUMP_PER_CAP_DB_TAG   : gbc.GB_NOT_FOUND,
    rpdbc.RP_PROD_COST_PER_TON_DOMINANT_CROP_DB_TAG : gbc.GB_NOT_FOUND,
    rpdbc.RP_DAILY_COST_FOOD_AID_PER_PERSON_DB_TAG  : gbc.GB_NOT_FOUND,
    rpdbc.RP_PERC_POP_ELECTRICITY_DB_TAG            : gbc.GB_NOT_FOUND,
    rpdbc.RP_ELECTRICITY_PROD_PER_CAP_DB_TAG        : gbc.GB_NOT_FOUND,
    rpdbc.RP_COST_PER_KWH_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_LCOE_0_PER_PERSON_PER_DAY_DB_TAG       : gbc.GB_NOT_FOUND,
    rpdbc.RP_CLEAN_COOKING_DB_TAG                   : gbc.GB_NOT_FOUND,
    rpdbc.RP_CLEAN_COOKING_CAP_COST_DB_TAG          : gbc.GB_NOT_FOUND,
    rpdbc.RP_UNEMPLOYMENT_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_GDP_PER_CAP_DB_TAG                     : gbc.GB_NOT_FOUND,
    rpdbc.RP_AVG_GROWTH_RATE_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_LABOR_MALES_DB_TAG                     : gbc.GB_NOT_FOUND,
    rpdbc.RP_LABOR_FEMALES_DB_TAG                   : gbc.GB_NOT_FOUND,
    rpdbc.RP_LABOR_BOTH_DB_TAG                      : gbc.GB_NOT_FOUND,
    rpdbc.RP_YOUTH_NEED_ALL_DB_TAG                  : gbc.GB_NOT_FOUND,
    rpdbc.RP_YOUTH_NEED_MALE_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_YOUTH_NEED_FEMALE_DB_TAG               : gbc.GB_NOT_FOUND,
    rpdbc.RP_1_DEGREE_GDP_DB_TAG                    : gbc.GB_NOT_FOUND,
    rpdbc.RP_1_DEGREE_STAPLE_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_2_DEGREE_STAPLE_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_3_DEGREE_STAPLE_DB_TAG                 : gbc.GB_NOT_FOUND,
    rpdbc.RP_1_DEGREE_ON_ENERGY_DB_TAG              : gbc.GB_NOT_FOUND,
}

# List of all columns that can be strings. All others should be numbers
RP_DB_string_type_tags = [
    rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_DB_TAG, 
    rpdbc.RP_COUNTRY_NAME_DB_TAG, 
    rpdbc.RP_CLASSIFICATION_DB_TAG,
    rpdbc.RP_REGION_DB_TAG,
    rpdbc.RP_DOMINANT_CROP_DB_TAG
]

def create_country_defaults_DB_RP(version = str):

    log('Creating RP country defaults DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_RP) + '\RPModData.xlsx')
    sheet = wb[rpc.RP_COUNTRY_DEFAULTS_DB_NAME]
    
    tag_row = 1
    first_data_row = 2
    first_data_col = 4
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # If the tag is valid, save the column we find it at. This will allow us to move the columns around without needing to change
        # the code.
        if RP_DB_tag_columns.get(tag):
            RP_DB_tag_columns[tag] = c

    # Get each row from the sheet and create a dictionary for each each country.
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}

        iso3 = row[RP_DB_tag_columns[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_DB_TAG] - 1]

        if iso3 != None:
            # For each tag in the database,... 
            for tag in RP_DB_tag_columns:
                # If a corresponding column was found in the database,...
                if RP_DB_tag_columns[tag] != gbc.GB_NOT_FOUND:
                    # If a field was set up for the key (fields may not be set up for columns we don't care about), save the
                    # value from the column we're at to the appropriate key in the country dict.
                    if rpdbc.RP_DB_tag_fields[tag] != gbc.GB_NOT_FOUND:
                        value = row[RP_DB_tag_columns[tag] - 1]

                        # If the tag is for a field with a non string-type value and the value is either a string or None, change it to 0
                        if not (tag in RP_DB_string_type_tags):
                            if isinstance(value, str) or (value == None):
                                value = 0

                        country_dict[rpdbc.RP_DB_tag_fields[tag]] = value

            country_list.append(country_dict)

    RP_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_COUNTRY_DEFAULTS_DB_DIR
    os.makedirs(RP_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[rpdbc.RP_ISO_ALPHA_3_COUNTRY_CODE_KEY_CDDB]
        with open(RP_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished RP country defaults DB')

def upload_country_defaults_DB_RP(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_RP) + '\\' + rpc.RP_COUNTRY_DEFAULTS_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_RP_CONTAINER, walk_path, version, rpc.RP_COUNTRY_DEFAULTS_DB_DIR + '\\')
    log('Uploaded RP country defaults DB')
            
