"""
AM Prevention Needs Database Creation Script

This module creates JSON database files for AIDS Module (AM) prevention needs data
by extracting data from an Excel workbook and converting it to structured JSON format.
The data includes key populations, VMMC, PrEP, and condom intervention data by country.
"""

import os
from sqlite3 import Row
import openpyxl
import ujson

from AvenirCommon.Database import GB_upload_file


def CreatePreventionNeedsDB(version):
    """
    Main function to create prevention needs database from Excel source data.
    
    Args:
        version (str): Version identifier for the database files
    
    Process:
        1. Loads AMModData.xlsx from SourceData
        2. Extracts data for key populations, VMMC, PrEP, and condoms
        3. Organizes data by country (ISO code)
        4. Saves individual JSON files for each country
    """
    # Load the source Excel workbook with calculated values only
    aim_FQName = os.getcwd() + '\\DefaultData\\SourceData\\aim\\AMModData.xlsx'
    xlsx = openpyxl.load_workbook(aim_FQName, read_only=False, keep_vba=False, data_only=True, keep_links=False)
   
    # Dictionary to store all country data, keyed by ISO country code
    countries = {}
    
    # Extract data from different worksheets
    create_key_populations_db(countries, xlsx)  # FSW, MSM, PWID, TG populations
    create_vmmc_db(countries, xlsx)             # Voluntary Male Medical Circumcision
    create_prep_db(countries, xlsx)             # Pre-Exposure Prophylaxis
    create_condom_db(countries, xlsx)           # Condom distribution programs

    # Save individual JSON files for each country
    for country_iso in countries:
        country = countries[country_iso]
        # Create output directory if it doesn't exist
        os.makedirs(f'{os.getcwd()}\\DefaultData\\JSONData\\aim\\PreventionNeedsDB\\', exist_ok=True)
        
        # Write country-specific JSON file with version identifier
        with open(f'{os.getcwd()}\\DefaultData\\JSONData\\aim\\PreventionNeedsDB\\{country_iso}_{version}.JSON', 'w+') as f:
            ujson.dump(country, f)


def UploadPreventionNeedsDB(version):
    """
    Uploads generated JSON files to the database.
    
    Args:
        version (str): Version identifier to filter which files to upload
    
    Process:
        1. Scans JSONData directory for files containing the version string
        2. Uploads each matching file to the database using GB_upload_file
    """
    # Get database connection string from environment
    connection = os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
    countries = []
    
    # Construct path to JSON output directory
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    json_path= default_path+'\\JSONData\\aim\\PreventionNeedsDB\\'
    
    # Walk through directory and upload files matching the version
    for subdir, dirs, files in os.walk(json_path):
        for file in files:
            FQName = os.path.join(subdir, file)
            if version in FQName:  # Only upload files for this version
                print(FQName)
                # Upload to 'aim' module under PreventionNeedsDB folder
                GB_upload_file(connection, 'aim', 'PreventionNeedsDB\\'+file, FQName)


def create_key_populations_db(countryies, xlsx):
    """
    Extract key populations data (FSW, MSM, PWID, TG) from KeyPopulations worksheet.
    
    Args:
        countryies (dict): Dictionary to store country data
        xlsx (Workbook): OpenPyXL workbook object
        
    Data extracted:
        - Population size estimates for each key population
        - Coverage rates for interventions
        - Data sources
        - Ratio projections for 2030/2024
    """
    # Read key populations sheet 
    sheet = xlsx['KeyPopulations']
    
    # Get column headers from first row
    headers = [cell.value for cell in sheet[1]]
    
    # Process each data row (starting from row 2)
    for row in range(2, sheet.max_row + 1):
        values = [cell.value for cell in sheet[row]]
        row_data = dict(zip(headers, values))  # Create dict mapping headers to values

        # Initialize country entry if not exists
        if row_data['ISO'] not in countryies:
            countryies[row_data['ISO']] = {}
            
        # Extract data for each key population type
        key_pop_data = []
        for kp in ['FSW', 'MSM', 'PWID', 'TG']:  # Female Sex Workers, Men who have Sex with Men, People Who Inject Drugs, Transgender
            key_pop_data.append({
                'population': row_data[kp + 'pop'],      # Population size estimate
                'coverage': row_data[kp + 'cov'],        # Current coverage rate
                'source': row_data[kp + 'source'],       # Data source reference
                'userCoverage': 90.0,                    # Default user-defined coverage target
                'id': kp                                 # Key population identifier
            })
        
        # Store key populations data and projection ratio
        countryies[row_data['ISO']]['KeyPopulations'] = {
            'data': key_pop_data,
            'ratio': row_data['Ratio_2030_2024']  # Projection scaling factor
        }


def create_vmmc_db(countryies, xlsx):
    """
    Extract Voluntary Male Medical Circumcision (VMMC) data from VMMC worksheet.
    
    Args:
        countryies (dict): Dictionary to store country data
        xlsx (Workbook): OpenPyXL workbook object
        
    Data extracted:
        - Regional/area-specific VMMC program data
        - Population targets and coverage rates
        - Traditional circumcision rates
        - Age distribution parameters
    """
    # Read VMMC sheet
    sheet = xlsx['VMMC']
    
    # Get column headers from first row
    headers = [cell.value for cell in sheet[1]]
 
    # Process each data row
    for row in range(2, sheet.max_row + 1):
        values = [cell.value for cell in sheet[row]]
        row_data = dict(zip(headers, values))

        # Initialize country entry if not exists
        if row_data['ISO'] not in countryies:
            countryies[row_data['ISO']] = {}

        # Initialize VMMC array if not exists (can have multiple regions per country)
        if 'VMMC' not in countryies[row_data['ISO']]:
            countryies[row_data['ISO']]['VMMC'] = []

        # Add VMMC program data for this area/region
        countryies[row_data['ISO']]['VMMC'].append({
            'name'         : row_data['Area'],          # Geographic area name
            'region'       : row_data['Region'],        # Region classification
            'popFirstYear' : row_data["MalePop2025"],   # Male population in 2025
            'popFinalYear' : row_data["MalePop2030"],   # Male population in 2030
            'covFirstYear' : row_data["CircCov2025"],   # Circumcision coverage in 2025
            'targetCov'    : row_data["TargetCov"],     # Target coverage rate
            'tradCircum'   : row_data["TradCirc"],      # Traditional circumcision rate
            'percAged15'   : row_data["PctAged15"],     # Percentage aged 15
            'percAged29'   : row_data["PctAged29"]      # Percentage aged 29
        })


def create_prep_db(countryies, xlsx):
    """
    Extract Pre-Exposure Prophylaxis (PrEP) data from PrEP worksheets.
    
    Args:
        countryies (dict): Dictionary to store country data
        xlsx (Workbook): OpenPyXL workbook object
        
    Data sources:
        - PrEPParameters: Global efficacy/delivery parameters by population and PrEP type
        - PrEPPopulations: Country-specific population sizes and coverage data
        
    PrEP types: Oral, Ring, 2-month injectable, 6-month injectable
    """
    # Read PrEP parameter and population sheets
    param_sheet = xlsx['PrEPParameters']
    pop_sheet = xlsx['PrEPPopulations']
    
    # Define population types eligible for PrEP
    pop_types  = ["Sero-different couples", "AGYW", "MSM", "Sex_workers", "PWID", "TG", "Pregnant and breastfeeding women"]
    prep_types = ["Oral", "Ring", "2-month", "6-month"]  # Available PrEP formulations
    years      = [2025, 2026, 2027, 2028, 2029, 2030]   # Projection years
    
    # Define column positions in PrEPPopulations sheet
    first_year_col = 5                              # Starting column for 2025 population data
    final_year_col = first_year_col+len(pop_types)+1  # Starting column for 2030 population data  
    coverage_col   = first_year_col+2*len(pop_types)+2  # Starting column for coverage data
    
    # Create lookup dictionary for population type indices
    pop_type_indices = { name:i for i, name in enumerate(pop_types) }
    
    # Initialize global parameters array
    global_params = [{} for _ in range(len(pop_types))]

    # Map parameter sheet population names to our standardized names
    param_pop_indices = {
        "Sero-different couples " : 0,
        "Adolescent girls and young women (AGYW)" : 1,
        "Men who have sex with men" : 2,
        "Female sex workers" : 3,
        "People who inject drugs" : 4,
        "Pregnant and breastfeeding people" : 6,
        "Transgender people" : 5
    }
    
    # Find columns for each PrEP type in parameter sheet
    oral_cols = [cell.col_idx-1 for cell in param_sheet[1] if cell.value.strip()=="Oral"]
    ring_cols = [cell.col_idx-1 for cell in param_sheet[1] if cell.value.strip()=="Ring"]
    two_month_cols = [cell.col_idx-1 for cell in param_sheet[1] if cell.value.strip()=="2-month"]
    six_month_cols = [cell.col_idx-1 for cell in param_sheet[1] if cell.value.strip()=="6-month"]
    
    # Extract global parameters for each population type
    for row in range(2, param_sheet.max_row + 1):
        param_values = [cell.value for cell in param_sheet[row]]
        pop_type = param_values[0]  # Population type name
        i = param_pop_indices[pop_type]  # Get index in our array
        
        # Store efficacy/delivery parameters for each PrEP type
        global_params[i] = {
            'id': pop_types[i],
            'targetCoverage': param_values[1],  # Default target coverage
            'Oral': [param_values[i] for i in oral_cols],        # Oral PrEP parameters by year
            'Ring': [param_values[i] for i in ring_cols],        # Ring parameters by year
            'TwoMonth': [param_values[i] for i in two_month_cols], # 2-month injectable by year
            'SixMonth': [param_values[i] for i in six_month_cols]  # 6-month injectable by year
        }

    # Extract country-specific population and coverage data
    for row in range(2, pop_sheet.max_row + 1):
        pop_values = [cell.value for cell in pop_sheet[row]]
        row_ISO = pop_sheet['C' + str(row)].value  # Country ISO code from column C
        
        # Initialize country entry if not exists
        if row_ISO not in countryies:
            countryies[row_ISO] = {}

        # Initialize PrEP data structure
        if 'PrEP' not in countryies[row_ISO]:
            countryies[row_ISO]['PrEP'] = {'data':[], 'popTypeIndices'  : {}}
        
        # Process each population type
        c = 0
        for pop_type in pop_types:
            # Special target coverage calculation for AGYW and pregnant women (doubled)
            if pop_type in ["AGYW", "Pregnant and breastfeeding women"]:
                target_cov = pop_values[coverage_col + c]*2
            else:
                target_cov = global_params[c]['targetCoverage']
                
            # Store population-specific data with global parameters
            # Note: The key names will need to change at a later 
            # date to be lowercase to match the rest of the codebase
            countryies[row_ISO]['PrEP']['data'].append({
                'id': pop_type,
                'PopFirstYear': pop_values[first_year_col + c],  # 2025 population
                'PopFinalYear': pop_values[final_year_col + c],  # 2030 population
                'PrEPCov': pop_values[coverage_col + c],         # Current coverage
                'TargetCov': target_cov,                        # Target coverage
                'Oral': global_params[c]['Oral'],               # Oral PrEP efficacy by year
                'Ring': global_params[c]['Ring'],               # Ring efficacy by year
                'TwoMonth': global_params[c]['TwoMonth'],       # 2-month injectable efficacy by year
                'SixMonth': global_params[c]['SixMonth']        # 6-month injectable efficacy by year
            })
            countryies[row_ISO]['PrEP']['popTypeIndices'][pop_type] = c
            c += 1


def create_condom_db(countryies, xlsx):
    """
    Extract condom distribution program data from Condom worksheets.
    
    Args:
        countryies (dict): Dictionary to store country data
        xlsx (Workbook): OpenPyXL workbook object
        
    Data sources:
        - CondomPopulations: Country-specific population and coverage data
        - CondomParameters: Global parameters for target coverage, sexual activity, wastage
        
    Condom programs target different populations and use cases including
    HIV prevention and family planning.
    """
    # Read condom population and parameter sheets
    condom_pop_sheet = xlsx['CondomPopulations']
    condom_param_sheet = xlsx['CondomParameters']
    
    # Define condom target populations (equivalent to TCondomType enum from Delphi)
    condom_types = [
        "PLHIV_couples",	            # People Living with HIV in couples
        "Sex_workers",	                # Female sex workers
        "MSM",	                        # Men who have sex with men
        "YP_15_24_nr_partners",	        # Young people 15-24 with non-regular partners
        "M&W_25_34_nr_partners",	    # Men & women 25-34 with non-regular partners
        "M&W_35_49_nr_partners",	    # Men & women 35-49 with non-regular partners
        "M&W_50_plus_nr_partners",	    # Men & women 50+ with non-regular partners
        "Couples_using_condoms_FP",	    # Couples using condoms for family planning
        "Unmet_need_for_FP",	        # Unmet need for family planning
        "PWID",	                        # People who inject drugs
        "TG"                            # Transgender people
    ]
    
    # Column arrays mapping to Excel sheet structure (converted from Delphi arrays, +1 for Python indexing)
    pop_first_year_cols = [6, 7, 8, 10, 11, 12, 13, 14, 15, 16, 17]    # 2025 population columns
    pop_final_year_cols = [19, 20, 21, 23, 24, 25, 26, 27, 28, 29, 30]  # 2030 population columns
    baseline_cov_cols = [32, 33, 34, 35, 36, 37, 38, 39, 40, 41, 42]   # Baseline coverage columns
    
    # Parameter sheet column positions (all populations use same columns)
    target_cov_col = 2  # Target coverage column
    sex_acts_col = 3    # Sexual acts per year column
    wastage_col = 4     # Condom wastage rate column
    
    # Read global parameters for each condom type
    global_params = {}
    for r, condom_type in enumerate(condom_types, start=1):
        if r <= condom_param_sheet.max_row:
            param_row = r + 1  # Skip header row
            
            # Special handling for family planning condom types
            # Their target coverage equals their baseline coverage (set later)
            if condom_type in ["Couples_using_condoms_FP", "Unmet_need_FP"]:
                target_cov = None  # Will be set to baseline coverage later
            else:
                target_cov = condom_param_sheet.cell(row=param_row, column=target_cov_col).value
            
            # Store global parameters for this condom type
            global_params[condom_type] = {
                'targetCov': target_cov,
                'sexActs': condom_param_sheet.cell(row=param_row, column=sex_acts_col).value,    # Sexual acts per year
                'wastage': condom_param_sheet.cell(row=param_row, column=wastage_col).value      # Wastage rate
            }
    
    # Read country-specific population and coverage data
    for row in range(2, condom_pop_sheet.max_row + 1):
        row_ISO = condom_pop_sheet.cell(row=row, column=3).value  # Country ISO code from column C
        
        # Skip rows without country code
        if not row_ISO:
            continue
            
        # Initialize country entry if not exists
        if row_ISO not in countryies:
            countryies[row_ISO] = {}

        # Initialize condom data structure
        if 'Condoms' not in countryies[row_ISO]:
            countryies[row_ISO]['Condoms'] = {'data': [], 'condomTypeIndices': {}}
        
        # Process each condom type/population
        for i, condom_type in enumerate(condom_types):
            try:
                # Extract population and coverage data from appropriate columns
                pop_first_year = condom_pop_sheet.cell(row=row, column=pop_first_year_cols[i]).value
                pop_final_year = condom_pop_sheet.cell(row=row, column=pop_final_year_cols[i]).value
                baseline_cov = condom_pop_sheet.cell(row=row, column=baseline_cov_cols[i]).value
                
                # Special case: family planning condom programs use baseline as target
                if condom_type in ["Couples_using_condoms_FP", "Unmet_need_FP"]:
                    target_cov = baseline_cov  # Set TargetCov to BaselineCov
                else:
                    target_cov = global_params[condom_type]['targetCov']
                
                # Store complete condom program data
                countryies[row_ISO]['Condoms']['data'].append({
                    'id': condom_type,
                    'popFirstYear': float(pop_first_year) if pop_first_year is not None else 0.0,   # 2025 population
                    'popFinalYear': float(pop_final_year) if pop_final_year is not None else 0.0,   # 2030 population
                    'baselineCov': float(baseline_cov) if baseline_cov is not None else 0.0,        # Current coverage
                    'targetCov': float(target_cov) if target_cov is not None else 0.0,              # Target coverage
                    'sexActs': float(global_params[condom_type]['sexActs']) if global_params[condom_type]['sexActs'] is not None else 0.0,  # Sexual acts/year
                    'wastage': float(global_params[condom_type]['wastage']) if global_params[condom_type]['wastage'] is not None else 0.0    # Wastage rate
                })
                
                # Store index mapping for quick lookup
                countryies[row_ISO]['Condoms']['condomTypeIndices'][condom_type] = i
                
            except (ValueError, TypeError, IndexError) as e:
                # Error handling: log error and add default values
                print(f"Error processing condom type {condom_type} for country {row_ISO}: {e}")
                
                # Add default values when data is missing or invalid
                countryies[row_ISO]['Condoms']['data'].append({
                    'id': condom_type,
                    'popFirstYear': 0.0,
                    'popFinalYear': 0.0,
                    'baselineCov': 0.0,
                    'targetCov': 0.0,
                    'sexActs': 0.0,
                    'wastage': 0.0
                })
                countryies[row_ISO]['Condoms']['condomTypeIndices'][condom_type] = i