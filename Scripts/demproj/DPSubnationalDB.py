import os
import openpyxl
from openpyxl.utils import get_column_letter
from AvenirCommon.Database import GB_upload_json, GB_upload_file
import ujson
import numpy as np
import cProfile
import pstats
import datetime

import logging
from SpectrumCommon.Const.TB import *
from SpectrumCommon.Const.GB import GB_Nan

def create_DP_subnationals(version):
    
    # TB_Path = os.getcwd()+'\\Tools\DefaultDataManager\\TB\\'
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    
    # profile = cProfile.Profile()
    # profile.enable()
    countries = TB_RUN_DEFAULT_COUNTRIES
    countries = ['NGA', ]
    age_map = {
        'Y000_004' : 0,
        'Y005_009' : 1, 
        'Y010_014' : 2,
        'Y015_019' : 3,
        'Y020_024' : 4,
        'Y025_029' : 5,
        'Y030_034' : 6,
        'Y035_039' : 7,
        'Y040_044' : 8,
        'Y045_049' : 9,
        'Y050_054' : 10,
        'Y055_059' : 11,
        'Y060_064' : 12,
        'Y065_069' : 13,
        'Y070_074' : 14,
        'Y075_079' : 15,
        'Y080_999' : 16,
        'Y015_049' : 17, 
    }
    indicators = ['population',
                  'prevalence',
                  'plhiv',
                  'art_coverage',
                  'art_current_residents',
                  'art_current',
                  'untreated_plhiv_num',
                  'incidence',
                  'infections',
                  'anc_art_coverage']


          
    for iso3 in countries: #81 GRL 137 ANT 174 SCG 197 TKL
        print(iso3)
        try:
       
            who_tb_fqname = f'{default_path}\\SourceData\\demproj\\NGA_SubNat_Data.xlsx'
            xlsx = openpyxl.load_workbook(who_tb_fqname, read_only=False, keep_vba=False, data_only=False, keep_links=True)
            # pages = ['TB_burden_countries']
            subnationals = {}
            sheet = xlsx['NaomiNGA']
            
            
            for row in range(2, xlsx['NaomiNGA'].max_row+1):
                subnational_name = sheet[f'D{row}'].value
                subnational_id = sheet[f'C{row}'].value
                if subnational_id not in subnationals:
                    subnationals[subnational_id] = {
                        'iso3': iso3,
                        'name': subnational_name,
                        'areaID': subnational_id,
                        'totalInfection': 0,
                        'totalHIVPop': 0,
                        
                        }
                    for indicator in indicators:
                        subnationals[subnational_id][indicator] = {
                            'mean': np.zeros((2, 18)).tolist(),
                            'se': np.zeros((2, 18)).tolist(),
                            'median': np.zeros((2, 18)).tolist(),
                            'mode': np.zeros((2, 18)).tolist(),
                            'lower': np.zeros((2, 18)).tolist(),
                            'upper': np.zeros((2, 18)).tolist(),
                        }
                    
                    
                    
                sex = 0 if sheet[f'E{row}'].value.lower() =="male" else 1
                age = age_map[sheet[f'F{row}'].value]
                indicator = sheet[f'J{row}'].value
                subnationals[subnational_id][indicator]['mean'][sex][age]   = sheet[f'L{row}'].value
                subnationals[subnational_id][indicator]['se'][sex][age]     = sheet[f'M{row}'].value
                subnationals[subnational_id][indicator]['median'][sex][age] = sheet[f'N{row}'].value
                subnationals[subnational_id][indicator]['mode'][sex][age]   = sheet[f'O{row}'].value
                subnationals[subnational_id][indicator]['lower'][sex][age]  = sheet[f'P{row}'].value
                subnationals[subnational_id][indicator]['upper'][sex][age]  = sheet[f'Q{row}'].value
                if ((age==17) and (indicator == 'infections')):
                    subnationals[subnational_id]['totalInfection']  +=sheet[f'L{row}'].value
                if ((age==17) and (indicator == 'plhiv')):
                    subnationals[subnational_id]['totalHIVPop']  +=sheet[f'L{row}'].value
                    
            os.makedirs(default_path+'\\JSONData\\demproj\\subnationals\\'+iso3+'\\', exist_ok=True)
            for subnational_id in subnationals:
                subnationals[subnational_id]['totalIncidence'] = subnationals[subnational_id]['totalInfection'] / subnationals[subnational_id]['totalHIVPop']
                with open(f'{default_path}\\JSONData\\demproj\\subnationals\\{iso3}\\{subnational_id}_{version}.JSON', 'w') as f:
                    ujson.dump(subnationals[subnational_id], f)
                
        except Exception as e:
            logging.exception(f'Error processing {iso3} {e}')   
            print(f'{iso3} exception {e}')            
                    
    # profile.disable()
    # ps = pstats.Stats(profile)
    # ps.sort_stats('calls', 'cumtime')
    # ps.print_stats(50)
    pass

def upload_DP_subnationals(version):
    connection =  os.environ['AVENIR_SW_DEFAULT_DATA_CONNECTION']
    countries = []
    
    default_path = os.getcwd()+'\\' + __name__.split('.')[0] 
    json_path= default_path+'\\JSONData\\demproj\\subnationals\\'
    for subdir, dirs, files in os.walk(json_path):
        for file in files:
            FQName = os.path.join(subdir, file)
            if version in FQName:
                print(FQName)
                GB_upload_file(connection, 'demproj', 'subnationals\\'+file, FQName)
        
    logging.debug('Uploaded TB who db json')