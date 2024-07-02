import numpy as np
import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_AREA_TYPE_ID         = '<Area Type ID>'  
PC_AREA_TYPE_NAME       = '<Area Type Name>'   
PC_AREA_ID              = '<Area ID>'  
PC_AREA_NAME            = '<Area Name>'   
PC_COSTING_SECT_ID      = '<Costing Section ID>'    
PC_COSTING_SECT_NAME    = '<Costing Section Name>' 
PC_COSTING_SUBSECT_ID   = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME = '<Costing Subsection Name>' 
PC_ACT_ID               = '<Activity ID>' 
PC_ACT_NAME             = '<Activity Name>' 
PC_COST_INPUT_ID        = '<Cost Input ID>' 
PC_COST_INPUT_NAME      = '<Cost Input Name>' 
PC_NUM_UNITS            = '<Number of Units>' 
PC_DURATION             = '<Duration>' 
PC_QUANTITY             = '<Quantity>'
PC_PHASE_ID             = '<Phase ID>'
PC_PHASE_NAME           = '<Phase Name>'
PC_ADMIN_LEVEL_ID       = '<Admin Level ID>'
PC_ADMIN_LEVEL_NAME     = '<Admin Level Name>'

def create_PPLI_activity_line_item_DB_PC(version = str):

    log('Creating PC PPLI activity line item DB')

    activity_line_item_list = []

    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_PPLI_ACTIVITY_LINE_ITEM_DB_NAME]

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 6
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # row = 1
    # row += 1
    # phase_ID_row = row
    # row += 1
    # phase_name_row = row 
    # row += 1
    # admin_level_ID_row = row
    # row += 1
    # admin_level_name_row = row 

    area_type_ID_col = gbc.GB_NOT_FOUND
    # area_type_name_col = gbc.GB_NOT_FOUND
    area_ID_col = gbc.GB_NOT_FOUND
    # area_name_col = gbc.GB_NOT_FOUND
    costing_sect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_name_col = gbc.GB_NOT_FOUND
    costing_subsect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_section_name_col = gbc.GB_NOT_FOUND
    act_ID_col = gbc.GB_NOT_FOUND
    #act_name_col = gbc.GB_NOT_FOUND
    act_line_item_ID_col = gbc.GB_NOT_FOUND
    # act_line_item_name_col = gbc.GB_NOT_FOUND
    num_units_start_col = gbc.GB_NOT_FOUND
    duration_start_col = gbc.GB_NOT_FOUND
    quantity_start_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_AREA_TYPE_ID:
            area_type_ID_col = c

        # elif tag == PC_AREA_TYPE_NAME:
        #     area_type_name_col = c

        elif tag == PC_AREA_ID:
            area_ID_col = c

        # elif tag == PC_AREA_NAME:
        #     area_name_col = c

        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c

        # elif tag == PC_COST_CAT_NAME:
        #     cost_cat_name_col = c

        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c

        # elif tag == PC_COST_CAT_SECTION_NAME:
        #     cost_cat_section_name_col = c

        elif tag == PC_ACT_ID:
            act_ID_col = c

        # elif tag == PC_ACT_NAME:
        #     act_name_col = c

        elif tag == PC_COST_INPUT_ID:
            act_line_item_ID_col = c

        # elif tag == PC_ACT_INPUT_NAME:
        #     act_input_name_col = c

        elif tag == PC_NUM_UNITS:
            num_units_start_col = c
        
        elif tag == PC_DURATION:
            duration_start_col = c

        elif tag == PC_QUANTITY:
            quantity_start_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        act_line_item_dict = {}
        act_line_item_dict[pcdbk.PC_AREA_TYPE_ID_KEY_PALIDB] = row[area_type_ID_col - 1]
        act_line_item_dict[pcdbk.PC_AREA_ID_KEY_PALIDB] = str(row[area_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COSTING_SECT_ID_KEY_PALIDB] = str(row[costing_sect_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COSTING_SUBSECT_ID_KEY_PALIDB] = str(row[costing_subsect_ID_col - 1])
        act_line_item_dict[pcdbk.PC_ACT_ID_KEY_PALIDB] = str(row[act_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COST_INPUT_ID_KEY_PALIDB] = str(row[act_line_item_ID_col - 1])

        act_line_item_dict[pcdbk.PC_NUM_UNITS_KEY_PALIDB] = np.full((pcc.PC_PPLI_NUM_ADMIN_LEVELS), 0.0).tolist()
        act_line_item_dict[pcdbk.PC_NUM_UNITS_KEY_PALIDB][pcc.PC_PPLI_NATIONAL_ADMIN_LEVEL_CURR_ID - 1] = row[num_units_start_col - 1]
        act_line_item_dict[pcdbk.PC_NUM_UNITS_KEY_PALIDB][pcc.PC_PPLI_PROVINCE_ADMIN_LEVEL_CURR_ID - 1] = row[num_units_start_col]
        act_line_item_dict[pcdbk.PC_NUM_UNITS_KEY_PALIDB][pcc.PC_PPLI_DISTRICT_ADMIN_LEVEL_CURR_ID - 1] = row[num_units_start_col + 1]

        act_line_item_dict[pcdbk.PC_DURATION_KEY_PALIDB] = np.full((pcc.PC_PPLI_NUM_ADMIN_LEVELS), 0.0).tolist()
        act_line_item_dict[pcdbk.PC_DURATION_KEY_PALIDB][pcc.PC_PPLI_NATIONAL_ADMIN_LEVEL_CURR_ID - 1] = row[duration_start_col - 1]
        act_line_item_dict[pcdbk.PC_DURATION_KEY_PALIDB][pcc.PC_PPLI_PROVINCE_ADMIN_LEVEL_CURR_ID - 1] = row[duration_start_col]
        act_line_item_dict[pcdbk.PC_DURATION_KEY_PALIDB][pcc.PC_PPLI_DISTRICT_ADMIN_LEVEL_CURR_ID - 1] = row[duration_start_col + 1]

        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB] = np.full((pcc.PC_NUM_PHASES, pcc.PC_PPLI_NUM_ADMIN_LEVELS), 0.0).tolist()
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_EARLY_IMP_PHASE_25_PPLI_CURR_ID - 1][pcc.PC_PPLI_NATIONAL_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col - 1]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_EARLY_IMP_PHASE_25_PPLI_CURR_ID - 1][pcc.PC_PPLI_PROVINCE_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_EARLY_IMP_PHASE_25_PPLI_CURR_ID - 1][pcc.PC_PPLI_DISTRICT_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 1]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_PARTIAL_IMP_PHASE_50_PPLI_CURR_ID - 1][pcc.PC_PPLI_NATIONAL_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 2]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_PARTIAL_IMP_PHASE_50_PPLI_CURR_ID - 1][pcc.PC_PPLI_PROVINCE_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 3]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_PARTIAL_IMP_PHASE_50_PPLI_CURR_ID - 1][pcc.PC_PPLI_DISTRICT_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 4]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_FULL_IMP_PHASE_100_PPLI_CURR_ID - 1][pcc.PC_PPLI_NATIONAL_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 5]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_FULL_IMP_PHASE_100_PPLI_CURR_ID - 1][pcc.PC_PPLI_PROVINCE_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 6]
        act_line_item_dict[pcdbk.PC_QUANTITY_KEY_PALIDB][pcc.PC_FULL_IMP_PHASE_100_PPLI_CURR_ID - 1][pcc.PC_PPLI_DISTRICT_ADMIN_LEVEL_CURR_ID - 1] = \
            row[quantity_start_col + 7]

        activity_line_item_list.append(act_line_item_dict)

    # Write activity input line items

    j = ujson.dumps(activity_line_item_list)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_PPLI_ACTIVITY_LINE_ITEM_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)
        
    log('Finished PC PPLI activity line item DB')

def upload_PPLI_activity_line_item_DB_PC(version):
    JSON_file_name = pcc.PC_PPLI_ACTIVITY_LINE_ITEM_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC PPLI activity line item DB') 

    
            
