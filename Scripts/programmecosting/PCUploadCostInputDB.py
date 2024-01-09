import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_COST_CAT_ID               = '<Cost Category ID>'  
PC_COST_CAT_NAME             = '<Cost Category Name>'   
PC_COST_INPUT_ID             = '<Cost Input ID>'    
PC_COST_INPUT_NAME           = '<Cost Input Name>' 
PC_COST_INPUT_STR_CONST      = '<Cost Input String Constant>'
PC_UNIT_COST                 = '<Unit Cost>'       
PC_UNIT_OF_MEASURE           = '<Unit of Measure>' 
PC_UNIT_OF_MEASURE_STR_CONST = '<Unit of Measure String Constant>'
PC_NOTE                      = '<Note>' 
PC_SOURCE                    = '<Source>'

def create_cost_input_DB_PC(version = str):
    log('Creating PC cost input DB')

    cost_input_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_COST_INPUT_DB_NAME]
    
    '''
        IMPORTANT: Max cost category master ID: xxx

        New cost category must have a master ID greater than this. Please update this maximum if you add a new one.
    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    cost_cat_ID_col = gbc.GB_NOT_FOUND
    cost_cat_name_col = gbc.GB_NOT_FOUND
    cost_input_ID_col = gbc.GB_NOT_FOUND
    cost_input_name_col = gbc.GB_NOT_FOUND
    cost_input_str_const_col = gbc.GB_NOT_FOUND
    unit_cost_col = gbc.GB_NOT_FOUND
    unit_of_measure_col = gbc.GB_NOT_FOUND
    unit_of_measure_str_const_col = gbc.GB_NOT_FOUND
    note_col = gbc.GB_NOT_FOUND
    source_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_COST_CAT_ID:
            cost_cat_ID_col = c

        elif tag == PC_COST_CAT_NAME:
            cost_cat_name_col = c

        elif tag == PC_COST_INPUT_ID:
            cost_input_ID_col = c

        elif tag == PC_COST_INPUT_NAME:
            cost_input_name_col = c

        elif tag == PC_COST_INPUT_STR_CONST:
            cost_input_str_const_col = c

        elif tag == PC_UNIT_COST:
            unit_cost_col = c

        elif tag == PC_UNIT_OF_MEASURE:
            unit_of_measure_col = c

        elif tag == PC_UNIT_OF_MEASURE_STR_CONST:
            unit_of_measure_str_const_col = c

        elif tag == PC_NOTE:
            note_col = c

        elif tag == PC_SOURCE:
            source_col = c

    for row in islice(sheet.values, first_data_row - 1, num_rows):

        cost_input_dict = {}

        note = row[note_col - 1]
        source = row[source_col - 1]

        cost_input_dict[pcdbk.PC_COST_CAT_ID_KEY_CIDB] = row[cost_cat_ID_col - 1]
        cost_input_dict[pcdbk.PC_COST_CAT_NAME_KEY_CIDB] = row[cost_cat_name_col - 1]
        cost_input_dict[pcdbk.PC_COST_INPUT_ID_KEY_CIDB] = row[cost_input_ID_col - 1]
        cost_input_dict[pcdbk.PC_COST_INPUT_NAME_KEY_CIDB] = row[cost_input_name_col - 1]
        cost_input_dict[pcdbk.PC_COST_INPUT_STR_CONST_KEY_CIDB] = row[cost_input_str_const_col - 1]
        cost_input_dict[pcdbk.PC_UNIT_COST_KEY_CIDB] = row[unit_cost_col - 1]
        cost_input_dict[pcdbk.PC_UNIT_OF_MEASURE_KEY_CIDB] = row[unit_of_measure_col - 1]
        cost_input_dict[pcdbk.PC_UNIT_OF_MEASURE_STR_CONST_KEY_CIDB] = row[unit_of_measure_str_const_col - 1]
        cost_input_dict[pcdbk.PC_NOTE_KEY_CIDB] = note if note != None else ''
        cost_input_dict[pcdbk.PC_SOURCE_KEY_CIDB] = source if source != None else ''
        cost_input_list.append(cost_input_dict)

    j = ujson.dumps(cost_input_list)
    # for debugging
    # for i, obj in enumerate(activity_list):
    #     log(obj)
    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_COST_INPUT_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished PC cost input DB')

def upload_cost_input_DB_PC(version):
    JSON_file_name = pcc.PC_COST_INPUT_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC cost input DB')
            
