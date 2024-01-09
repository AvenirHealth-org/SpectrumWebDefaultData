import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_COST_CAT_ID        = '<Cost Category ID>'  
PC_COST_CAT_NAME      = '<Cost Category Name>'  
PC_COST_CAT_STR_CONST = '<Cost Category String Constant>' 

def create_cost_category_DB_PC(version = str):

    log('Creating PC cost category DB')

    cost_cat_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_COST_CAT_DB_NAME]
    
    '''
        IMPORTANT: Max cost category master ID: xxx

        New cost category must have a master ID greater than this. Please update this maximum if you add a new one.
    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    cost_cat_ID_col = gbc.GB_NOT_FOUND
    cost_cat_name_col = gbc.GB_NOT_FOUND
    cost_cat_str_const = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_COST_CAT_ID:
            cost_cat_ID_col = c

        elif tag == PC_COST_CAT_NAME:
            cost_cat_name_col = c

        elif tag == PC_COST_CAT_STR_CONST:
            cost_cat_str_const = c

    for row in islice(sheet.values, first_data_row - 1, num_rows):

        cost_cat_dict = {}

        cost_cat_dict[pcdbk.PC_COST_CAT_ID_KEY_CCDB] = row[cost_cat_ID_col - 1]
        cost_cat_dict[pcdbk.PC_COST_CAT_NAME_KEY_CCDB] = row[cost_cat_name_col - 1]
        cost_cat_dict[pcdbk.PC_COST_CAT_STR_CONST_KEY_CCDB] = row[cost_cat_str_const - 1]

        cost_cat_list.append(cost_cat_dict)

    j = ujson.dumps(cost_cat_list)
    # for debugging
    # for i, obj in enumerate(cost_cat_list):
    #     log(obj)
    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_COST_CAT_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished PC cost category DB')

def upload_cost_category_DB_PC(version):
    JSON_file_name = pcc.PC_COST_CAT_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC cost category DB')
            
