import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_AREA_TYPE_ID         = '<Area Type ID>'  
PC_AREA_TYPE_NAME       = '<Area Type Name>'   
PC_AREA_ID              = '<Area ID>'  
PC_AREA_NAME            = '<Area Name>'   
PC_COSTING_SECT_ID      = '<Costing Section ID>'    
PC_COSTING_SECT_NAME    = '<Costing Section Name>' 
PC_COSTING_SUBSECT_ID   = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME = '<Costing Subsection Name>' 
PC_ACT_ID               = '<Activity ID>' 
PC_ACT_NAME             = '<Activity Name>' 
PC_ACT_STR_CONST        = '<Activity String Constant>'

def create_activity_DB_PC(version = str):

    log('Creating PC activity DB')

    activity_list = []

    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_ACTIVITY_DB_NAME]

    '''
        <Notes>
        
            New groups within a particular area, cost category, and cost category section must be given a master ID that is greater than any existing 
            group master ID within those confines.
  
    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    area_type_ID_col = gbc.GB_NOT_FOUND
    area_ID_col = gbc.GB_NOT_FOUND
    # area_name_col = gbc.GB_NOT_FOUND
    costing_sect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_name_col = gbc.GB_NOT_FOUND
    costing_subsect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_section_name_col = gbc.GB_NOT_FOUND
    act_ID_col = gbc.GB_NOT_FOUND
    act_name_col = gbc.GB_NOT_FOUND
    act_string_const_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_AREA_TYPE_ID:
            area_type_ID_col = c

        elif tag == PC_AREA_ID:
            area_ID_col = c

        # elif tag == PC_AREA_NAME:
        #     area_name_col = c

        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c

        # elif tag == PC_COST_CAT_NAME:
        #     cost_cat_name_col = c

        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c

        # elif tag == PC_COST_CAT_SECTION_NAME:
        #     cost_cat_section_name_col = c

        elif tag == PC_ACT_ID:
            act_ID_col = c

        elif tag == PC_ACT_NAME:
            act_name_col = c

        elif tag == PC_ACT_STR_CONST:
            act_string_const_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        # If the dict with the current intervention master ID, delivery channel master ID, item type master ID, and group master ID 
        # does not exist, create a new group master ID. Otherwise, leave the treatment input groups alone.  and add to it. Otherwise, add to the one from the previous loop.

        act_dict = {}
        act_dict[pcdbk.PC_AREA_TYPE_ID_KEY_ADB] = row[area_type_ID_col - 1]
        act_dict[pcdbk.PC_AREA_ID_KEY_ADB] = str(row[area_ID_col - 1])
        act_dict[pcdbk.PC_COSTING_SECT_ID_KEY_ADB] = str(row[costing_sect_ID_col - 1])
        act_dict[pcdbk.PC_COSTING_SUBSECT_ID_KEY_ADB] = str(row[costing_subsect_ID_col - 1])
        act_dict[pcdbk.PC_ACT_ID_KEY_ADB] = str(row[act_ID_col - 1])
        act_dict[pcdbk.PC_ACT_NAME_KEY_ADB] = row[act_name_col - 1]
        act_dict[pcdbk.PC_ACT_STR_CONST_KEY_ADB] = row[act_string_const_col - 1]
        activity_list.append(act_dict)
        
    # Write activities

    j = ujson.dumps(activity_list)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_ACTIVITY_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)
        
    log('Finished PC activity DB')

def upload_activity_DB_PC(version):
    JSON_file_name = pcc.PC_ACTIVITY_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC activity DB') 

    
            
