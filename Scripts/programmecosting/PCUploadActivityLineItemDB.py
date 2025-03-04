import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_AREA_TYPE_ID         = '<Area Type ID>'  
PC_AREA_TYPE_NAME       = '<Area Type Name>'   
PC_AREA_ID              = '<Area ID>'  
PC_AREA_NAME            = '<Area Name>'   
PC_COSTING_SECT_ID      = '<Costing Section ID>'    
PC_COSTING_SECT_NAME    = '<Costing Section Name>' 
PC_COSTING_SUBSECT_ID   = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME = '<Costing Subsection Name>' 
PC_ACT_ID               = '<Activity ID>' 
PC_ACT_NAME             = '<Activity Name>' 
PC_COST_INPUT_ID        = '<Cost Input ID>' 
PC_COST_INPUT_NAME      = '<Cost Input Name>' 
PC_NUM_UNITS            = '<Number of Units>' 
PC_DURATION             = '<Duration>' 
PC_FREQUENCY            = '<Frequency>' 

# To create a new database for testing, pass 'Testing' as the version. To use, set the database version to 'Testing'. Don't forget to change it back later! 
def create_activity_line_item_DB_PC(version : str = pcc.PC_ACTIVITY_LINE_ITEM_DB_CURR_VERSION):

    log('Creating PC activity line item DB')

    activity_line_item_list = []

    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = None
    if version == 'Testing':
        sheet = wb[pcc.PC_ACTIVITY_LINE_ITEM_DB_NAME + 'Testing']
    else:
        sheet = wb[pcc.PC_ACTIVITY_LINE_ITEM_DB_NAME]
    '''
        <Notes>
        
            New groups within a particular area, cost category, and cost category section must be given a master ID that is greater than any existing 
            group master ID within those confines.
  
    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    area_type_ID_col = gbc.GB_NOT_FOUND
    area_ID_col = gbc.GB_NOT_FOUND
    # area_name_col = gbc.GB_NOT_FOUND
    costing_sect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_name_col = gbc.GB_NOT_FOUND
    costing_subsect_ID_col = gbc.GB_NOT_FOUND
    # cost_cat_section_name_col = gbc.GB_NOT_FOUND
    act_ID_col = gbc.GB_NOT_FOUND
    #act_name_col = gbc.GB_NOT_FOUND
    cost_input_ID_col = gbc.GB_NOT_FOUND
    frequency_col = gbc.GB_NOT_FOUND
    num_units_col = gbc.GB_NOT_FOUND
    duration_col = gbc.GB_NOT_FOUND
    frequency_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_AREA_TYPE_ID:
            area_type_ID_col = c

        elif tag == PC_AREA_ID:
            area_ID_col = c

        # elif tag == PC_AREA_NAME:
        #     area_name_col = c

        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c

        # elif tag == PC_COST_CAT_NAME:
        #     cost_cat_name_col = c

        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c

        # elif tag == PC_COST_CAT_SECTION_NAME:
        #     cost_cat_section_name_col = c

        elif tag == PC_ACT_ID:
            act_ID_col = c

        # elif tag == PC_ACT_NAME:
        #     act_name_col = c

        elif tag == PC_COST_INPUT_ID:
            cost_input_ID_col = c

        # elif tag == PC_ACT_INPUT_NAME:
        #     act_input_name_col = c

        elif tag == PC_NUM_UNITS:
            num_units_col = c
        
        elif tag == PC_DURATION:
            duration_col = c

        elif tag == PC_FREQUENCY:
            frequency_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        # If the dict with the current intervention master ID, delivery channel master ID, item type master ID, and group master ID 
        # does not exist, create a new group master ID. Otherwise, leave the treatment input groups alone.  and add to it. Otherwise, add to the one from the previous loop.

        act_line_item_dict = {}
        act_line_item_dict[pcdbk.PC_AREA_TYPE_ID_KEY_ALIDB] = row[area_type_ID_col - 1]
        act_line_item_dict[pcdbk.PC_AREA_ID_KEY_ALIDB] = str(row[area_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COSTING_SECT_ID_KEY_ALIDB] = str(row[costing_sect_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COSTING_SUBSECT_ID_KEY_ALIDB] = str(row[costing_subsect_ID_col - 1])
        act_line_item_dict[pcdbk.PC_ACT_ID_KEY_ALIDB] = str(row[act_ID_col - 1])
        act_line_item_dict[pcdbk.PC_COST_INPUT_ID_KEY_ALIDB] = str(row[cost_input_ID_col - 1])
        act_line_item_dict[pcdbk.PC_NUM_UNITS_KEY_ALIDB] = row[num_units_col - 1]
        act_line_item_dict[pcdbk.PC_DURATION_KEY_ALIDB] = row[duration_col - 1]
        act_line_item_dict[pcdbk.PC_FREQUENCY_KEY_ALIDB] = row[frequency_col - 1]
        activity_line_item_list.append(act_line_item_dict)

    # Write activity input line items

    j = ujson.dumps(activity_line_item_list)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_ACTIVITY_LINE_ITEM_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)
        
    log('Finished PC activity line item DB')

def upload_activity_line_item_DB_PC(version : str = pcc.PC_ACTIVITY_LINE_ITEM_DB_CURR_VERSION):
    JSON_file_name = pcc.PC_ACTIVITY_LINE_ITEM_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC activity line item DB') 

    
            
