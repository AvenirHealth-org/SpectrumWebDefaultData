import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_AREA_TYPE_ID              = '<Area Type ID>'  
PC_AREA_TYPE_NAME            = '<Area Type Name>'   
PC_AREA_ID                   = '<Area ID>'  
PC_AREA_NAME                 = '<Area Name>'   
PC_COSTING_SECT_ID           = '<Costing Section ID>'    
PC_COSTING_SECT_NAME         = '<Costing Section Name>' 
PC_COSTING_SECT_STR_CONST    = '<Costing Section String Constant>'
PC_COSTING_SUBSECT_ID        = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME      = '<Costing Subsection Name>' 
PC_COSTING_SUBSECT_STR_CONST = '<Costing Subsection String Constant>' 
PC_COSTS_CS                  = '<Costs - CS>'

def create_costing_section_DB_PC(version = str):

    log('Creating PC costing section DB')

    costing_sect_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_COSTING_SECT_DB_NAME]
    
    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    area_type_ID_col = gbc.GB_NOT_FOUND
    area_ID_col = gbc.GB_NOT_FOUND
    area_name_col = gbc.GB_NOT_FOUND
    costing_sect_ID_col = gbc.GB_NOT_FOUND
    costing_sect_name_col = gbc.GB_NOT_FOUND
    costing_sect_str_const_col = gbc.GB_NOT_FOUND
    costing_subsect_ID_col = gbc.GB_NOT_FOUND
    costing_subsect_name_col = gbc.GB_NOT_FOUND
    costing_subsect_str_const_col = gbc.GB_NOT_FOUND
    costs_CS_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_AREA_TYPE_ID:
            area_type_ID_col = c

        elif tag == PC_AREA_ID:
            area_ID_col = c

        elif tag == PC_AREA_NAME:
            area_name_col = c

        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c

        elif tag == PC_COSTING_SECT_NAME:
            costing_sect_name_col = c

        elif tag == PC_COSTING_SECT_STR_CONST:
            costing_sect_str_const_col = c

        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c

        elif tag == PC_COSTING_SUBSECT_NAME:
            costing_subsect_name_col = c

        elif tag == PC_COSTING_SUBSECT_STR_CONST:
            costing_subsect_str_const_col = c

        elif tag == PC_COSTS_CS:
            costs_CS_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        costing_sect_dict = {}

        cost_cat_section_name = row[costing_subsect_name_col - 1]
        costs_CS = row[costs_CS_col - 1]

        costing_sect_dict[pcdbk.PC_AREA_TYPE_ID_KEY_CSDB] = row[area_type_ID_col - 1]
        costing_sect_dict[pcdbk.PC_AREA_ID_KEY_CSDB] = row[area_ID_col - 1]
        costing_sect_dict[pcdbk.PC_AREA_NAME_KEY_CSDB] = row[area_name_col - 1]
        costing_sect_dict[pcdbk.PC_COSTING_SECT_ID_KEY_CSDB] = row[costing_sect_ID_col - 1]
        costing_sect_dict[pcdbk.PC_COSTING_SECT_NAME_KEY_CSDB] = row[costing_sect_name_col - 1]
        costing_sect_dict[pcdbk.PC_COSTING_SECT_STR_CONST_KEY_CSDB] = row[costing_sect_str_const_col - 1]
        costing_sect_dict[pcdbk.PC_COSTING_SUBSECT_ID_KEY_CSDB] = row[costing_subsect_ID_col - 1]
        costing_sect_dict[pcdbk.PC_COSTING_SUBSECT_NAME_KEY_CSDB] =  cost_cat_section_name if cost_cat_section_name != None else ''
        costing_sect_dict[pcdbk.PC_COSTING_SUBSECT_STR_CONST_KEY_CSDB] =  row[costing_subsect_str_const_col - 1]
        costing_sect_dict[pcdbk.PC_COSTS_CS_KEY_CSDB] = costs_CS if costs_CS != None else 0.0

        costing_sect_list.append(costing_sect_dict)

    j = ujson.dumps(costing_sect_list)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC)
    JSON_file_name = pcc.PC_COSTING_SECT_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    with open(PC_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished PC costing section DB')

def upload_costing_section_DB_PC(version = str):
    JSON_file_name = pcc.PC_COSTING_SECT_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_PC_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + JSON_file_name)
    log('Uploaded PC costing section DB')
            
