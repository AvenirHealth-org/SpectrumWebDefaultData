import os
import ujson
from itertools import islice
from openpyxl import load_workbook

# from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log
from AvenirCommon.Util import GBRange

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
PC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
PC_COUNTRY_NAME             = '<Country Name>'
PC_AREA_TYPE_ID             = '<Area Type ID>'  
PC_AREA_TYPE_NAME           = '<Area Type Name>'   
PC_AREA_ID                  = '<Area ID>'  
PC_AREA_NAME                = '<Area Name>'   
PC_COSTING_SECT_ID          = '<Costing Section ID>'    
PC_COSTING_SECT_NAME        = '<Costing Section Name>' 
PC_COSTING_SUBSECT_ID       = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME     = '<Costing Subsection Name>' 
PC_PHASE_ID                 = '<Phase ID>'
PC_PHASE_NAME               = '<Phase Name>'

def create_PPLI_phase_DB_PC(version = str):

    log('Creating PPLI phase DB PC')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_PPLI_PHASE_DB_NAME]

    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    iso_alpha_3_country_code_col = gbc.GB_NOT_FOUND
    area_type_ID_col = gbc.GB_NOT_FOUND
    area_ID_col = gbc.GB_NOT_FOUND
    # area_name_col = gbc.GB_NOT_FOUND
    costing_sect_ID_col = gbc.GB_NOT_FOUND
    # costing_sect_name_col = gbc.GB_NOT_FOUND
    # costing_sect_str_const_col = gbc.GB_NOT_FOUND
    costing_subsect_ID_col = gbc.GB_NOT_FOUND
    # costing_subsect_name_col = gbc.GB_NOT_FOUND
    phase_ID_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_ISO_ALPHA_3_COUNTRY_CODE:
            iso_alpha_3_country_code_col = c

        elif tag == PC_AREA_TYPE_ID:
            area_type_ID_col = c

        elif tag == PC_AREA_ID:
            area_ID_col = c

        # elif tag == PC_AREA_NAME:
        #     area_name_col = c

        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c

        # elif tag == PC_COSTING_SECT_NAME:
        #     costing_sect_name_col = c

        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c

        # elif tag == PC_COSTING_SUBSECT_NAME:
        #     costing_subsect_name_col = c

        elif tag == PC_PHASE_ID:
            phase_ID_col = c

    row = 2
    first_data_row = row

    curr_country_alpha_3 = ''

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        row_country_alpha_3 = row[iso_alpha_3_country_code_col - 1]

        # If we're on a new country, initialize a new country dict and add it to the country list
        if curr_country_alpha_3 != row_country_alpha_3:
            country_dict = {}
            country_dict[pcdbk.PC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PPLIPDB] = row[iso_alpha_3_country_code_col - 1]
            country_dict[pcdbk.PC_PHASE_INTENSITY_RECORDS] = []
            country_list.append(country_dict)
            curr_country_alpha_3 = row_country_alpha_3

        phase_intensity_rec = {
            pcdbk.PC_AREA_TYPE_ID_KEY_PPLIPDB           : row[area_type_ID_col - 1],
            pcdbk.PC_AREA_ID_KEY_PPLIPDB                : row[area_ID_col - 1],
            pcdbk.PC_COSTING_SECT_ID_KEY_PPLIPDB        : row[costing_sect_ID_col - 1],
            pcdbk.PC_COSTING_SUBSECT_ID_KEY_PPLIPDB     : row[costing_subsect_ID_col - 1],
            pcdbk.PC_PHASE_ID_KEY_PPLIPDB               : row[phase_ID_col - 1],
        }
        country_dict[pcdbk.PC_PHASE_INTENSITY_RECORDS].append(phase_intensity_rec)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + pcc.PC_PPLI_PHASES_DB_DIR
    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[pcdbk.PC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PPLIPDB]
        with open(PC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished PPLI phase DB PC')

def upload_PPLI_phase_DB_PC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + pcc.PC_PPLI_PHASES_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_PC_CONTAINER, walk_path, version, pcc.PC_PPLI_PHASES_DB_DIR + '\\')
    log('Uploaded PPLI phase DB PC')


