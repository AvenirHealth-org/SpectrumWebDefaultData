import os
import ujson
from itertools import islice
from openpyxl import load_workbook

# from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log
from AvenirCommon.Util import GBRange

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.PC.PCConst as pcc
import SpectrumCommon.Const.PC.PCDatabaseKeys as pcdbk 

# Column tag 'constants' for identifying columns.

PC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
PC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
PC_COUNTRY_NAME             = '<Country Name>'
PC_PRICE_ELASTICITY_PARAMS  = '<Price Elasticity Parameters'

def create_PPLI_taxation_DB_PC(version = str):

    log('Creating PPLI taxation DB PC')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_PC) + '\PCModData.xlsx')
    sheet = wb[pcc.PC_PPLI_TAXATION_DB_NAME]

    # first row of intervention data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    iso_alpha_3_country_code_col = gbc.GB_NOT_FOUND
    price_elasticity_params_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == PC_ISO_ALPHA_3_COUNTRY_CODE:
            iso_alpha_3_country_code_col = c

        elif tag == PC_PRICE_ELASTICITY_PARAMS:
            price_elasticity_params_col = c

    row = 2
    first_data_row = row

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}
        country_dict[pcdbk.PC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PPLITDB] = row[iso_alpha_3_country_code_col - 1]
        country_dict[pcdbk.PC_PRICE_ELASTICITY_PARAMS_KEY_PPLITDB] = row[price_elasticity_params_col - 1]
        country_list.append(country_dict)

    PC_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + pcc.PC_PPLI_TAXATION_DB_DIR
    os.makedirs(PC_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[pcdbk.PC_ISO_ALPHA_3_COUNTRY_CODE_KEY_PPLITDB]
        with open(PC_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished PPLI taxation DB PC')

def upload_PPLI_taxation_DB_PC(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_PC) + '\\' + pcc.PC_PPLI_TAXATION_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_PC_CONTAINER, walk_path, version, pcc.PC_PPLI_TAXATION_DB_DIR + '\\')
    log('Uploaded PPLI taxation DB PC')


