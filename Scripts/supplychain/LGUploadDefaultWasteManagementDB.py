import os
import SpectrumCommon.Const.LG.LGConst as lgc
import SpectrumCommon.Const.GB.GBConst as gbc
from AvenirCommon.Database import GB_upload_file
import DefaultDataUtil as ddu
from AvenirCommon.Logger import log
from openpyxl import load_workbook
from AvenirCommon.Util import GBRange
from itertools import islice
import SpectrumCommon.Const.LG.LGDatabaseKeys as lgdbk
import ujson

# Column tag 'constants' for identifying columns.

LG_WASTE_MANAGEMENT_ID = "<Waste Management ID>"
LG_NAME = "<Name>"
LG_IS_DEFAULT = "<Is Default>"
LG_UNIT_COST = "<Unit Cost>"
LG_WORKING_LIFE = "<Working Life>"
LG_MAINT_PERC = "<Maint Perc>"
LG_ELECTRICITY_COST = "<Electricity Cost>"
LG_WATER_COST = "<Water Cost>"
LG_FUEL_COST = "<Fuel Cost>"


def create_default_waste_management_DB_LG(version: str):
    log("Creating LG default waste management DB")

    default_waste_management_records = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_LG) + "\LGModData.xlsx")
    sheet = wb[lgc.LG_WASTE_MANAGEMENT_DB_NAME]

    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    waste_management_id_col = gbc.GB_NOT_FOUND
    name_col = gbc.GB_NOT_FOUND
    is_default_col = gbc.GB_NOT_FOUND
    unit_cost_col = gbc.GB_NOT_FOUND
    working_life_col = gbc.GB_NOT_FOUND
    maint_perc_col = gbc.GB_NOT_FOUND
    electricity_cost_col = gbc.GB_NOT_FOUND
    water_cost_col = gbc.GB_NOT_FOUND
    fuel_cost_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == LG_WASTE_MANAGEMENT_ID:
            waste_management_id_col = c

        elif tag == LG_NAME:
            name_col = c

        elif tag == LG_IS_DEFAULT:
            is_default_col = c

        elif tag == LG_UNIT_COST:
            unit_cost_col = c

        elif tag == LG_WORKING_LIFE:
            working_life_col = c

        elif tag == LG_MAINT_PERC:
            maint_perc_col = c

        elif tag == LG_ELECTRICITY_COST:
            electricity_cost_col = c

        elif tag == LG_WATER_COST:
            water_cost_col = c

        elif tag == LG_FUEL_COST:
            fuel_cost_col = c

    for row in islice(sheet.values, first_data_row - 1, num_rows):
        waste_management_rec_dict = {}

        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_ID_KEY] = str(row[waste_management_id_col - 1])
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_NAME_KEY] = str(row[name_col - 1])
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_IS_DEFAULT_KEY] = row[is_default_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_UNIT_COST_KEY] = row[unit_cost_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_WORKING_LIFE_KEY] = row[working_life_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_MAINT_COST_PERC_KEY] = row[maint_perc_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_ELECTRICITY_COST_KEY] = row[electricity_cost_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_WATER_COST_KEY] = row[water_cost_col - 1]
        waste_management_rec_dict[lgdbk.LG_WASTE_MANAGEMENT_FUEL_COST_KEY] = row[fuel_cost_col - 1]

        default_waste_management_records.append(waste_management_rec_dict)

    j = ujson.dumps(default_waste_management_records)

    LG_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_LG)
    JSON_file_name = lgc.LG_WASTE_MANAGEMENT_DB_NAME + "_" + version + "." + gbc.GB_JSON

    os.makedirs(LG_JSON_data_path + "\\", exist_ok=True)
    with open(LG_JSON_data_path + "\\" + JSON_file_name, "w") as f:
        f.write(j)
    log("Finished default waste management DB")


def upload_default_waste_management_DB_LG(version: str):
    JSON_file_name = lgc.LG_WASTE_MANAGEMENT_DB_NAME + "_" + version + "." + gbc.GB_JSON
    GB_upload_file(
        os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV],
        gbc.GB_LG_CONTAINER,
        JSON_file_name,
        ddu.get_JSON_data_path(gbc.GB_LG) + "\\" + JSON_file_name,
    )
    log("Uploaded LG default waste management DB")
