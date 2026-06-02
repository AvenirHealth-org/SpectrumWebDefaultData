import os
import SpectrumCommon.Const.LG.LGConst as lgc
import SpectrumCommon.Const.GB.GBConst as gbc
from AvenirCommon.Database import GB_upload_file
import DefaultDataUtil as ddu
from AvenirCommon.Logger import log
from openpyxl import load_workbook
from AvenirCommon.Util import GBRange
from itertools import islice
import SpectrumCommon.Const.LG.LGDatabaseKeys as lgdbk
import ujson

# Column tag 'constants' for identifying columns.

LG_COLD_CHAIN_ID = "<Cold Chain ID>"
LG_COLD_CHAIN_TYPE = "<Cold Chain Type ID>"
LG_NAME = "<Name>"
LG_DESCRIPTION = "<Description>"
LG_UNIT_COST = "<Unit Cost>"
LG_CAPACITY = "<Capacity"


def create_default_cold_chain_DB_LG(version: str):
    log("Creating LG default cold chain DB")

    default_cold_chain_records = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_LG) + "\LGModData.xlsx")
    sheet = wb[lgc.LG_COLD_CHAIN_DB_NAME]

    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    cold_chain_ID_col = gbc.GB_NOT_FOUND
    cold_chain_type_col = gbc.GB_NOT_FOUND
    name_col = gbc.GB_NOT_FOUND
    # description_col = gbc.GB_NOT_FOUND
    unit_cost_col = gbc.GB_NOT_FOUND
    capacity_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == LG_COLD_CHAIN_ID:
            cold_chain_ID_col = c

        elif tag == LG_COLD_CHAIN_TYPE:
            cold_chain_type_col = c

        elif tag == LG_NAME:
            name_col = c

        # elif tag == LG_DESCRIPTION:
        #     description_col = c

        elif tag == LG_UNIT_COST:
            unit_cost_col = c

        elif tag == LG_CAPACITY:
            capacity_col = c

    for row in islice(sheet.values, first_data_row - 1, num_rows):
        cold_chain_rec_dict = {}

        cold_chain_rec_dict[lgdbk.LG_COLD_CHAIN_ID_KEY] = str(row[cold_chain_ID_col - 1])
        cold_chain_rec_dict[lgdbk.LG_COLD_CHAIN_TYPE_KEY] = row[cold_chain_type_col - 1]
        cold_chain_rec_dict[lgdbk.LG_COLD_CHAIN_NAME_KEY] = str(row[name_col - 1])
        cold_chain_rec_dict[lgdbk.LG_COLD_CHAIN_UNIT_COST_KEY] = row[unit_cost_col - 1]
        cold_chain_rec_dict[lgdbk.LG_COLD_CHAIN_CAPACITY_KEY] = row[capacity_col - 1]

        default_cold_chain_records.append(cold_chain_rec_dict)

    j = ujson.dumps(default_cold_chain_records)

    LG_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_LG)
    JSON_file_name = lgc.LG_COLD_CHAIN_DB_NAME + "_" + version + "." + gbc.GB_JSON

    os.makedirs(LG_JSON_data_path + "\\", exist_ok=True)
    with open(LG_JSON_data_path + "\\" + JSON_file_name, "w") as f:
        f.write(j)
    log("Finished default cold chain DB")


def upload_default_cold_chain_DB_LG(version: str):
    JSON_file_name = lgc.LG_COLD_CHAIN_DB_NAME + "_" + version + "." + gbc.GB_JSON
    GB_upload_file(
        os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV],
        gbc.GB_LG_CONTAINER,
        JSON_file_name,
        ddu.get_JSON_data_path(gbc.GB_LG) + "\\" + JSON_file_name,
    )
    log("Uploaded LG default cold chain DB")
