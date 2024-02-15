import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.HW.HWConst as hwc
import SpectrumCommon.Const.HW.HWDatabaseConst as hwdbc 

# Column tag 'constants' for identifying columns.

IC_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
IC_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>' 
IC_COUNTRY_NAME             = '<Country Name>'
IC_EDU_CODE_1               = '<Educational Code 1>'
IC_EDU_CODE_2               = '<Educational Code 2>'
IC_EDU_CODE_3               = '<Educational Code 3>'
IC_EDU_CODE_4               = '<Educational Code 4>'
IC_EDU_CODE_5               = '<Educational Code 5>'

def create_staff_salaries_DB_HW(version = str):

    log('Creating HW staff salaries DB')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_HW) + '\HWModData.xlsx')
    sheet = wb[hwc.HW_STAFF_SALARIES_DB_NAME]
    
    num_rows = sheet.max_row
    #num_cols = sheet.max_column

    # Establish tag columns.  Not many, so don't bother searching for them.
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    edu_code_1_salary_col = col
    col += 1
    edu_code_2_salary_col = col
    col += 1
    edu_code_3_salary_col = col
    col += 1
    edu_code_4_salary_col = col
    col += 1
    edu_code_5_salary_col = col

    # Establish tag rows. Not many, so don't bother searching for them.
    row = 2
    first_data_row = row

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):
        country_dict = {}

        iso3 = row[iso_alpha_3_country_code_col - 1]

        if iso3 != None:
            country_dict[hwdbc.HW_ISO_ALPHA_3_COUNTRY_CODE_KEY_SSDB] = iso3
            country_dict[hwdbc.HW_EDU_CODE_1_SALARY_KEY_SSDB] = row[edu_code_1_salary_col - 1]
            country_dict[hwdbc.HW_EDU_CODE_2_SALARY_KEY_SSDB] = row[edu_code_2_salary_col - 1]
            country_dict[hwdbc.HW_EDU_CODE_3_SALARY_KEY_SSDB] = row[edu_code_3_salary_col - 1]
            country_dict[hwdbc.HW_EDU_CODE_4_SALARY_KEY_SSDB] = row[edu_code_4_salary_col - 1]
            country_dict[hwdbc.HW_EDU_CODE_5_SALARY_KEY_SSDB] = row[edu_code_5_salary_col - 1]
            country_list.append(country_dict)

    HW_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_HW) + '\\' + hwc.HW_STAFF_SALARIES_DB_DIR
    os.makedirs(HW_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[hwdbc.HW_ISO_ALPHA_3_COUNTRY_CODE_KEY_SSDB]
        with open(HW_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished HW staff salaries DB')

def upload_staff_salaries_DB_HW(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_HW) + '\\' + hwc.HW_STAFF_SALARIES_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_HW_CONTAINER, walk_path, version, hwc.HW_STAFF_SALARIES_DB_DIR + '\\')
    log('Uploaded HW staff salaries DB')
