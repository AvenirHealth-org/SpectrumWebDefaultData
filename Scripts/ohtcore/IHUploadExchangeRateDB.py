import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc
import SpectrumCommon.Const.IH.IHConst as ihc
import SpectrumCommon.Const.IH.IHDatabaseKeys as ihdbk 

# Column tag 'constants' for identifying columns.

IH_ISO_NUMERIC_COUNTRY_CODE = '<ISO Numeric Country Code>'  
IH_ISO_ALPHA_3_COUNTRY_CODE = '<ISO Alpha-3 Country Code>'
IH_COUNTRY_NAME             = '<Country Name>'
IH_DATA                     = '<Data>'
IH_YEARS                    = '<Years>'

def create_exchange_rate_DB_IH(version = str):

    log('Creating Exchange rate IH')

    country_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IH) + '\IHModData.xlsx')
    sheet = wb[ihc.IH_EXCHANGE_RATE_DB_NAME]

    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # Establish tag columns.
    col = 1
    #iso_numeric_country_code_col = col
    col += 1    
    #country_name_col = col
    col += 1
    iso_alpha_3_country_code_col = col
    col += 1
    data_col = col

    row = 3
    first_data_row = row

    # Get each row from the sheet and create a dictionary for each each country
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        country_dict = {}
        country_dict[ihdbk.IH_ISO_ALPHA_3_COUNTRY_CODE_KEY_EXDB] = row[iso_alpha_3_country_code_col - 1]
        country_dict[ihdbk.IH_EXCHANGE_RATE_KEY_EXDB] = row[data_col - 1]

        country_list.append(country_dict)

    IH_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IH) + '\\' + ihc.IH_EXCHANGE_RATE_DB_DIR
    os.makedirs(IH_JSON_data_path + '\\', exist_ok = True)
    
    for country_obj in country_list:    
        iso3 = country_obj[ihdbk.IH_ISO_ALPHA_3_COUNTRY_CODE_KEY_EXDB]
        with open(IH_JSON_data_path + '\\' + iso3 + '_' + version + '.' + gbc.GB_JSON, 'w') as f:
            ujson.dump(country_obj, f)

    log('Finished Exchange rate IH')

def upload_exchange_rate_DB_IH(version):
    walk_path = ddu.get_JSON_data_path(gbc.GB_IH) + '\\' + ihc.IH_EXCHANGE_RATE_DB_DIR + '\\'
    ddu.uploadFilesInDir(gbc.GB_IH_CORE_CONTAINER, walk_path, version)
    log('Uploaded Exchange rate IH')


