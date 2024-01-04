import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.BG.BGConst as bgc
import SpectrumCommon.Const.BG.BGDatabaseKeys as bgdbk 

# Column tag 'constants' for identifying columns.

BG_COSTING_MODE_NAMES     = '<Costing Mode Names>'
BG_COSTING_MODES          = '<Costing Modes>'  
BG_BUDGET_NAME_ENGLISH    = '<Budget Name - English>'
BG_BUDGET_NAME_STR_CONST  = '<Budget String Constant>'
BG_BUDGET_ID              = '<Budget ID>'

def create_budget_DB_BG(version_str):

    log('Creating BG budget DB')
    BG_path = os.getcwd() + '\Tools\DefaultDataManager\BG\\'
    BG_mod_data_FQ_name = BG_path + 'ModData\BGModData.xlsx'

    budget_list = []
    wb = load_workbook(BG_mod_data_FQ_name)
    sheet = wb[bgc.BG_BUDGET_DB_NAME]
    
    '''
        IMPORTANT: Max target population master ID:338

        New target populations must have a master ID greater than this. Please update this maximum if you add a new one.

    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # costing_mode_names_col = gbc.GB_NOT_FOUND
    costing_modes_col = gbc.GB_NOT_FOUND
    budget_name_english_col = gbc.GB_NOT_FOUND
    budget_name_str_const_col = gbc.GB_NOT_FOUND
    budget_ID_col =  gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # if tag == BG_COSTING_MODE_NAMES:
        #     costing_mode_names_col = c

        if tag == BG_COSTING_MODES:
            costing_modes_col = c

        elif tag == BG_BUDGET_NAME_ENGLISH:
            budget_name_english_col = c

        elif tag == BG_BUDGET_NAME_STR_CONST:
            budget_name_str_const_col = c

        elif tag == BG_BUDGET_ID:
            budget_ID_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        budget_dict = {}

        budget_dict[bgdbk.BG_COSTING_MODES_KEY_BDB] = str(row[costing_modes_col - 1]).split(', ')
        budget_dict[bgdbk.BG_ENGLISH_STRING_KEY_BDB] = row[budget_name_english_col - 1]
        budget_dict[bgdbk.BG_STRING_CONSTANT_KEY_BDB] = row[budget_name_str_const_col - 1]
        budget_dict[bgdbk.BG_ID_KEY_BDB] = row[budget_ID_col - 1]

        budget_list.append(budget_dict)

    j = ujson.dumps(budget_list)

    os.makedirs(BG_path + 'JSON\\', exist_ok = True)
    with open(BG_path + 'JSON\\' + bgc.BG_BUDGET_DB_NAME + '_' + version_str + '.' + gbc.GB_JSON, 'w') as f:
        f.write(j)
    log('Finished BG budget DB')

def upload_budget_DB_BG(version):
    connection =  os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV]
    FQName = os.getcwd() + '\Tools\DefaultDataManager\BG\JSON\\' + bgc.BG_BUDGET_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    container_name = gbc.GB_BG_CONTAINER
    GB_upload_file(connection, container_name, bgc.BG_BUDGET_DB_NAME + '_' + version + '.' + gbc.GB_JSON, FQName) 
    log('Uploaded BG budget DB')
            
