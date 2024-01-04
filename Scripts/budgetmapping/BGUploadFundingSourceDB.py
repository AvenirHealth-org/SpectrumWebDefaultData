import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.BG.BGConst as bgc
import SpectrumCommon.Const.BG.BGDatabaseKeys as bgdbk 

# Column tag 'constants' for identifying columns.

BG_COSTING_MODE_NAMES          = '<Costing Mode Names>'
BG_COSTING_MODES               = '<Costing Modes>'  
BG_FUND_FW_NAME                = '<Funding Framework Name>'
BG_FUND_FW_ID                  = '<Funding Framework ID>'
BG_FUND_SOURCE_NAME_ENGLISH    = '<Funding Source Name - English>'
BG_FUND_SOURCE_NAME_STR_CONST  = '<Funding Source String Constant>'
BG_FUND_SOURCE_ID              = '<Funding Source ID>'

def create_funding_source_DB_BG(version = str):

    log('Creating BG funding source DB')

    fund_source_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_BG) + '\BGModData.xlsx')
    sheet = wb[bgc.BG_FUND_SOURCE_DB_NAME]
    
    '''
        IMPORTANT: Max target population master ID:338

        New target populations must have a master ID greater than this. Please update this maximum if you add a new one.

    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    # costing_mode_names_col = gbc.GB_NOT_FOUND
    costing_modes_col = gbc.GB_NOT_FOUND
    # fund_framework_name_col = gbc.GB_NOT_FOUND
    fund_framework_ID_col = gbc.GB_NOT_FOUND
    fund_source_name_english_col = gbc.GB_NOT_FOUND
    fund_source_name_str_const_col = gbc.GB_NOT_FOUND
    fund_source_ID_col =  gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        # if tag == BG_COSTING_MODE_NAMES:
        #     costing_mode_names_col = c

        if tag == BG_COSTING_MODES:
            costing_modes_col = c

        elif tag == BG_FUND_FW_ID:
            fund_framework_ID_col = c

        elif tag == BG_FUND_SOURCE_NAME_ENGLISH:
            fund_source_name_english_col = c

        elif tag == BG_FUND_SOURCE_NAME_STR_CONST:
            fund_source_name_str_const_col = c

        elif tag == BG_FUND_SOURCE_ID:
            fund_source_ID_col = c

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        fund_source_dict = {}

        fund_source_dict[bgdbk.BG_COSTING_MODES_KEY_FSDB] = str(row[costing_modes_col - 1]).split(', ')
        fund_source_dict[bgdbk.BG_FUND_FW_ID_KEY_FSDB] = row[fund_framework_ID_col - 1]
        fund_source_dict[bgdbk.BG_ENGLISH_STRING_KEY_FSDB] = row[fund_source_name_english_col - 1]
        fund_source_dict[bgdbk.BG_STRING_CONSTANT_KEY_FSDB] = row[fund_source_name_str_const_col - 1]
        fund_source_dict[bgdbk.BG_ID_KEY_FSDB] = row[fund_source_ID_col - 1]

        fund_source_list.append(fund_source_dict)

    j = ujson.dumps(fund_source_list)

    BG_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_BG)
    JSON_file_name = bgc.BG_FUND_SOURCE_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(BG_JSON_data_path + '\\', exist_ok = True)
    with open(BG_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished BG funding source DB')

def upload_funding_source_DB_BG(version):
    JSON_file_name = bgc.BG_FUND_SOURCE_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_BG_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_BG) + '\\' + JSON_file_name) 
    log('Uploaded BG funding source DB')
            
