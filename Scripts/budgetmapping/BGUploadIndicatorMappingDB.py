import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.BG.BGConst as bgc
import SpectrumCommon.Const.BG.BGDatabaseKeys as bgdbk 

# Column tag 'constants' for identifying columns.

BG_COSTING_MODES             = '<Costing Modes>'  
BG_COSTING_MODE_NAMES        = '<Costing Mode Names>'
BG_BUDGET_ID                 = '<Budget ID>'
BG_BUDGET_NAME               = '<Budget Name>'
BG_MODULE_ID                 = '<Module ID>'
BG_MODULE_NAME               = '<Module Name>'
BG_INDICATOR_TYPE_ID         = '<Indicator Type ID>'
BG_INDICATOR_TYPE_NAME       = '<Indicator Type Name>'
BG_INDICATOR_ID              = '<Indicator ID>'
BG_INDICATOR_NAME            = '<Indicator Name>'
PC_COSTING_SECT_ID           = '<Costing Section ID>'    
PC_COSTING_SECT_NAME         = '<Costing Section Name>' 
PC_COSTING_SUBSECT_ID        = '<Costing Subsection ID>'       
PC_COSTING_SUBSECT_NAME      = '<Costing Subsection Name>' 
PC_ACT_ID                    = '<Activity ID>' 
PC_ACT_NAME                  = '<Activity Name>'
BG_BUDGET_CAT_ID             = '<Budget Category ID>' 
BG_BUDGET_CAT_NAME_ENGLISH   = '<Budget Category Name>'

def create_indicator_mapping_DB_BG(version = str):

    log('Creating BG indicator mapping DB')

    ind_mapping_list = []
    wb = load_workbook(ddu.get_source_data_path(gbc.GB_BG) + '\BGModData.xlsx')
    sheet = wb[bgc.BG_INDICATOR_MAPPING_DB_NAME]

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    costing_modes_col = gbc.GB_NOT_FOUND
    # costing_mode_names_col = gbc.GB_NOT_FOUND

    budget_ID_col =  gbc.GB_NOT_FOUND
    # budget_name_col = gbc.GB_NOT_FOUND

    mod_ID_col =  gbc.GB_NOT_FOUND
    #mod_name_col = gbc.GB_NOT_FOUND

    ind_type_ID_col =  gbc.GB_NOT_FOUND
    #ind_type_name_col = gbc.GB_NOT_FOUND

    ind_ID_col =  gbc.GB_NOT_FOUND
    #ind_name_col = gbc.GB_NOT_FOUND

    costing_sect_ID_col =  gbc.GB_NOT_FOUND
    #costing_sect_name_col = gbc.GB_NOT_FOUND

    costing_subsect_ID_col =  gbc.GB_NOT_FOUND
    #costing_subsect_name_col = gbc.GB_NOT_FOUND

    act_ID_col =  gbc.GB_NOT_FOUND
    #act_name_col = gbc.GB_NOT_FOUND

    budget_cat_ID_col =  gbc.GB_NOT_FOUND
    #budget_cat_name_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == BG_COSTING_MODES:
            costing_modes_col = c

        # elif tag == BG_COSTING_MODE_NAMES:
        #     costing_mode_names_col = c
            
        elif tag == BG_BUDGET_ID:
            budget_ID_col = c
            
        # elif tag == BG_BUDGET_NAME:
        #     budget_name_col = c
            
        elif tag == BG_MODULE_ID:
            mod_ID_col = c
            
        # elif tag == BG_MODULE_NAME:
        #     mod_name_col = c
            
        elif tag == BG_INDICATOR_TYPE_ID:
            ind_type_ID_col = c
            
        # elif tag == BG_INDICATOR_TYPE_NAME:
        #     ind_type_name_col = c
            
        elif tag == BG_INDICATOR_ID:
            ind_ID_col = c
            
        # elif tag == BG_INDICATOR_NAME:
        #     ind_name_col = c
            
        elif tag == PC_COSTING_SECT_ID:
            costing_sect_ID_col = c
              
        # elif tag == PC_COSTING_SECT_NAME:
        #     costing_sect_name_col = c
            
        elif tag == PC_COSTING_SUBSECT_ID:
            costing_subsect_ID_col = c
                 
        # elif tag == PC_COSTING_SUBSECT_NAME:
        #     costing_subsect_name_col = c
            
        elif tag == PC_ACT_ID:
            act_ID_col = c
            
        # elif tag == PC_ACT_NAME:
        #     act_name_col = c
            
        elif tag == BG_BUDGET_CAT_ID:
            budget_cat_ID_col = c
            
        # elif tag == BG_BUDGET_CAT_NAME_ENGLISH:
        #     budget_cat_name_col = c
            
    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        ind_mapping_dict = {}

        ind_mapping_dict[bgdbk.BG_COSTING_MODES_KEY_IMDB] = str(row[costing_modes_col - 1]).split(', ')
        ind_mapping_dict[bgdbk.BG_BUDGET_ID_KEY_IMDB] = row[budget_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_MOD_ID_KEY_IMDB] = row[mod_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_IND_TYPE_IDS_KEY_IMDB] = str(row[ind_type_ID_col - 1]).split(', ')
        ind_mapping_dict[bgdbk.BG_IND_ID_KEY_IMDB] = row[ind_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_COSTING_SECT_ID_KEY_IMDB] = row[costing_sect_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_COSTING_SUBSECT_ID_KEY_IMDB] = row[costing_subsect_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_ACTIVITY_ID_KEY_IMDB] = row[act_ID_col - 1]
        ind_mapping_dict[bgdbk.BG_BUDGET_CAT_ID_KEY_IMDB] = row[budget_cat_ID_col - 1]

        ind_mapping_list.append(ind_mapping_dict)

    j = ujson.dumps(ind_mapping_list)

    BG_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_BG)
    JSON_file_name = bgc.BG_INDICATOR_MAPPING_DB_NAME + '_' + version + '.' + gbc.GB_JSON

    os.makedirs(BG_JSON_data_path + '\\', exist_ok = True)
    with open(BG_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    log('Finished BG indicator mapping DB')

def upload_indicator_mapping_DB_BG(version):
    JSON_file_name = bgc.BG_INDICATOR_MAPPING_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_BG_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_BG) + '\\' + JSON_file_name) 
    log('Uploaded BG indicator mapping DB')
            
