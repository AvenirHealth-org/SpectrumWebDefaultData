import os
import ujson
from itertools import islice
from openpyxl import load_workbook

from AvenirCommon.Util import GBRange
from AvenirCommon.Database import GB_upload_file
from AvenirCommon.Logger import log

import DefaultData.DefaultDataUtil as ddu

import SpectrumCommon.Const.GB as gbc

import SpectrumCommon.Const.IS.ISConst as isc
import SpectrumCommon.Const.IS.ISDatabaseKeys as isdbk 

# Column tag 'constants' for identifying columns.

IS_MOD_ID                     = '<Module ID>'
IS_MOD_NAME                   = '<Module Name>'
IS_FACILITY_TYPE_MST_ID       = '<Facility Type Master ID>'
IS_FACILITY_TYPE_NAME         = '<Facility Type Name>'
IS_GROUP_MST_ID               = '<Group Master ID>'
IS_GROUP_NAME                 = '<Group Name>'
IS_GROUP_STR_CONST            = '<Group String Constant>'
IS_FACILITY_EQUIP_TYPE_MST_ID = '<Facility Equipment Type Master ID>'
IS_EQUIP_MST_ID               = '<Equipment Master ID>'
IS_EQUIP_NAME                 = '<Equipment Name>'
IS_EQUIP_STR_CONST            = '<Equipment String Constant>'
IS_UNIT_COST                  = '<Unit Cost>'
IS_UNITS_PER_FACILITY_TYPE    = '<Units per Facility Type>'

def create_facility_equipment_and_group_DBs_IS(version = str):

    log('Creating Facility Equipment DB, IS')

    facility_equip_list = []
    facility_med_equip_group_list = []

    wb = load_workbook(ddu.get_source_data_path(gbc.GB_IS) + '\ISModData.xlsx')
    sheet = wb[isc.IS_FACILITY_EQUIP_DB_NAME]

    '''
        <Notes>
        
            New groups within a particular facility type must be given a master ID that is greater than any existing 
            group master ID within those confines.

            1111 is the max equipment ID. New equipment must have a larger ID than this. Equipment IDs do not reset with groups, so no two
            equipment IDs will be the same across any modules, facilities, groups, etc.

    '''

    # first row of data after col descriptions and col tags
    tag_row = 1
    first_data_row = 2
    first_data_col = 1
    num_rows = sheet.max_row
    num_cols = sheet.max_column

    mod_ID_col = gbc.GB_NOT_FOUND
    #mod_name_col = gbc.GB_NOT_FOUND
    facility_type_mst_ID_col = gbc.GB_NOT_FOUND
    #facility_type_name_col = gbc.GB_NOT_FOUND
    group_mst_ID_col = gbc.GB_NOT_FOUND
    group_name_col = gbc.GB_NOT_FOUND
    group_str_const_col =  gbc.GB_NOT_FOUND
    facility_equip_type_mst_ID_col = gbc.GB_NOT_FOUND
    equipment_mst_ID_col = gbc.GB_NOT_FOUND
    equipment_name_col = gbc.GB_NOT_FOUND
    equipment_str_const_col = gbc.GB_NOT_FOUND
    unit_cost_col = gbc.GB_NOT_FOUND
    units_per_facility_col = gbc.GB_NOT_FOUND

    for c in GBRange(first_data_col, num_cols):
        tag = sheet.cell(tag_row, c).value

        if tag == IS_MOD_ID:
            mod_ID_col = c

        elif tag == IS_FACILITY_TYPE_MST_ID:
            facility_type_mst_ID_col = c

        #elif tag == IS_FACILITY_TYPE_NAME_ID:
        #    facility_type_name_col = c

        elif tag == IS_GROUP_MST_ID:
            group_mst_ID_col = c

        elif tag == IS_GROUP_NAME:
            group_name_col = c

        elif tag == IS_GROUP_STR_CONST:
            group_str_const_col = c

        elif tag == IS_FACILITY_EQUIP_TYPE_MST_ID:
            facility_equip_type_mst_ID_col = c

        elif tag == IS_EQUIP_MST_ID:
            equipment_mst_ID_col = c

        elif tag == IS_EQUIP_NAME:
            equipment_name_col = c

        elif tag == IS_EQUIP_STR_CONST:
            equipment_str_const_col = c

        elif tag == IS_UNIT_COST:
            unit_cost_col = c

        elif tag == IS_UNITS_PER_FACILITY_TYPE:
            units_per_facility_col = c

    current_mod_ID = 0
    current_facility_type_mst_ID = 0
    currrent_group_mst_ID = 0

    # Get each row from the sheet and create a dictionary for each each group
    for row in islice(sheet.values, first_data_row - 1, num_rows):

        mod_ID = row[mod_ID_col - 1]
        facility_type_mst_ID = row[facility_type_mst_ID_col - 1]
        #facility_type_name = row[facility_type_name_col - 1]
        group_mst_ID = row[group_mst_ID_col - 1]
        group_name = row[group_name_col - 1]
        group_str_const = row[group_str_const_col - 1]
        facility_equip_type_mst_ID = row[facility_equip_type_mst_ID_col - 1]
        equipment_mst_ID = row[equipment_mst_ID_col - 1]
        equipment_name = row[equipment_name_col - 1]
        equipment_str_const = row[equipment_str_const_col - 1]
        unit_cost = row[unit_cost_col - 1]
        units_per_facility = row[units_per_facility_col - 1]

        equipment_dict = {}
        equipment_dict[isdbk.IS_MOD_ID_KEY_FEIDB] = mod_ID
        equipment_dict[isdbk.IS_FACILITY_TYPE_MST_ID_KEY_FEIDB] = str(facility_type_mst_ID)
        equipment_dict[isdbk.IS_GROUP_MST_ID_KEY_FEIDB] = str(group_mst_ID)
        equipment_dict[isdbk.IS_FACILITY_EQUIP_TYPE_MST_ID_KEY_FEIDB] = facility_equip_type_mst_ID
        equipment_dict[isdbk.IS_EQUIP_MST_ID_KEY_FEIDB] = str(equipment_mst_ID)
        equipment_dict[isdbk.IS_EQUIP_NAME_MST_ID_KEY_FEIDB] = equipment_name
        equipment_dict[isdbk.IS_EQUIP_STR_CONST_KEY_FEIDB] = equipment_str_const
        equipment_dict[isdbk.IS_UNIT_COST_KEY_FEIDB] = unit_cost
        equipment_dict[isdbk.IS_UNITS_PER_FACILITY_TYPE_KEY_FEIDB] = units_per_facility
        facility_equip_list.append(equipment_dict)

        # Create a new group dict if the equipment type is medical equipment and the module, facility type, or group has changed.
        if (facility_equip_type_mst_ID == isc.IS_FACIL_MEDICAL_EQUIP_MST_ID) and \
            (not ((current_mod_ID == mod_ID) and (current_facility_type_mst_ID == facility_type_mst_ID) and (currrent_group_mst_ID == group_mst_ID))):
                facility_med_equip_group_dict = {}
                facility_med_equip_group_dict[isdbk.IS_MOD_ID_KEY_FEIDB] = mod_ID
                facility_med_equip_group_dict[isdbk.IS_FACILITY_TYPE_MST_ID_KEY_FEIDB] = str(facility_type_mst_ID)
                facility_med_equip_group_dict[isdbk.IS_GROUP_MST_ID_KEY_FEIDB] = str(group_mst_ID)
                facility_med_equip_group_dict[isdbk.IS_GROUP_NAME_ID_KEY_FEIDB] = group_name
                facility_med_equip_group_dict[isdbk.IS_GROUP_STR_CONST_KEY_FEIDB] = group_str_const
                facility_med_equip_group_list.append(facility_med_equip_group_dict)
        
        current_mod_ID = mod_ID
        current_facility_type_mst_ID = facility_type_mst_ID
        currrent_group_mst_ID = group_mst_ID


    IS_JSON_data_path = ddu.get_JSON_data_path(gbc.GB_IS)
    os.makedirs(IS_JSON_data_path + '\\', exist_ok = True)

    # Write facility equipment

    j = ujson.dumps(facility_equip_list)
    # for debugging
    # for i, obj in enumerate(facility_equip_list):
    #     log(obj)
    JSON_file_name = isc.IS_FACILITY_EQUIP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    with open(IS_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)

    # Write facility medical equipment groups

    j = ujson.dumps(facility_med_equip_group_list)
    JSON_file_name = isc.IS_FACILITY_MED_EQUIP_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    with open(IS_JSON_data_path + '\\' + JSON_file_name, 'w') as f:
        f.write(j)
        
    log('Finished Facility Equipment DB, IS')

def upload_facility_equipment_and_group_DBs_IS(version):
    JSON_file_name = isc.IS_FACILITY_EQUIP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_IS_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_IS) + '\\' + JSON_file_name) 
    
    JSON_file_name = isc.IS_FACILITY_MED_EQUIP_GROUP_DB_NAME + '_' + version + '.' + gbc.GB_JSON
    GB_upload_file(os.environ[gbc.GB_SPECT_MOD_DATA_CONN_ENV], gbc.GB_IS_CONTAINER, JSON_file_name, ddu.get_JSON_data_path(gbc.GB_IS) + '\\' + JSON_file_name) 

    log('Uploaded Facility Equipment DB, IS') 

    
            
